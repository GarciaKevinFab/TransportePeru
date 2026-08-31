from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Query, Body, Request, Header
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient

# Postgres: la unica base del backend desde el corte 013. Las 50 tablas
# cruzaron y ya no queda ningun acceso a Mongo; ver db/tablas_en_postgres.txt.
import db_pg
import tenant_host

import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from enum import Enum
import re
import secrets
import asyncio

# Rate limiting (slowapi). Add "slowapi" to backend/requirements.txt.
try:
    from slowapi import Limiter, _rate_limit_exceeded_handler
    from slowapi.util import get_remote_address
    from slowapi.errors import RateLimitExceeded
    limiter = Limiter(key_func=get_remote_address)
    SLOWAPI_AVAILABLE = True
except ImportError:
    # slowapi no instalado: limiter no-op para que el server siga arrancando
    SLOWAPI_AVAILABLE = False

    class _NoopLimiter:
        def limit(self, *args, **kwargs):
            def decorator(func):
                return func
            return decorator

    limiter = _NoopLimiter()

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# ============== TABLAS EN POSTGRES: EMPRESAS Y USUARIOS ==============
# companies y users ya cortaron (db/migrations/004_corte_companies_users.sql).
# Son la raiz del modelo: todo lo demas apunta a ellas.

COMPANY_COLS = {
    "id": "uuid", "name": "text", "ruc": "text", "address": "text",
    # Subdominio de la empresa (migracion 016). Tiene que estar en el mapa: sin
    # la clave, build_insert lo omitiria del INSERT y chocaria contra el NOT
    # NULL de la columna en cuanto se cree la primera empresa.
    "slug": "text",
    "phone": "text", "email": "text", "logo_url": "text", "brand_color": "text",
    # Suscripcion (migracion 015). Sin estas tres, el alta por la web dejaria
    # la empresa con los defaults del esquema y no habria prueba que vencer.
    "plan": "text", "subscription_status": "text", "trial_ends_at": "ts",
    "config": "json", "sunat_config": "json",
    "created_at": "ts", "updated_at": "ts",
}

# Dias de prueba de un alta nueva. La landing lo anuncia, asi que si cambia
# aqui tiene que cambiar alli: es la misma promesa.
DIAS_DE_PRUEBA = 14

# Vehiculos que admite cada plan. La landing publica estas cifras, asi que
# tienen que cumplirse: un limite anunciado y no aplicado es una cifra
# decorativa, y el cliente lo descubre el dia que quiere pagar mas.
#
# Lo que NO esta en el mapa no tiene limite, y es intencionado: 'trial' -para
# que se pueda probar con la flota entera, que es de lo que depende la venta- y
# 'activa', que es como quedaron las empresas que ya operaban antes de que
# existieran los planes. Ante un plan desconocido se deja pasar, por lo mismo
# que el corte por suscripcion falla abierto.
LIMITE_VEHICULOS = {"gratis": 3, "pro": 20}


async def _exigir_cupo_de_vehiculos(current_user: dict):
    """403 cuando la empresa ya llego al tope de su plan."""
    async with db_pg.tx(current_user) as conn:
        fila = await conn.fetchrow(
            "select c.plan, count(v.id) as usados "
            "  from companies c left join vehicles v on v.company_id = c.id "
            " where c.id = $1 group by c.plan",
            db_pg.as_uuid(current_user["company_id"]),
        )
    if not fila:
        return
    tope = LIMITE_VEHICULOS.get(fila["plan"])
    if tope is not None and fila["usados"] >= tope:
        raise HTTPException(
            status_code=403,
            detail=("El plan %s admite hasta %d vehiculos y ya tienes %d. "
                    "Cambia de plan para agregar mas."
                    % (fila["plan"], tope, fila["usados"])),
        )

USER_COLS = {
    "id": "uuid", "company_id": "uuid", "email": "text", "dni": "text",
    "name": "text", "role": "enum:user_role",
    "password_hash": "text", "pin_hash": "text", "is_active": "bool",
    "failed_attempts": "int", "locked_until": "ts",
    "force_password_change": "bool",
    "license_number": "text", "license_expiry": "ts",
    "phone": "text", "whatsapp_number": "text",
    "epp": "json", "push_subscription": "json",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}

# JWT Settings
JWT_SECRET = os.environ.get("JWT_SECRET")
if not JWT_SECRET:
    if os.environ.get("ENV", "development").lower() == "production":
        raise RuntimeError("JWT_SECRET no configurado en producción")
    JWT_SECRET = "dev-only-insecure-secret"  # solo desarrollo
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.environ.get("ACCESS_TOKEN_EXPIRE_MINUTES", "15"))
REFRESH_TOKEN_EXPIRE_DAYS = int(os.environ.get("REFRESH_TOKEN_EXPIRE_DAYS", "7"))

# Upload directory
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Create the main app
app = FastAPI(title="TransportePeru SaaS API", version="1.0.0")

# Register rate limiter (no-op si slowapi no está instalado)
if SLOWAPI_AVAILABLE:
    app.state.limiter = limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Create router with /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# ============== ENUMS ==============
class UserRole(str, Enum):
    SUPERADMIN = "superadmin"
    OWNER = "owner"
    ADMIN = "admin"
    OPERACIONES = "operaciones"
    FLOTA = "flota"
    MANTENIMIENTO = "mantenimiento"
    ALMACEN = "almacen"
    CONTABILIDAD = "contabilidad"
    CHOFER = "chofer"

class VehicleType(str, Enum):
    TRACTO = "tracto"
    CARRETA = "carreta"

class VehicleStatus(str, Enum):
    DISPONIBLE = "disponible"
    EN_VIAJE = "en_viaje"
    EN_MANTENIMIENTO = "en_mantenimiento"
    FUERA_SERVICIO = "fuera_servicio"

class DocumentStatus(str, Enum):
    VIGENTE = "vigente"
    POR_VENCER = "por_vencer"
    VENCIDO = "vencido"
    PENDIENTE = "pendiente"
    APROBADO = "aprobado"
    OBSERVADO = "observado"
    RECHAZADO = "rechazado"

class TripStatus(str, Enum):
    PROGRAMADO = "programado"
    EN_CURSO = "en_curso"
    COMPLETADO = "completado"
    CANCELADO = "cancelado"

class TireStatus(str, Enum):
    NUEVO = "nuevo"
    EN_USO = "en_uso"
    REENCAUCHE = "reencauche"
    BAJA = "baja"
    ALMACEN = "almacen"

class ChecklistResult(str, Enum):
    PENDING = "pending"
    OK = "ok"
    OBSERVADO = "observado"
    CRITICO = "critico"

class WorkOrderStatus(str, Enum):
    ABIERTA = "abierta"
    EN_PROCESO = "en_proceso"
    COMPLETADA = "completada"
    CANCELADA = "cancelada"

class WorkOrderPriority(str, Enum):
    BAJA = "baja"
    NORMAL = "normal"
    ALTA = "alta"
    CRITICA = "critica"

class IssueType(str, Enum):
    INCIDENTE = "incidente"
    MULTA = "multa"
    SINIESTRO = "siniestro"
    CHECKLIST_CRITICO = "checklist_critico"
    LLANTA_CRITICA = "llanta_critica"
    OTRO = "otro"

class IssueSeverity(str, Enum):
    BAJA = "baja"
    MEDIA = "media"
    ALTA = "alta"
    CRITICA = "critica"

class ExpenseCategory(str, Enum):
    ALIMENTACION = "alimentacion"
    HOSPEDAJE = "hospedaje"
    MOVILIDAD = "movilidad"
    PEAJES = "peajes"
    PARQUEO = "parqueo"
    COMBUSTIBLE = "combustible"
    # La app del chofer ya ofrecia "Ticket Balanza" y este enum no lo tenia:
    # las tres capas (aca, el enum de Postgres y el frontend) discrepaban.
    BALANZA = "balanza"
    OTROS = "otros"

class SettlementStatus(str, Enum):
    PENDIENTE = "pendiente"
    EN_REVISION = "en_revision"
    APROBADO = "aprobado"
    CERRADO = "cerrado"

class StockMoveType(str, Enum):
    ENTRADA = "entrada"
    SALIDA = "salida"
    AJUSTE = "ajuste"
    CONSUMO_OT = "consumo_ot"

class BlockRule(str, Enum):
    BLOQUEA_ASIGNACION = "bloquea_asignacion"
    BLOQUEA_INICIO = "bloquea_inicio"
    SOLO_ALERTA = "solo_alerta"

# ============== MODELS ==============
class Company(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    # Opcional en el modelo aunque la columna sea NOT NULL: se decide con
    # tenant_host.slug_libre(), que consulta la base, y por lo tanto no puede
    # salir de un default de Pydantic. Quien construye un Company lo asigna
    # antes de insertar.
    slug: Optional[str] = None
    ruc: str
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    logo_url: Optional[str] = None
    brand_color: str = "#f97316"
    config: Dict[str, Any] = Field(default_factory=dict)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    email: Optional[str] = None
    dni: Optional[str] = None
    name: str
    role: UserRole
    password_hash: Optional[str] = None
    pin_hash: Optional[str] = None
    is_active: bool = True
    failed_attempts: int = 0
    locked_until: Optional[datetime] = None
    force_password_change: bool = False
    license_number: Optional[str] = None
    license_expiry: Optional[datetime] = None
    phone: Optional[str] = None
    whatsapp_number: Optional[str] = None  # E.164 normalizado (+51...), usado por el bot de WhatsApp
    # EPP (Equipo de Protección Personal) - solo para choferes
    # Estructura: {"casco": {"assigned": true, "date": "2024-01-15", "condition": "bueno", "size": "M"}, ...}
    epp: Dict[str, Any] = Field(default_factory=dict)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Vehicle(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    plate: str
    vehicle_type: VehicleType
    brand: Optional[str] = None
    model: Optional[str] = None
    year: Optional[int] = None
    vin: Optional[str] = None
    color: Optional[str] = None
    status: VehicleStatus = VehicleStatus.DISPONIBLE
    odometer: int = 0
    fuel_capacity: Optional[float] = None
    tire_config: str = "6"  # Number of tires
    # Configuración de ejes: [{"name": str, "type": "direccional"|"traccion"|"muerto"|"levantable", "dual": bool, "is_spare": bool}]
    axle_config: Optional[List[Dict[str, Any]]] = None
    axle_config_history: Optional[List[Dict[str, Any]]] = None
    assigned_driver_id: Optional[str] = None
    photo_url: Optional[str] = None
    proveedor_id: Optional[str] = None    # Proveedor/transportista dueño de esta unidad
                                            # (ver liquidacion_flete.py) — None hasta que se asigne
    viatico_fijo: Optional[float] = None   # Viático fijo por viaje para esta placa
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class VehicleEquipment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    vehicle_id: str
    items: List[Dict[str, Any]] = []  # [{name, quantity, condition, expiry_date}]
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_by: Optional[str] = None

class CouplingHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    tracto_id: str
    carreta_id: str
    trip_id: Optional[str] = None
    start_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    end_date: Optional[datetime] = None
    created_by: Optional[str] = None

class DocumentType(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    name: str
    applies_to: str  # vehiculo, chofer, empresa
    is_critical: bool = False
    requires_expiry: bool = True
    alert_days: List[int] = Field(default_factory=lambda: [60, 30, 15, 7, 3, 1])
    block_rule: BlockRule = BlockRule.SOLO_ALERTA
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Document(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    document_type_id: str
    entity_type: str  # vehicle, user, company
    entity_id: str
    number: Optional[str] = None
    issue_date: Optional[datetime] = None
    expiry_date: Optional[datetime] = None
    status: DocumentStatus = DocumentStatus.PENDIENTE
    file_url: Optional[str] = None
    notes: Optional[str] = None
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Alert(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    alert_type: str
    entity_type: str
    entity_id: str
    message: str
    severity: str = "warning"  # info, warning, critical
    is_read: bool = False
    resolved: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class OperationalBlock(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    entity_type: str
    entity_id: str
    reason: str
    block_type: str
    document_id: Optional[str] = None
    document_type_id: Optional[str] = None
    is_active: bool = True
    resolved_at: Optional[datetime] = None
    resolved_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Route(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    name: str
    origin: str
    destination: str
    distance_km: float
    estimated_hours: float
    toll_cost: float = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Trip(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    trip_number: Optional[str] = None
    tracto_id: str
    carreta_id: Optional[str] = None
    driver_id: str
    route_id: Optional[str] = None
    client_name: Optional[str] = None
    cargo_description: Optional[str] = None
    cargo_weight: Optional[float] = None
    status: TripStatus = TripStatus.PROGRAMADO
    is_round_trip: bool = True
    scheduled_date: datetime
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    km_start: Optional[int] = None
    km_end: Optional[int] = None
    checklist_id: Optional[str] = None
    checklist_approved: bool = False
    total_advance: float = 0
    total_expenses: float = 0
    settlement_status: str = "pending"
    viatico_budget: Optional[float] = None
    viatico_days: Optional[int] = None
    viatico_daily: Optional[float] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class TripAdvance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    trip_id: str
    amount: float
    payment_method: str
    delivered_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    delivered_by: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TripExpense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    trip_id: str
    category: str
    description: Optional[str] = None
    amount: float
    provider: Optional[str] = None
    ruc: Optional[str] = None
    has_igv: bool = False
    receipt_url: Optional[str] = None
    expense_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Checklist(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    trip_id: str
    vehicle_id: str
    driver_id: str
    items: List[Dict[str, Any]] = Field(default_factory=list)
    tire_checks: List[Dict[str, Any]] = Field(default_factory=list)
    result: str = "pending"  # ok, observado, critico
    signature_url: Optional[str] = None
    location: Optional[Dict[str, float]] = None
    completed_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FuelVoucher(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    vehicle_id: str
    trip_id: Optional[str] = None
    voucher_number: str
    provider: str
    limit_amount: Optional[float] = None
    limit_liters: Optional[float] = None
    valid_from: datetime
    valid_until: datetime
    is_used: bool = False
    approved_by: Optional[str] = None
    voucher_photo_url: Optional[str] = None
    invoice_photo_url: Optional[str] = None
    invoice_number: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FuelLoad(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    vehicle_id: str
    voucher_id: Optional[str] = None
    trip_id: Optional[str] = None
    voucher_number: Optional[str] = None
    invoice_number: Optional[str] = None
    liters: float
    price_per_liter: float
    total_amount: float
    odometer: int
    provider: str
    receipt_url: Optional[str] = None
    voucher_photo_url: Optional[str] = None
    invoice_photo_url: Optional[str] = None
    location: Optional[Dict[str, float]] = None
    load_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Tire(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    serial: str
    brand: str
    model: Optional[str] = None
    dimension: str
    position_type: Optional[str] = "toda_posicion"  # direccional|traccion|toda_posicion|mixto
    purchase_cost: float = 0
    purchase_date: Optional[datetime] = None
    supplier: Optional[str] = None
    status: TireStatus = TireStatus.NUEVO
    life_number: int = 1  # VN=1, R1=2, R2=3...
    initial_depth: Optional[float] = None  # baseline mm at current life (for cost_per_mm / desgaste)
    current_vehicle_id: Optional[str] = None
    current_position: Optional[str] = None
    total_km: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TireMount(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    tire_id: str
    vehicle_id: str
    position_code: str
    mount_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    mount_odometer: int
    unmount_date: Optional[datetime] = None
    unmount_odometer: Optional[int] = None
    reason: Optional[str] = None
    created_by: Optional[str] = None

class TireInspection(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    tire_id: str
    vehicle_id: str
    position_code: str
    depths: List[float] = Field(default_factory=list)  # 2-4 measurements in mm
    pressure: float
    irregular_wear: bool = False
    wear_type: Optional[str] = None
    photos: List[str] = Field(default_factory=list)
    odometer: int
    notes: Optional[str] = None
    inspection_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class MaintenancePlan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    name: str
    vehicle_type: VehicleType
    component: str
    interval_km: Optional[int] = None
    interval_days: Optional[int] = None
    interval_hours: Optional[int] = None
    tasks: List[str] = Field(default_factory=list)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaintenanceMatrixPlan(BaseModel):
    """Matrix-based maintenance plan (intervals x tasks like E MAX 540 plan)"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    name: str  # "E MAX 540 MT"
    vehicle_model: Optional[str] = None
    applies_to_vehicle_ids: List[str] = Field(default_factory=list)
    # Intervals: [{code:"M1", hours:500, km:30, labor_hours:4}, ...]
    intervals: List[Dict[str, Any]] = Field(default_factory=list)
    # Sections: [{code:"A", name:"MOTOR", tasks:[{n:1, description:"...", component_type:"FILTRO", quantity:1, actions:{"M1":"C","M2":"C"}}]}]
    sections: List[Dict[str, Any]] = Field(default_factory=list)
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Issue(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    issue_number: Optional[str] = None
    trip_id: Optional[str] = None
    vehicle_id: Optional[str] = None
    driver_id: Optional[str] = None
    checklist_id: Optional[str] = None
    tire_id: Optional[str] = None
    issue_type: str  # incidente, multa, siniestro, checklist_critico, llanta_critica
    severity: str = "media"  # baja, media, alta, critica
    status: str = "abierto"  # abierto, en_proceso, cerrado
    title: str = ""
    description: str
    location: Optional[Dict[str, float]] = None
    photos: List[str] = Field(default_factory=list)
    cost: float = 0
    responsible: Optional[str] = None
    resolution: Optional[str] = None
    work_order_id: Optional[str] = None
    resolved_by: Optional[str] = None
    resolved_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

# ============== CHECKLIST MODELS ==============
class ChecklistTemplate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    name: str
    vehicle_type: Optional[str] = None
    items: List[Dict[str, Any]] = Field(default_factory=list)
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class ChecklistRun(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    template_id: str
    trip_id: str
    tracto_id: str
    carreta_id: Optional[str] = None
    driver_id: str
    responses: List[Dict[str, Any]] = Field(default_factory=list)
    tire_checks: List[Dict[str, Any]] = Field(default_factory=list)
    result: str = "pending"  # pending, ok, observado, critico
    signature_url: Optional[str] = None
    location: Optional[Dict[str, float]] = None
    started_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completed_at: Optional[datetime] = None
    photos: List[str] = Field(default_factory=list)
    created_by: Optional[str] = None

# ============== SETTLEMENT MODEL ==============
class TripSettlement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    trip_id: str
    total_advances: float = 0
    total_expenses: float = 0
    deductions: float = 0
    deduction_notes: Optional[str] = None
    balance: float = 0
    balance_type: str = "favor_empresa"
    status: str = "pendiente"
    reviewed_by: Optional[str] = None
    reviewed_at: Optional[datetime] = None
    closed_by: Optional[str] = None
    closed_at: Optional[datetime] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ============== MAINTENANCE MODELS ==============
class WorkOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    order_number: str
    vehicle_id: str
    order_type: str  # preventivo, correctivo
    priority: str = "normal"  # baja, normal, alta, critica
    status: str = "abierta"  # abierta, en_proceso, completada, cancelada
    description: str
    maintenance_plan_id: Optional[str] = None
    issue_id: Optional[str] = None
    items: List[Dict[str, Any]] = Field(default_factory=list)
    labor_cost: float = 0
    parts_cost: float = 0
    total_cost: float = 0
    workshop: Optional[str] = None
    technician: Optional[str] = None
    scheduled_date: Optional[datetime] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    odometer_at_service: Optional[int] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None
    closed_by: Optional[str] = None

class DowntimeRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    vehicle_id: str
    work_order_id: Optional[str] = None
    reason: str
    start_time: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    end_time: Optional[datetime] = None
    duration_hours: float = 0
    created_by: Optional[str] = None

# ============== INVENTORY MODELS ==============
class InventoryItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    code: str
    name: str
    description: Optional[str] = None
    category: str
    unit: str = "unidad"
    min_stock: int = 0
    max_stock: Optional[int] = None
    current_stock: int = 0
    unit_cost: float = 0
    location: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StockMove(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    item_id: str
    move_type: str  # entrada, salida, ajuste, consumo_ot
    quantity: int
    unit_cost: float = 0
    total_cost: float = 0
    reference_type: Optional[str] = None
    reference_id: Optional[str] = None
    work_order_id: Optional[str] = None
    notes: Optional[str] = None
    move_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Supplier(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    name: str
    ruc: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    contact_person: Optional[str] = None
    category: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PurchaseOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    order_number: str
    supplier_id: str
    status: str = "borrador"
    items: List[Dict[str, Any]] = Field(default_factory=list)
    subtotal: float = 0
    tax: float = 0
    total: float = 0
    notes: Optional[str] = None
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    received_by: Optional[str] = None
    received_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

# ============== AUDIT LOG ==============
class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    user_id: str
    user_name: str
    action: str
    entity_type: str
    entity_id: str
    details: Dict[str, Any] = Field(default_factory=dict)
    ip_address: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ============== TIRE EXTENDED MODELS ==============
class TireLifeEvent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    tire_id: str
    life_number: int
    event_type: str
    cost: float = 0
    supplier: Optional[str] = None
    notes: Optional[str] = None
    event_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class TireRotation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    vehicle_id: str
    changes: List[Dict[str, str]] = Field(default_factory=list)
    reason: Optional[str] = None
    odometer: int = 0
    rotation_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class AlignmentRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    vehicle_id: str
    axle: str
    workshop: Optional[str] = None
    cost: float = 0
    notes: Optional[str] = None
    alignment_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

# ============== REQUEST/RESPONSE MODELS ==============
class LoginRequest(BaseModel):
    email: Optional[str] = None
    password: Optional[str] = None
    dni: Optional[str] = None
    pin: Optional[str] = None

class TokenResponse(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str = "bearer"
    user: Dict[str, Any]

class RefreshRequest(BaseModel):
    refresh_token: str

class CreateUserRequest(BaseModel):
    email: Optional[str] = None
    dni: Optional[str] = None
    name: str
    role: UserRole
    password: Optional[str] = None
    pin: Optional[str] = None
    license_number: Optional[str] = None
    license_expiry: Optional[datetime] = None
    phone: Optional[str] = None

class CreateVehicleRequest(BaseModel):
    plate: str
    vehicle_type: VehicleType
    brand: Optional[str] = None
    model: Optional[str] = None
    year: Optional[int] = None
    vin: Optional[str] = None
    color: Optional[str] = None
    fuel_capacity: Optional[float] = None
    tire_config: str = "6"
    axle_config: Optional[List[Dict[str, Any]]] = None

class CreateTripRequest(BaseModel):
    tracto_id: str
    carreta_id: Optional[str] = None
    driver_id: str
    route_id: Optional[str] = None
    client_name: Optional[str] = None
    cargo_description: Optional[str] = None
    cargo_weight: Optional[float] = None
    scheduled_date: datetime
    is_round_trip: bool = True
    notes: Optional[str] = None

class CreateDocumentRequest(BaseModel):
    document_type_id: str
    entity_type: str
    entity_id: str
    number: Optional[str] = None
    issue_date: Optional[datetime] = None
    expiry_date: Optional[datetime] = None
    notes: Optional[str] = None

class CreateTireRequest(BaseModel):
    serial: str
    brand: str
    model: Optional[str] = None
    dimension: str
    position_type: Optional[str] = "toda_posicion"  # direccional|traccion|toda_posicion|mixto
    purchase_cost: float = 0
    purchase_date: Optional[datetime] = None
    supplier: Optional[str] = None
    initial_depth: Optional[float] = None

class MountTireRequest(BaseModel):
    tire_id: str
    vehicle_id: str
    position_code: str
    mount_odometer: int

class CreateInspectionRequest(BaseModel):
    tire_id: str
    vehicle_id: str
    position_code: str
    depths: List[float]
    pressure: float
    irregular_wear: bool = False
    wear_type: Optional[str] = None
    odometer: int
    notes: Optional[str] = None

# ============== HELPER FUNCTIONS ==============
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire, "type": "access"})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire, "type": "refresh"})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

async def get_current_user(
    request: Request,
    credentials: HTTPAuthorizationCredentials = Depends(security),
) -> dict:
    token = credentials.credentials
    payload = decode_token(token)
    if payload.get("type") != "access":
        raise HTTPException(status_code=401, detail="Token inválido")
    
    # tx_global y no tx(): resolver la identidad es justamente lo que pasa antes
    # de saber la empresa. Ademas, el company_id del token NO sirve para filtrar:
    # un superadmin que uso /companies/{id}/switch lleva en el token la empresa a
    # la que entro, no la suya, y filtrar por ese valor no encontraria su usuario.
    # Se devuelve la fila tal cual, igual que hacia la version con Mongo.
    async with db_pg.tx_global("autenticacion: resolver la identidad antes de conocer la empresa") as conn:
        fila = await conn.fetchrow(
            "select * from users where id = $1", db_pg.as_uuid(payload["user_id"])
        )
    if not fila:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    user = db_pg.to_api(fila)
    if not user.get("is_active"):
        raise HTTPException(status_code=401, detail="Usuario desactivado")
    await _aplicar_empresa_del_token(payload, user)
    await _exigir_host_de_la_empresa(request, user)
    await _exigir_suscripcion_al_dia(request, user)
    return user


async def _aplicar_empresa_del_token(payload: dict, user: dict):
    """Hace efectivo el /companies/{id}/switch de un superadmin.

    Hasta ahora ese endpoint emitia un token con la empresa destino y NADIE
    leia ese campo: get_current_user devolvia la fila del usuario y todos los
    endpoints usan current_user["company_id"]. O sea que el superadmin
    "cambiaba de empresa" y seguia viendo la suya, sin ningun error - la peor
    forma de fallar, porque parece que funciona.

    Aqui es donde el campo empieza a valer. A partir de esta linea,
    current_user["company_id"] es la empresa en la que se esta trabajando, y
    todo lo demas -incluido el SET LOCAL de db_pg.tx() que fija el contexto de
    RLS- va detras de ella sin enterarse de que hubo un cambio.

    TRES CANDADOS, y ninguno sobra:

      1. El rol se lee de la FILA, no del token. Un token viejo emitido cuando
         el usuario era superadmin deja de servir en cuanto se le baja el rol
         en la base; si mirasemos el `role` del token, seguiria entrando en
         cualquier empresa hasta que caducara.
      2. La empresa destino tiene que existir. Si se borro mientras habia una
         sesion dentro, la sesion muere en vez de quedar apuntando al vacio.
      3. RLS NO se abre. El contexto se fija en la empresa destino y las
         politicas siguen aplicando con la misma fuerza: el superadmin ve esa
         empresa y solo esa, igual que cualquier usuario suyo. Entrar a una
         empresa no es apagar el aislamiento, es mudarse a otro lado de el.

    Para todos los demas roles, el campo del token se ignora por completo.
    """
    empresa_del_token = payload.get("company_id")
    if not empresa_del_token or user.get("role") != "superadmin":
        return
    if str(empresa_del_token) == str(user.get("company_id")):
        return  # su propia empresa: no hay nada que cambiar

    try:
        destino = db_pg.as_uuid(empresa_del_token)
    except (ValueError, AttributeError, TypeError):
        raise HTTPException(status_code=401, detail="Token invalido")

    async with db_pg.tx_global("superadmin: comprobar la empresa en la que esta trabajando") as conn:
        existe = await conn.fetchval("select 1 from companies where id = $1", destino)
    if not existe:
        raise HTTPException(
            status_code=401,
            detail="La empresa de esta sesion ya no existe. Vuelve a entrar.",
        )

    # Se conserva de donde viene para poder volver, y para que la interfaz
    # pueda avisar de que se esta dentro de otra empresa. Un superadmin que se
    # olvida de que esta en los datos de un cliente es un incidente esperando.
    user["company_id_propio"] = user["company_id"]
    user["en_otra_empresa"] = True
    user["company_id"] = str(empresa_del_token)


async def _exigir_host_de_la_empresa(request: Request, user: dict):
    """403 cuando la direccion es de una empresa y la sesion es de otra.

    Solo actua si el host resuelve a un inquilino. En la landing, en el acceso
    de rescate (fletepro.sisac.pe) y en local no hay empresa en el host y no
    hay nada que comparar: se sigue como siempre.

    ESTO NO ES LO QUE IMPIDE VER DATOS AJENOS. De eso se encargan el company_id
    del usuario y las politicas RLS, y siguen intactos: un token de la empresa
    A servido desde el origen de B nunca vio datos de B. Lo que evita es que la
    sesion de un inquilino siga viva en el origen de otro, que confunde a
    cualquiera que lo lea en un log y no tiene ninguna razon de ser.

    Sin excepcion para superadmin, a proposito. Su empresa es la del sistema,
    asi que un subdominio de cliente le responde 403 y su sitio es la consola
    en fletepro.sisac.pe. Una excepcion aqui significaria que el superadmin ve
    en gye.sisac.pe los datos de la empresa del sistema, que es peor que el
    403: parece que la direccion funciona cuando no lo hace.
    """
    empresa = await tenant_host.empresa_desde_host(request.headers.get("host"))
    if empresa is None:
        return
    if str(user.get("company_id")) == str(empresa["id"]):
        return
    raise HTTPException(
        status_code=403,
        detail="Esta sesion es de otra empresa. Vuelve a entrar desde la "
               "direccion de la tuya.",
        # El interceptor del frontend lo usa para cerrar la sesion y mandar al
        # login en vez de mostrar un error que el usuario no puede resolver.
        # Se comprueba la cabecera y no el texto para no atar el frontend a la
        # redaccion de un mensaje.
        headers={"X-Tenant-Mismatch": "1"},
    )


# Metodos que solo leen. Una suscripcion vencida deja mirar pero no escribir:
# los datos son del cliente, y cortarle la lectura de su propia operacion seria
# retenerselos, no cobrarle.
_METODOS_DE_LECTURA = {"GET", "HEAD", "OPTIONS"}


async def _exigir_suscripcion_al_dia(request: Request, user: dict):
    """402 cuando la empresa dejo de estar al dia Y la peticion escribe.

    FALLA ABIERTO a proposito. Solo corta ante un vencimiento comprobado; ante
    cualquier otra cosa -empresa que no aparece, columna sin valor- deja pasar.
    Un fallo aca bloquearia a clientes que si pagaron, y eso es mucho peor que
    regalar un dia de mas.
    """
    if user.get("role") == "superadmin":
        return  # gestiona las empresas: no puede quedar fuera por una de ellas
    if request.method in _METODOS_DE_LECTURA:
        return
    company_id = user.get("company_id")
    if not company_id:
        return

    async with db_pg.tx_global("comprobar la suscripcion de la empresa") as conn:
        fila = await conn.fetchrow(
            "select subscription_status, trial_ends_at from companies where id = $1",
            db_pg.as_uuid(company_id),
        )
    if not fila:
        return

    estado = fila["subscription_status"]
    fin = fila["trial_ends_at"]
    if estado in ("vencida", "cancelada"):
        vencida = True
    elif estado == "trial" and fin is not None:
        vencida = fin < datetime.now(timezone.utc)
    else:
        vencida = False

    if vencida:
        raise HTTPException(
            status_code=402,
            detail="La prueba gratuita termino. Activa un plan para seguir "
                   "registrando informacion; tus datos siguen disponibles "
                   "para consulta.",
        )

def _slug_pedido(valor: str) -> str:
    """Valida un slug elegido a mano y traduce el motivo a un 400 legible.

    tenant_host levanta SlugInvalido (un ValueError) porque no sabe nada de
    HTTP; la traduccion vive aca, que es donde empieza la web.
    """
    try:
        return tenant_host.validar_slug(valor)
    except tenant_host.SlugInvalido as e:
        raise HTTPException(status_code=400, detail=str(e))


def require_roles(*roles):
    """Dependency reutilizable para exigir uno de los roles dados.
    superadmin siempre está autorizado."""
    async def checker(current_user: dict = Depends(get_current_user)):
        if current_user["role"] not in roles and current_user["role"] != "superadmin":
            raise HTTPException(status_code=403, detail="No autorizado")
        return current_user
    return checker

def serialize_doc(doc: dict) -> dict:
    """Remove MongoDB _id and convert datetimes to ISO strings"""
    if doc is None:
        return None
    result = {k: v for k, v in doc.items() if k != "_id"}
    for key, value in result.items():
        if isinstance(value, datetime):
            result[key] = value.isoformat()
    return result

def _normalize_text(s: Optional[str]) -> str:
    """Lowercase + strip accents for tolerant name matching."""
    if not s:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()

def _is_revision_tecnica(name: Optional[str]) -> bool:
    """True if the document type name refers to Revisión Técnica (CITV)."""
    n = _normalize_text(name)
    if not n:
        return False
    return "citv" in n or ("revision" in n and "tecnica" in n)

def _revision_tecnica_no_aplica(doc_type: dict, entity: dict, entity_type: str) -> bool:
    """Regla: Revisión Técnica no aplica a vehículos con <= 4 años de antigüedad."""
    if entity_type != "vehicle":
        return False
    if not _is_revision_tecnica((doc_type or {}).get("name")):
        return False
    year = (entity or {}).get("year")
    try:
        year = int(year) if year is not None else None
    except (TypeError, ValueError):
        year = None
    if not year:
        return False
    current_year = datetime.now(timezone.utc).year
    return (current_year - year) <= 4

# ============== TABLAS EN POSTGRES: DOCUMENTOS Y BLOQUEOS ==============
# document_types, documents y blocks cortaron con la migracion 009. Van juntas
# porque documents apunta a document_types y blocks apunta a las dos.
#
# entity_id es polimorfico -segun entity_type apunta a vehicles o a users- y
# por eso va sin FK, igual que estaba.

DOCUMENT_TYPE_COLS = {
    "id": "uuid", "company_id": "uuid", "name": "text", "applies_to": "text",
    "is_critical": "bool", "requires_expiry": "bool",
    "alert_days": "int[]", "block_rule": "enum:block_rule",
    "created_at": "ts",
}

DOCUMENT_COLS = {
    "id": "uuid", "company_id": "uuid", "document_type_id": "uuid",
    "entity_type": "text", "entity_id": "uuid", "number": "text",
    "issue_date": "ts", "expiry_date": "ts",
    "status": "enum:document_status", "file_url": "text", "notes": "text",
    "approved_by": "uuid", "approved_at": "ts",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}

BLOCK_COLS = {
    "id": "uuid", "company_id": "uuid", "entity_type": "text",
    "entity_id": "uuid", "reason": "text", "block_type": "text",
    "document_id": "uuid", "document_type_id": "uuid",
    "is_active": "bool", "resolved_at": "ts", "resolved_by": "uuid",
    "created_at": "ts",
}


async def check_entity_blocks(company_id: str, entity_type: str, entity_id: str, block_type: str = None) -> List[dict]:
    """Check for active blocks on an entity"""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
    f.agregar("entity_type = $?", entity_type)
    f.agregar("entity_id = $?", db_pg.as_uuid(entity_id))
    f.crudo("is_active")
    f.si(block_type, "block_type = $?", block_type)
    async with db_pg.tx({"company_id": company_id}) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from blocks where " + f.where + " limit 100", *f.values
        ))

async def validate_trip_can_be_assigned(company_id: str, tracto_id: str, carreta_id: str, driver_id: str) -> dict:
    """Validate that a trip can be assigned (no blocking conditions)"""
    errors = []
    
    # Check tracto blocks
    tracto_blocks = await check_entity_blocks(company_id, "vehicle", tracto_id, "bloquea_asignacion")
    if tracto_blocks:
        errors.append(f"Tracto bloqueado: {tracto_blocks[0].get('reason', 'Sin razón')}")
    
    # Check carreta blocks
    if carreta_id:
        carreta_blocks = await check_entity_blocks(company_id, "vehicle", carreta_id, "bloquea_asignacion")
        if carreta_blocks:
            errors.append(f"Carreta bloqueada: {carreta_blocks[0].get('reason', 'Sin razón')}")
    
    # Check driver blocks
    driver_blocks = await check_entity_blocks(company_id, "user", driver_id, "bloquea_asignacion")
    if driver_blocks:
        errors.append(f"Chofer bloqueado: {driver_blocks[0].get('reason', 'Sin razón')}")
    
    # Check critical work orders
    unidades = [
        db_pg.as_uuid(v)
        for v in ([tracto_id, carreta_id] if carreta_id else [tracto_id])
        if v
    ]
    async with db_pg.tx({"company_id": company_id}) as conn:
        critical_wos = db_pg.rows_to_api(await conn.fetch(
            "select * from work_orders where company_id = $1 "
            "and vehicle_id = any($2::uuid[]) "
            "and status in ('abierta', 'en_proceso') and priority = 'critica' "
            "limit 10",
            db_pg.as_uuid(company_id), unidades,
        ))
    if critical_wos:
        errors.append(f"OT crítica pendiente: {critical_wos[0].get('description', '')[:50]}")
    
    return {"valid": len(errors) == 0, "errors": errors}

async def validate_trip_can_start(company_id: str, trip_id: str, trip: dict) -> dict:
    """Validate that a trip can be started"""
    errors = []
    
    # Check if checklist is required and approved
    company = await _empresa_pg(company_id)
    config = company.get("config", {}) if company else {}
    
    if config.get("require_checklist_for_start", True):
        if not trip.get("checklist_id"):
            errors.append("Se requiere completar el checklist pre-viaje")
        elif trip.get("checklist_result") == "critico":
            errors.append("Checklist con resultado CRÍTICO - no se puede iniciar")
    
    # Check blocks with rule "bloquea_inicio"
    tracto_blocks = await check_entity_blocks(company_id, "vehicle", trip["tracto_id"], "bloquea_inicio")
    if tracto_blocks:
        errors.append(f"Tracto bloqueado para inicio: {tracto_blocks[0].get('reason', '')}")
    
    if trip.get("carreta_id"):
        carreta_blocks = await check_entity_blocks(company_id, "vehicle", trip["carreta_id"], "bloquea_inicio")
        if carreta_blocks:
            errors.append(f"Carreta bloqueada para inicio: {carreta_blocks[0].get('reason', '')}")
    
    driver_blocks = await check_entity_blocks(company_id, "user", trip["driver_id"], "bloquea_inicio")
    if driver_blocks:
        errors.append(f"Chofer bloqueado para inicio: {driver_blocks[0].get('reason', '')}")
    
    return {"valid": len(errors) == 0, "errors": errors}

async def create_audit_log(company_id: str, user_id: str, user_name: str, action: str, 
                           entity_type: str, entity_id: str, details: dict = None):
    """Create an audit log entry"""
    log = AuditLog(
        company_id=company_id,
        user_id=user_id,
        user_name=user_name,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details or {}
    )
    sql, values = db_pg.build_insert(
        "audit_logs", AUDIT_LOG_COLS, _modelo_a_fila(log.model_dump())
    )
    async with db_pg.tx({"company_id": company_id}) as conn:
        await conn.execute(sql, *values)

async def create_block_for_expired_doc(company_id: str, doc_type: dict, document: dict, entity_type: str, entity_id: str):
    """Create operational block for expired document"""
    block = OperationalBlock(
        company_id=company_id,
        entity_type=entity_type,
        entity_id=entity_id,
        reason=f"Documento vencido: {doc_type.get('name', 'Desconocido')}",
        block_type=doc_type.get("block_rule", "solo_alerta"),
        document_id=document["id"],
        document_type_id=doc_type.get("id")
    )
    sql, values = db_pg.build_insert(
        "blocks", BLOCK_COLS, _modelo_a_fila(block.model_dump())
    )
    async with db_pg.tx({"company_id": company_id}) as conn:
        await conn.execute(sql, *values)
    return block

# ============== BUSINESS RULE HELPERS (viáticos / llantas / mantenimiento / push) ==============
# Defaults configurables vía company.config
DEFAULT_VIATICO_POR_VIAJE = 540
DEFAULT_MAINT_ANTICIPATION_KM = 500
DEFAULT_TIRE_REVIEW_KM = 5000
DEFAULT_TIRE_CRITICAL_DEPTH = 3
DEFAULT_TIRE_WARNING_DEPTH = 5
# Detracciones (SPOT): transporte de carga = 4% cuando el comprobante supera S/ 400
DEFAULT_DETRACCION_RATE = 4.0
DEFAULT_DETRACCION_MIN_AMOUNT = 400
DEFAULT_DETRACCION_CODIGO = "027"  # servicio de transporte de carga


def _sin_secretos(user: dict) -> dict:
    """Quita los hashes de credenciales antes de devolver un usuario por la API.

    Con Mongo esto se hacia con una proyeccion ({password_hash: 0}) en cada
    consulta. En SQL seria enumerar las 20 columnas restantes en cada select y
    olvidarse de una al agregar la proxima, asi que se filtra a la salida y en
    un solo lugar.
    """
    if not user:
        return user
    return {k: v for k, v in user.items() if k not in ("password_hash", "pin_hash")}


async def _empresa_pg(company_id: str):
    """Fila de companies de la empresa dada, o None.

    Con RLS activo, el contexto de la transaccion ya limita a esa empresa; el
    where por id esta igual para que la consulta se lea sola.
    """
    if not company_id:
        return None
    async with db_pg.tx({"company_id": company_id}) as conn:
        return db_pg.to_api(await conn.fetchrow(
            "select * from companies where id = $1", db_pg.as_uuid(company_id)
        ))


async def _viaje_pg(company_id: str, trip_id):
    """Fila de trips dentro de la empresa, o None."""
    if not trip_id:
        return None
    async with db_pg.tx({"company_id": company_id}) as conn:
        return db_pg.to_api(await conn.fetchrow(
            "select * from trips where id = $1 and company_id = $2",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(company_id),
        ))


async def _contar_viajes(company_id: str, status=None) -> int:
    """Cantidad de viajes de la empresa. `status` acepta uno o varios."""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
    if isinstance(status, (list, tuple)):
        f.agregar("status = any($?::trip_status[])", list(status))
    elif status:
        f.agregar("status = $?::trip_status", status)
    async with db_pg.tx({"company_id": company_id}) as conn:
        return await conn.fetchval(
            "select count(*) from trips where " + f.where, *f.values
        )


async def _actualizar_viaje(company_id: str, trip_id, cambios: dict) -> bool:
    """Aplica `cambios` a un viaje. False si no existe (para el 404)."""
    datos = dict(cambios)
    datos["updated_at"] = datetime.now(timezone.utc)
    datos["id"] = trip_id
    datos["company_id"] = company_id
    sql, values = db_pg.build_update("trips", TRIP_COLS, datos, ["id", "company_id"])
    if not sql:
        return False
    async with db_pg.tx({"company_id": company_id}) as conn:
        return await conn.fetchval(sql + " returning id", *values) is not None


async def _vehiculo_pg(company_id: str, vehicle_id):
    """Fila de vehicles dentro de la empresa, o None."""
    if not vehicle_id:
        return None
    async with db_pg.tx({"company_id": company_id}) as conn:
        return db_pg.to_api(await conn.fetchrow(
            "select * from vehicles where id = $1 and company_id = $2",
            db_pg.as_uuid(vehicle_id), db_pg.as_uuid(company_id),
        ))


async def _actualizar_vehiculo(company_id: str, vehicle_id, cambios: dict) -> bool:
    """Aplica `cambios` a un vehiculo de la empresa.

    Devuelve False si no existe, para que quien llama pueda responder 404
    igual que hacia con matched_count. VEHICLE_COLS actua de lista blanca:
    lo que no sea una columna declarada se ignora.
    """
    datos = dict(cambios)
    datos["updated_at"] = datetime.now(timezone.utc)
    datos["id"] = vehicle_id
    datos["company_id"] = company_id
    sql, values = db_pg.build_update("vehicles", VEHICLE_COLS, datos, ["id", "company_id"])
    if not sql:
        return False
    async with db_pg.tx({"company_id": company_id}) as conn:
        return await conn.fetchval(sql + " returning id", *values) is not None


async def _contar_vehiculos(company_id: str, status: str = None) -> int:
    """Cantidad de vehiculos de la empresa, opcionalmente por estado."""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
    f.si(status, "status = $?::vehicle_status", status)
    async with db_pg.tx({"company_id": company_id}) as conn:
        return await conn.fetchval(
            "select count(*) from vehicles where " + f.where, *f.values
        )


async def _contar_usuarios(company_id: str, role: str = None) -> int:
    """Cantidad de usuarios de la empresa, opcionalmente filtrando por rol."""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
    f.si(role, "role = $?::user_role", role)
    async with db_pg.tx({"company_id": company_id}) as conn:
        return await conn.fetchval(
            "select count(*) from users where " + f.where, *f.values
        )


async def _usuario_pg(company_id: str, user_id):
    """Fila de users dentro de la empresa, o None si no existe o es de otra."""
    if not user_id:
        return None
    async with db_pg.tx({"company_id": company_id}) as conn:
        return db_pg.to_api(await conn.fetchrow(
            "select * from users where id = $1 and company_id = $2",
            db_pg.as_uuid(user_id), db_pg.as_uuid(company_id),
        ))


async def _company_config(company_id: str) -> dict:
    """Devuelve el dict de configuración de la empresa (o {})."""
    company = await _empresa_pg(company_id)
    if not company:
        return {}
    return company.get("config", {}) or {}


async def notify_users(company_id: str, title: str, message: str, notif_type: str = "info",
                       target_role: str = None, user_id: str = None,
                       entity_type: str = None, entity_id: str = None):
    """Crea un registro de notificación y dispara push (si hay suscripciones). No propaga errores."""
    try:
        notification = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "title": title,
            "message": message,
            "type": notif_type,
            "target_role": target_role,
            "user_id": user_id,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await _crear_notificacion(company_id, notification)
        try:
            await send_push_notifications(company_id, title, message, target_role, user_id)
        except Exception as e:
            logging.error(f"notify_users push error: {e}")
        return notification["id"]
    except Exception as e:
        logging.error(f"notify_users error: {e}")
        return None


# ============== TABLAS EN POSTGRES: ALERTAS, AVISOS Y BITACORA ==============
# alerts, notifications, audit_logs, vehicle_equipment y los planes matriciales
# cortaron con la migracion 012. No dependen unas de otras: van en un mismo
# corte precisamente porque ninguna esta acoplada al resto.

ALERT_COLS = {
    "id": "uuid", "company_id": "uuid", "alert_type": "text",
    "entity_type": "text", "entity_id": "uuid", "message": "text",
    "severity": "text", "is_read": "bool", "resolved": "bool",
    "created_at": "ts",
}

NOTIFICATION_COLS = {
    "id": "uuid", "company_id": "uuid", "title": "text", "message": "text",
    "type": "enum:notification_type", "target_role": "text", "user_id": "uuid",
    "entity_type": "text", "entity_id": "uuid", "is_read": "bool",
    "read_at": "ts", "created_by": "uuid", "created_at": "ts",
}

AUDIT_LOG_COLS = {
    "id": "uuid", "company_id": "uuid", "user_id": "uuid", "user_name": "text",
    "action": "text", "entity_type": "text", "entity_id": "uuid",
    "details": "json", "ip_address": "text", "created_at": "ts",
}

VEHICLE_EQUIPMENT_COLS = {
    "id": "uuid", "company_id": "uuid", "vehicle_id": "uuid",
    "items": "json", "updated_at": "ts", "updated_by": "uuid",
}

MATRIX_PLAN_COLS = {
    "id": "uuid", "company_id": "uuid", "name": "text",
    "vehicle_model": "text", "intervals": "json", "sections": "json",
    "notes": "text", "created_at": "ts", "updated_at": "ts",
    "created_by": "uuid",
}


async def _insertar_alerta(company_id: str, alerta) -> str:
    """Guarda una Alert ya construida. Devuelve su id.

    Varios sitios creaban la alerta y repetian las tres lineas de insercion;
    con una sola tabla destino no hay motivo para tenerlo escrito seis veces.
    """
    datos = alerta if isinstance(alerta, dict) else _modelo_a_fila(alerta.model_dump())
    sql, values = db_pg.build_insert("alerts", ALERT_COLS, datos)
    async with db_pg.tx({"company_id": company_id}) as conn:
        await conn.execute(sql, *values)
    return datos["id"]


async def _crear_notificacion(company_id: str, notificacion: dict) -> str:
    """Guarda una notificacion armada como dict (no hay modelo Notification)."""
    sql, values = db_pg.build_insert(
        "notifications", NOTIFICATION_COLS, notificacion
    )
    async with db_pg.tx({"company_id": company_id}) as conn:
        await conn.execute(sql, *values)
    return notificacion["id"]


async def _guardar_vehiculos_del_plan(conn, plan_id, company_id, vehicle_ids):
    """Reescribe la tabla puente de un plan matricial con la lista de vehiculos.

    applies_to_vehicle_ids es una lista en el modelo, pero en Postgres esa
    relacion vive normalizada en maintenance_matrix_plan_vehicles. Sin esto la
    lista blanca de columnas la descartaria en silencio y el plan quedaria sin
    sus vehiculos, sin ningun error.

    Se borra y se reinserta en vez de calcular el diff: son unas pocas filas y
    corre dentro de la misma transaccion que el plan.
    """
    pid = db_pg.as_uuid(plan_id)
    await conn.execute("delete from maintenance_matrix_plan_vehicles where plan_id = $1", pid)
    ids = [u for u in (db_pg.as_uuid(v) for v in (vehicle_ids or [])) if u]
    if not ids:
        return
    # El vehiculo tiene que ser de la misma empresa que el plan: el select
    # filtrado lo garantiza sin comprobarlo fila por fila.
    await conn.execute(
        "insert into maintenance_matrix_plan_vehicles (plan_id, vehicle_id) "
        "select $1, v.id from vehicles v "
        "where v.company_id = $2 and v.id = any($3::uuid[]) "
        "on conflict do nothing",
        pid, db_pg.as_uuid(company_id), ids,
    )


async def _plan_matriz_a_api(conn, plan: dict) -> dict:
    """Devuelve el plan con applies_to_vehicle_ids reconstruido desde la tabla
    puente, que es la forma que el frontend ya sabe leer."""
    if not plan:
        return plan
    filas = await conn.fetch(
        "select vehicle_id from maintenance_matrix_plan_vehicles where plan_id = $1",
        db_pg.as_uuid(plan["id"]),
    )
    salida = dict(plan)
    salida["applies_to_vehicle_ids"] = [str(f["vehicle_id"]) for f in filas]
    return salida


async def create_alert_once(company_id: str, alert_type: str, entity_type: str, entity_id: str,
                            message: str, severity: str = "warning"):
    """Crea una Alert sólo si no existe una del mismo tipo sin resolver para la entidad. Devuelve el id o None."""
    alerta = Alert(
        company_id=company_id, alert_type=alert_type, entity_type=entity_type,
        entity_id=entity_id, message=message, severity=severity,
    )
    sql, values = db_pg.build_insert(
        "alerts", ALERT_COLS, _modelo_a_fila(alerta.model_dump())
    )
    # Comprobar y crear en la misma transaccion: si no, dos barridos a la vez
    # pasarian los dos por el "no existe" y crearian la alerta por duplicado.
    async with db_pg.tx({"company_id": company_id}) as conn:
        existe = await conn.fetchval(
            "select id from alerts where company_id = $1 and alert_type = $2 "
            "and entity_id = $3 and not resolved",
            db_pg.as_uuid(company_id), alert_type, db_pg.as_uuid(entity_id),
        )
        if existe:
            return None
        await conn.execute(sql, *values)
    return alerta.id


async def _vehicle_km_per_day(company_id: str, vehicle_id: str):
    """Estima km/día promedio a partir de viajes completados de la unidad. None si no es calculable."""
    async with db_pg.tx({"company_id": company_id}) as conn:
        trips = db_pg.rows_to_api(await conn.fetch(
            "select km_start, km_end, start_date, end_date from trips "
            "where company_id = $1 and tracto_id = $2 "
            "and status = 'completado'::trip_status "
            "order by end_date desc nulls last limit 50",
            db_pg.as_uuid(company_id), db_pg.as_uuid(vehicle_id),
        ))
    total_km = 0.0
    total_days = 0.0
    for t in trips:
        ks = t.get("km_start")
        ke = t.get("km_end")
        if ks is None or ke is None or ke <= ks:
            continue
        total_km += (ke - ks)
        sd = t.get("start_date")
        ed = t.get("end_date")
        days = 1.0
        try:
            if sd and ed:
                sd_dt = datetime.fromisoformat(str(sd).replace("Z", "+00:00")) if isinstance(sd, str) else sd
                ed_dt = datetime.fromisoformat(str(ed).replace("Z", "+00:00")) if isinstance(ed, str) else ed
                days = max((ed_dt - sd_dt).total_seconds() / 86400.0, 0.5)
        except Exception:
            days = 1.0
        total_days += days
    if total_days > 0 and total_km > 0:
        return round(total_km / total_days, 2)
    return None


async def compute_tire_projection(company_id: str, tire: dict, vehicle: dict = None,
                                  config: dict = None, km_recorridos: int = None,
                                  latest_inspection: dict = None):
    """Proyección de vida de llanta. Devuelve wear_rate_mm_per_km, km_remaining,
    estimated_change_date, needs_review."""
    if config is None:
        config = await _company_config(company_id)
    min_legal = config.get("tire_critical_depth", DEFAULT_TIRE_CRITICAL_DEPTH)
    review_threshold = config.get("tire_review_km_threshold", DEFAULT_TIRE_REVIEW_KM)

    initial_depth = tire.get("initial_depth")
    last_depth = tire.get("last_depth")
    if last_depth is None and latest_inspection and latest_inspection.get("depths"):
        last_depth = min(latest_inspection["depths"])

    # km recorridos: usa el override o lo calcula desde el montaje activo
    km = km_recorridos
    if km is None:
        if vehicle is None and tire.get("current_vehicle_id"):
            vehicle = await _vehiculo_pg(company_id, tire["current_vehicle_id"])
        veh_odo = vehicle.get("odometer") if vehicle else None
        async with db_pg.tx({"company_id": company_id}) as conn:
            mount = db_pg.to_api(await conn.fetchrow(
                "select * from tire_mounts where tire_id = $1 and company_id = $2 "
                "and unmount_date is null order by mount_date desc limit 1",
                db_pg.as_uuid(tire["id"]), db_pg.as_uuid(company_id),
            ))
        mount_odo = mount.get("mount_odometer") if mount else None
        if veh_odo is not None and mount_odo is not None:
            diff = veh_odo - mount_odo
            km = diff if diff >= 0 else None

    wear = None
    if initial_depth is not None and last_depth is not None:
        wear = initial_depth - last_depth

    rate = None
    if km and km > 0 and wear and wear > 0:
        rate = wear / km

    km_remaining = None
    if rate and last_depth is not None:
        km_remaining = (last_depth - min_legal) / rate
        if km_remaining < 0:
            km_remaining = 0
        km_remaining = int(round(km_remaining))

    needs_review = False
    if last_depth is not None and last_depth <= min_legal:
        needs_review = True
    elif km_remaining is not None and km_remaining < review_threshold:
        needs_review = True

    estimated_change_date = None
    if km_remaining is not None and vehicle:
        kmpd = await _vehicle_km_per_day(company_id, vehicle["id"])
        if kmpd and kmpd > 0:
            try:
                est = datetime.now(timezone.utc) + timedelta(days=km_remaining / kmpd)
                estimated_change_date = est.isoformat()
            except Exception:
                estimated_change_date = None

    return {
        "wear_rate_mm_per_km": round(rate, 6) if rate else None,
        "km_remaining": km_remaining,
        "estimated_change_date": estimated_change_date,
        "needs_review": needs_review,
    }


async def compute_maintenance_status(company_id: str, vehicle: dict) -> dict:
    """Estado de mantenimiento: faltan X km para el próximo servicio."""
    config = await _company_config(company_id)
    anticipation = config.get("maintenance_anticipation_km", DEFAULT_MAINT_ANTICIPATION_KM)
    current_odo = vehicle.get("odometer", 0) or 0
    last_maint = vehicle.get("last_maintenance_km", 0) or 0

    interval_km = None
    plan_name = None

    # 1) Plan matricial asignado a esta unidad (km guardados en miles: 30 = 30000)
    # En Mongo esto buscaba el id dentro del array applies_to_vehicle_ids; aca
    # esa relacion esta normalizada, asi que es un join contra la tabla puente.
    async with db_pg.tx({"company_id": company_id}) as conn:
        matrix = db_pg.to_api(await conn.fetchrow(
            "select p.* from maintenance_matrix_plans p "
            "join maintenance_matrix_plan_vehicles pv on pv.plan_id = p.id "
            "where p.company_id = $1 and pv.vehicle_id = $2 limit 1",
            db_pg.as_uuid(company_id), db_pg.as_uuid(vehicle["id"]),
        ))
    if matrix:
        plan_name = matrix.get("name")
        kms = []
        for itv in matrix.get("intervals", []):
            km = itv.get("km")
            if km:
                km = km * 1000 if km < 1000 else km
                kms.append(km)
        if kms:
            interval_km = min(kms)

    # 2) Fallback: MaintenancePlan por tipo de vehículo con interval_km
    if interval_km is None:
        vt = vehicle.get("vehicle_type")
        # vehicle_type::text: vt sale de la fila del vehiculo y podria no ser un
        # valor del enum; comparando como texto eso devuelve "sin plan" en vez
        # de un 500 en medio del calculo del proximo servicio.
        async with db_pg.tx({"company_id": company_id}) as conn:
            plan = db_pg.to_api(await conn.fetchrow(
                "select * from maintenance_plans where company_id = $1 "
                "and vehicle_type::text = $2 and interval_km > 0 "
                "order by interval_km limit 1",
                db_pg.as_uuid(company_id), vt,
            ))
        if plan:
            plan_name = plan.get("name")
            interval_km = plan.get("interval_km")

    next_service_km = None
    km_remaining = None
    due_soon = False
    if interval_km:
        next_service_km = last_maint + interval_km
        km_remaining = next_service_km - current_odo
        due_soon = km_remaining <= anticipation

    return {
        "current_odometer": current_odo,
        "next_service_km": next_service_km,
        "km_remaining": km_remaining,
        "plan_name": plan_name,
        "due_soon": due_soon,
        "interval_km": interval_km,
    }


async def check_maintenance_due(company_id: str, vehicle: dict):
    """Si el mantenimiento está próximo, crea alerta (dedup) y notifica al chofer asignado + admin."""
    try:
        status = await compute_maintenance_status(company_id, vehicle)
        if status.get("km_remaining") is not None and status.get("due_soon"):
            km_rem = status["km_remaining"]
            plate = vehicle.get("plate", "")
            msg = (f"Mantenimiento próximo para {plate}: faltan {km_rem} km"
                   f"{' (' + status['plan_name'] + ')' if status.get('plan_name') else ''}")
            created = await create_alert_once(
                company_id, "maintenance_due", "vehicle", vehicle["id"], msg,
                "critical" if km_rem <= 0 else "warning"
            )
            if created:
                await notify_users(company_id, "Mantenimiento próximo", msg, "warning",
                                   target_role="admin", entity_type="vehicle", entity_id=vehicle["id"])
                driver_id = vehicle.get("assigned_driver_id")
                if driver_id:
                    await notify_users(company_id, "Mantenimiento próximo", msg, "warning",
                                       user_id=driver_id, entity_type="vehicle", entity_id=vehicle["id"])
    except Exception as e:
        logging.error(f"check_maintenance_due error: {e}")


async def check_tire_reviews(company_id: str, vehicle: dict):
    """Revisa las llantas montadas y crea alerta tire_review_due cuando necesitan revisión."""
    try:
        config = await _company_config(company_id)
        async with db_pg.tx({"company_id": company_id}) as conn:
            tires = db_pg.rows_to_api(await conn.fetch(
                "select * from tires where company_id = $1 and current_vehicle_id = $2 "
                "limit 50",
                db_pg.as_uuid(company_id), db_pg.as_uuid(vehicle["id"]),
            ))
        for tire in tires:
            proj = await compute_tire_projection(company_id, tire, vehicle, config)
            if proj.get("needs_review"):
                km_rem = proj.get("km_remaining")
                msg = (f"Llanta {tire.get('serial', '')} requiere revisión"
                       f"{f': ~{km_rem} km restantes' if km_rem is not None else ' (profundidad crítica)'}")
                created = await create_alert_once(
                    company_id, "tire_review_due", "tire", tire["id"], msg, "warning"
                )
                if created:
                    await notify_users(company_id, "Revisión de llanta", msg, "warning",
                                       target_role="admin", entity_type="tire", entity_id=tire["id"])
                    driver_id = vehicle.get("assigned_driver_id")
                    if driver_id:
                        await notify_users(company_id, "Revisión de llanta", msg, "warning",
                                           user_id=driver_id, entity_type="tire", entity_id=tire["id"])
    except Exception as e:
        logging.error(f"check_tire_reviews error: {e}")


async def apply_odometer_update(company_id: str, vehicle_id: str, new_odometer, actor_user_id: str = None):
    """Actualiza vehicle.odometer = max(actual, nuevo) y dispara checks de mantenimiento/llanta."""
    vehicle = await _vehiculo_pg(company_id, vehicle_id)
    if not vehicle:
        return
    current = vehicle.get("odometer", 0) or 0
    try:
        candidate = int(new_odometer) if new_odometer is not None else current
    except (TypeError, ValueError):
        candidate = current
    final_odo = max(current, candidate)
    if final_odo != current:
        await _actualizar_vehiculo(company_id, vehicle_id, {"odometer": final_odo})
        vehicle["odometer"] = final_odo
    # Disparar checks (no bloqueantes)
    await check_maintenance_due(company_id, vehicle)
    await check_tire_reviews(company_id, vehicle)


async def check_viatico_alert(company_id: str, trip_id: str):
    """Crea alerta viatico_low si (viatico_budget - total_expenses) < viatico_por_viaje."""
    try:
        trip = await _viaje_pg(company_id, trip_id)
        if not trip:
            return
        config = await _company_config(company_id)
        per_trip = config.get("viatico_por_viaje", DEFAULT_VIATICO_POR_VIAJE)
        budget = trip.get("viatico_budget")
        if budget is None:
            budget = per_trip
        spent = trip.get("total_expenses", 0) or 0
        remaining = budget - spent
        if remaining < per_trip:
            msg = (f"Viáticos bajos en viaje {trip.get('trip_number', trip_id)}: "
                   f"quedan S/ {round(remaining, 2)} (mínimo por viaje S/ {per_trip})")
            created = await create_alert_once(
                company_id, "viatico_low", "trip", trip_id, msg,
                "critical" if remaining < 0 else "warning"
            )
            if created:
                await notify_users(company_id, "Viáticos bajos", msg, "warning",
                                   target_role="admin", entity_type="trip", entity_id=trip_id)
                driver_id = trip.get("driver_id")
                if driver_id:
                    await notify_users(company_id, "Viáticos bajos", msg, "warning",
                                       user_id=driver_id, entity_type="trip", entity_id=trip_id)
    except Exception as e:
        logging.error(f"check_viatico_alert error: {e}")


async def _generate_document_alerts(company_id: str) -> int:
    """Genera alertas (y notificaciones críticas) para documentos por vencer. Devuelve nº creadas."""
    alerts_created = 0
    now = datetime.now(timezone.utc)
    alert_days = [60, 30, 15, 7, 3, 1, 0]
    async with db_pg.tx({"company_id": company_id}) as conn:
        documents = db_pg.rows_to_api(await conn.fetch(
            "select * from documents where company_id = $1 "
            "and expiry_date is not null limit 1000",
            db_pg.as_uuid(company_id),
        ))

    # Documento más reciente por (entidad, tipo): sólo éste rige bloqueo/resolución
    latest_by_type = {}
    for d in documents:
        if not d.get("expiry_date"):
            continue
        k = (d.get("entity_id"), d.get("document_type_id"))
        cur = latest_by_type.get(k)
        if cur is None or str(d.get("expiry_date", "")) > str(cur.get("expiry_date", "")):
            latest_by_type[k] = d

    for doc in documents:
        if not doc.get("expiry_date"):
            continue
        expiry = doc["expiry_date"]
        if isinstance(expiry, str):
            expiry = datetime.fromisoformat(expiry.replace("Z", "+00:00"))
        days_until = (expiry - now).days

        async with db_pg.tx({"company_id": company_id}) as conn:
            doc_type = db_pg.to_api(await conn.fetchrow(
                "select * from document_types where id = $1 and company_id = $2",
                db_pg.as_uuid(doc["document_type_id"]), db_pg.as_uuid(company_id),
            ))
        # vehicles sigue en Mongo; users ya corto a Postgres.
        entity = (
            await _vehiculo_pg(company_id, doc["entity_id"])
            or await _usuario_pg(company_id, doc["entity_id"])
        )

        # Regla Revisión Técnica: no aplica a unidades <= 4 años -> excluir de alertas y bloqueos
        if _revision_tecnica_no_aplica(doc_type, entity, doc.get("entity_type")):
            continue

        # Bloqueo automático por vencimiento / resolución al renovar (sólo el doc más reciente del tipo)
        is_latest = latest_by_type.get((doc.get("entity_id"), doc.get("document_type_id")), {}).get("id") == doc.get("id")
        block_rule = (doc_type or {}).get("block_rule", "solo_alerta")
        if is_latest and doc_type and block_rule != "solo_alerta":
            async with db_pg.tx({"company_id": company_id}) as conn:
                existing_block = await conn.fetchval(
                    "select id from blocks where company_id = $1 and entity_type = $2 "
                    "and entity_id = $3 and document_type_id = $4 and is_active",
                    db_pg.as_uuid(company_id), doc.get("entity_type"),
                    db_pg.as_uuid(doc.get("entity_id")),
                    db_pg.as_uuid(doc.get("document_type_id")),
                )
                if days_until > 0:
                    # Documento vigente nuevamente: cerrar bloqueos activos.
                    # resolved_by queda NULL a proposito: la columna es uuid con
                    # FK a users y el "system" que se escribia antes no es
                    # ningun usuario. Un bloqueo con resolved_at puesto y
                    # resolved_by nulo es, por definicion, uno que cerro el
                    # barrido automatico.
                    await conn.execute(
                        "update blocks set is_active = false, resolved_at = $1, "
                        "resolved_by = null where company_id = $2 and entity_type = $3 "
                        "and entity_id = $4 and document_type_id = $5 and is_active",
                        now, db_pg.as_uuid(company_id), doc.get("entity_type"),
                        db_pg.as_uuid(doc.get("entity_id")),
                        db_pg.as_uuid(doc.get("document_type_id")),
                    )
            if days_until <= 0 and not existing_block:
                await create_block_for_expired_doc(
                    company_id, doc_type, doc, doc.get("entity_type"), doc.get("entity_id")
                )

        for alert_day in alert_days:
            if days_until <= alert_day:
                async with db_pg.tx({"company_id": company_id}) as conn:
                    existing = await conn.fetchval(
                        "select id from alerts where company_id = $1 "
                        "and entity_id = $2 and alert_type = 'document_expiry' "
                        "and not resolved",
                        db_pg.as_uuid(company_id), db_pg.as_uuid(doc["id"]),
                    )
                if not existing:
                    severity = "critical" if days_until <= 0 else "warning" if days_until <= 7 else "info"
                    alert = {
                        "id": str(uuid.uuid4()),
                        "company_id": company_id,
                        "alert_type": "document_expiry",
                        "entity_type": doc["entity_type"],
                        "entity_id": doc["id"],
                        "message": f"{doc_type['name'] if doc_type else 'Documento'} de {(entity or {}).get('plate') or (entity or {}).get('name', 'N/A')} {'VENCIDO' if days_until <= 0 else f'vence en {days_until} días'}",
                        "severity": severity,
                        "is_read": False,
                        "resolved": False,
                        "created_at": now.isoformat()
                    }
                    await _insertar_alerta(company_id, alert)
                    alerts_created += 1
                    if severity == "critical":
                        await _crear_notificacion(company_id, {
                            "id": str(uuid.uuid4()),
                            "company_id": company_id,
                            "title": "⚠️ Documento Vencido",
                            "message": alert["message"],
                            "type": "alert",
                            "target_role": "admin",
                            "entity_type": "alert",
                            "entity_id": alert["id"],
                            "is_read": False,
                            "created_at": now.isoformat()
                        })
                break
    return alerts_created


async def run_maintenance_sweep():
    """Barrido periódico por empresa: documentos + mantenimiento + llantas + viáticos."""
    try:
        async with db_pg.tx_global("tarea de fondo: recorre todas las empresas") as conn:
            companies = db_pg.rows_to_api(
                await conn.fetch("select id from companies limit 1000")
            )
    except Exception as e:
        logging.error(f"run_maintenance_sweep companies error: {e}")
        return
    for c in companies:
        cid = c.get("id")
        if not cid:
            continue
        try:
            await _generate_document_alerts(cid)
        except Exception as e:
            logging.error(f"sweep documents error ({cid}): {e}")
        try:
            async with db_pg.tx({"company_id": cid}) as conn:
                vehicles = db_pg.rows_to_api(await conn.fetch(
                    "select * from vehicles where company_id = $1 limit 1000",
                    db_pg.as_uuid(cid),
                ))
            for v in vehicles:
                await check_maintenance_due(cid, v)
                await check_tire_reviews(cid, v)
        except Exception as e:
            logging.error(f"sweep vehicles error ({cid}): {e}")
        try:
            async with db_pg.tx({"company_id": cid}) as conn:
                trips = db_pg.rows_to_api(await conn.fetch(
                    "select id from trips where company_id = $1 "
                    "and status = any($2::trip_status[]) limit 1000",
                    db_pg.as_uuid(cid),
                    ["en_curso", "programado", "checklist_pendiente"],
                ))
            for t in trips:
                await check_viatico_alert(cid, t["id"])
        except Exception as e:
            logging.error(f"sweep trips error ({cid}): {e}")


# ============== AUTH ROUTES ==============
class SignupRequest(BaseModel):
    company_name: str
    ruc: str
    name: str
    email: str
    password: str
    phone: Optional[str] = None


@api_router.get("/tenant")
async def tenant_del_host(request: Request):
    """La empresa duena de esta direccion, para que el acceso lleve su marca.

    Publico y sin token a proposito: se consulta ANTES de autenticar, que es
    justo el problema que resuelve -hasta ahora la pantalla de acceso no podia
    saber de quien era-.

    Devuelve solo nombre, logo y color. NO devuelve subscription_status ni nada
    operativo, y desde luego no sunat_config, que lleva el token de la API de
    facturacion electronica.

    404 cuando el host no es de nadie, que es lo que responde en la landing, en
    fletepro.sisac.pe y en local. El frontend lo lee como "aqui no hay empresa,
    pinta la marca del producto", no como un fallo.

    Si: esto permite averiguar si un subdominio existe. Es el mismo dato que ya
    revela el DNS y que ve cualquiera que abra la direccion en el navegador, y
    es el precio de que un chofer reconozca la pantalla donde pone su PIN.
    """
    empresa = await tenant_host.empresa_desde_host(request.headers.get("host"))
    if not empresa:
        raise HTTPException(
            status_code=404, detail="Esta direccion no corresponde a ninguna empresa"
        )
    # El logo NO viaja aqui dentro. companies.logo_url guarda el PNG como
    # data-URI en base64, y el de G&E pesa 286 KB: devolverlo en este JSON
    # significaba 286 KB sin autenticar en CADA carga del login, imposibles de
    # cachear (el navegador no cachea un campo de un JSON) y regalados a
    # cualquiera que haga un bucle sobre el endpoint.
    #
    # Se manda una ruta, y el navegador la pide una vez y la guarda. Si el logo
    # ya es una URL normal -no un data-URI-, se devuelve tal cual: ahi no hay
    # nada que proxyar.
    logo = empresa.get("logo_url") or ""
    return {
        "slug": empresa["slug"],
        "name": empresa["name"],
        "logo_url": ("/api/tenant/logo" if logo.startswith("data:") else (logo or None)),
        "brand_color": empresa.get("brand_color"),
    }


# Tipos que este endpoint acepta servir. Es una lista blanca, no un filtro:
# logo_url lo escribe el admin de la empresa, y devolver el MIME que venga de
# la base permitiria guardar `data:text/html;...` y hacer que NUESTRO origen
# sirva HTML del inquilino. Eso es XSS almacenado, y en el origen donde vive su
# propia sesion. La cabecera nosniff global ayuda, pero no sustituye a esto.
_TIPOS_DE_LOGO = {"image/png", "image/jpeg", "image/gif", "image/webp", "image/svg+xml"}


@api_router.get("/tenant/logo")
async def tenant_logo_del_host(request: Request):
    """El logo de la empresa de este host, como imagen y cacheable.

    Publico, igual que /api/tenant y por el mismo motivo: se pinta antes de
    autenticar. La diferencia es que esto SI se puede cachear, asi que el coste
    es una peticion por navegador y no una por carga de pagina.
    """
    import base64
    import binascii
    import hashlib

    from fastapi import Response

    empresa = await tenant_host.empresa_desde_host(request.headers.get("host"))
    logo = (empresa or {}).get("logo_url") or ""
    if not logo.startswith("data:"):
        raise HTTPException(status_code=404, detail="Esta empresa no tiene logo")

    cabecera, _, datos = logo.partition(",")
    tipo = cabecera[len("data:"):].split(";")[0].strip().lower()
    if tipo not in _TIPOS_DE_LOGO or ";base64" not in cabecera:
        raise HTTPException(status_code=404, detail="Esta empresa no tiene logo")

    try:
        crudo = base64.b64decode(datos, validate=True)
    except (binascii.Error, ValueError):
        # Un data-URI corrupto en la base no es un 500: es una empresa sin logo.
        raise HTTPException(status_code=404, detail="Esta empresa no tiene logo")

    # ETag sobre el contenido: si el admin cambia el logo, cambia el ETag y el
    # navegador se lo baja de nuevo sin esperar a que venza la cache.
    etag = '"' + hashlib.sha256(crudo).hexdigest()[:16] + '"'
    if request.headers.get("if-none-match") == etag:
        return Response(status_code=304, headers={"ETag": etag})

    return Response(
        content=crudo,
        media_type=tipo,
        headers={
            "ETag": etag,
            # setdefault en el middleware de seguridad deja que este gane.
            "Cache-Control": "public, max-age=3600",
        },
    )


@api_router.post("/auth/signup", response_model=TokenResponse)
@limiter.limit("5/hour")
async def signup(request: Request, datos: SignupRequest):
    """Alta de una transportista con prueba gratuita. Publico y sin token.

    Crea la empresa y su usuario dueno en UNA transaccion: una empresa sin
    dueno no se puede administrar ni borrar por la interfaz, y quedaria como
    basura que solo se limpia a mano en la base.

    Va con tx_global porque todavia no hay empresa a la que fijar el contexto
    -se esta creando en esta misma llamada-, que es el caso 2 de los que
    documenta db_pg.tx_global.
    """
    nombre_empresa = (datos.company_name or "").strip()
    ruc = re.sub(r"\D", "", datos.ruc or "")
    nombre = (datos.name or "").strip()
    email = (datos.email or "").strip().lower()

    if len(nombre_empresa) < 2:
        raise HTTPException(status_code=400, detail="Falta el nombre de la empresa")
    if len(ruc) != 11:
        raise HTTPException(status_code=400, detail="El RUC debe tener 11 digitos")
    if len(nombre) < 2:
        raise HTTPException(status_code=400, detail="Falta tu nombre")
    if "@" not in email or "." not in email.split("@")[-1]:
        raise HTTPException(status_code=400, detail="El correo no es valido")
    if len(datos.password or "") < 8:
        raise HTTPException(status_code=400, detail="La contrasena necesita al menos 8 caracteres")

    ahora = datetime.now(timezone.utc)
    company = Company(
        name=nombre_empresa,
        ruc=ruc,
        email=email,
        phone=datos.phone,
    )
    fila_empresa = _modelo_a_fila(company.model_dump())
    fila_empresa["plan"] = "trial"
    fila_empresa["subscription_status"] = "trial"
    fila_empresa["trial_ends_at"] = ahora + timedelta(days=DIAS_DE_PRUEBA)

    owner = User(
        company_id=company.id,
        email=email,
        name=nombre,
        role=UserRole.OWNER,
        password_hash=hash_password(datos.password),
    )

    # El subdominio se elige ANTES de abrir la transaccion, no dentro: ver que
    # empresas ocupan que slug es una pregunta que cruza inquilinos, y aca
    # dentro la transaccion ya estaria acotada a la que se esta creando. La
    # ventana entre elegir y usar la cierra el indice unico de slug.
    fila_empresa["slug"] = await tenant_host.slug_libre(nombre_empresa, ruc)

    sql_empresa, val_empresa = db_pg.build_insert("companies", COMPANY_COLS, fila_empresa)
    sql_owner, val_owner = db_pg.build_insert("users", USER_COLS, _modelo_a_fila(owner.model_dump()))

    async with db_pg.tx_global("alta de una empresa nueva: aun no hay contexto de empresa") as conn:
        # El correo se comprueba en TODO el sistema, no dentro de la empresa.
        # El indice unico es (company_id, email), pero /auth/login busca
        # `where email = $1` sin empresa y se queda con la primera fila: dos
        # usuarios con el mismo correo en empresas distintas dejarian a uno de
        # los dos sin poder entrar nunca.
        if await conn.fetchval("select 1 from users where lower(email) = $1", email):
            raise HTTPException(status_code=409, detail="Ese correo ya tiene una cuenta")
        if await conn.fetchval("select 1 from companies where ruc = $1", ruc):
            raise HTTPException(status_code=409, detail="Ese RUC ya esta registrado")
        await conn.execute(sql_empresa, *val_empresa)
        await conn.execute(sql_owner, *val_owner)

    # La cache de hosts guarda tambien los negativos, para que un barrido de
    # subdominios inexistentes no sea una consulta por intento. Sin invalidar,
    # quien abriera su direccion recien creada dentro de esa ventana se
    # encontraria un 404 en la puerta de su propia empresa.
    tenant_host.invalidar_cache(fila_empresa["slug"])

    payload = {"user_id": owner.id, "company_id": company.id, "role": UserRole.OWNER.value}
    return TokenResponse(
        access_token=create_access_token(payload),
        refresh_token=create_refresh_token(payload),
        user={
            "id": owner.id, "name": nombre, "email": email,
            "role": UserRole.OWNER.value, "company_id": company.id,
            "company_name": nombre_empresa,
            # La direccion propia de la empresa. El alta ocurre en el host de
            # la marca, asi que es el unico momento en que se le puede decir al
            # cliente cual es la suya.
            "company_slug": fila_empresa["slug"],
            "trial_ends_at": fila_empresa["trial_ends_at"].isoformat(),
        },
    )


@api_router.post("/auth/login", response_model=TokenResponse)
@limiter.limit("10/minute")
async def login(request: Request, login_data: LoginRequest):
    user = None

    # El host acota la busqueda cuando la direccion es la de una empresa
    # (<slug>.sisac.pe). En la landing, en el acceso de rescate y en local
    # devuelve None y se busca en todo el sistema, igual que antes.
    #
    # Donde mas se nota es en el login por DNI: users_dni_idx NO es unico, asi
    # que sin empresa el fetchrow se queda con la fila que le toque y el
    # segundo chofer que comparta DNI con otro no puede entrar nunca, sin
    # ningun mensaje que lo explique.
    empresa_host = await tenant_host.empresa_desde_host(request.headers.get("host"))
    empresa_id = db_pg.as_uuid(empresa_host["id"]) if empresa_host else None

    # Admin login (email + password)
    if login_data.email and login_data.password:
        # Sin empresa en el host, el email identifica al usuario en TODO el
        # sistema. Con empresa, dentro de ella - que es lo que el indice unico
        # (company_id, email) siempre dio por supuesto.
        async with db_pg.tx_global("autenticacion: resolver la identidad antes de conocer la empresa") as conn:
            user = db_pg.to_api(await conn.fetchrow(
                "select * from users where email = $1 "
                "  and ($2::uuid is null or company_id = $2)",
                login_data.email, empresa_id,
            ))
        if not user:
            raise HTTPException(status_code=401, detail="Credenciales inválidas")
        if not user.get("password_hash"):
            raise HTTPException(status_code=401, detail="Credenciales inválidas")
        if not verify_password(login_data.password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="Credenciales inválidas")

    # Driver login (DNI + PIN)
    elif login_data.dni and login_data.pin:
        async with db_pg.tx_global("autenticacion: resolver la identidad antes de conocer la empresa") as conn:
            user = db_pg.to_api(await conn.fetchrow(
                "select * from users where dni = $1 "
                "  and ($2::uuid is null or company_id = $2)",
                login_data.dni, empresa_id,
            ))
        if not user:
            raise HTTPException(status_code=401, detail="Credenciales inválidas")

        # Check lockout
        if user.get("locked_until"):
            locked_until = user["locked_until"]
            if isinstance(locked_until, str):
                locked_until = datetime.fromisoformat(locked_until)
            if datetime.now(timezone.utc) < locked_until:
                raise HTTPException(status_code=403, detail="Cuenta bloqueada temporalmente")

        if not user.get("pin_hash"):
            raise HTTPException(status_code=401, detail="PIN no configurado")

        if not verify_password(login_data.pin, user["pin_hash"]):
            # Increment failed attempts
            failed_attempts = (user.get("failed_attempts") or 0) + 1
            bloqueo = (
                datetime.now(timezone.utc) + timedelta(minutes=15)
                if failed_attempts >= 5 else None
            )
            async with db_pg.tx_global("autenticacion: el intento fallido se registra antes de conocer la empresa") as conn:
                await conn.execute(
                    "update users set failed_attempts = $1, "
                    "locked_until = coalesce($2, locked_until) where id = $3",
                    failed_attempts, bloqueo, db_pg.as_uuid(user["id"]),
                )
            raise HTTPException(status_code=401, detail="Credenciales inválidas")

        # Reset failed attempts on successful login
        async with db_pg.tx_global("autenticacion: el intento fallido se registra antes de conocer la empresa") as conn:
            await conn.execute(
                "update users set failed_attempts = 0, locked_until = null where id = $1",
                db_pg.as_uuid(user["id"]),
            )
    else:
        raise HTTPException(status_code=400, detail="Se requiere email/password o DNI/PIN")
    
    if not user.get("is_active"):
        raise HTTPException(status_code=403, detail="Usuario desactivado")
    
    # Create tokens
    token_data = {
        "user_id": user["id"],
        "company_id": user["company_id"],
        "role": user["role"]
    }
    
    access_token = create_access_token(token_data)
    refresh_token = create_refresh_token(token_data)
    
    # Return user info (without sensitive data)
    user_response = {
        "id": user["id"],
        "company_id": user["company_id"],
        "name": user["name"],
        "email": user.get("email"),
        "dni": user.get("dni"),
        "role": user["role"],
        "force_password_change": user.get("force_password_change", False)
    }
    
    return TokenResponse(
        access_token=access_token,
        refresh_token=refresh_token,
        user=user_response
    )

@api_router.post("/auth/refresh", response_model=TokenResponse)
async def refresh_token(request: RefreshRequest):
    payload = decode_token(request.refresh_token)
    if payload.get("type") != "refresh":
        raise HTTPException(status_code=401, detail="Token inválido")
    
    async with db_pg.tx_global("autenticacion: resolver la identidad antes de conocer la empresa") as conn:
        fila = await conn.fetchrow(
            "select * from users where id = $1", db_pg.as_uuid(payload["user_id"])
        )
    user = db_pg.to_api(fila)
    if not user or not user.get("is_active"):
        raise HTTPException(status_code=401, detail="Usuario no válido")

    # Sin esto, el switch del superadmin se evaporaba a los 15 minutos: el
    # access token caduca, el frontend renueva en silencio, y token_data se
    # reconstruia desde la fila del usuario -o sea, desde su empresa propia-.
    # Volveria a la empresa del sistema a mitad de trabajo y sin avisar, que es
    # justo la clase de cosa que hace desconfiar de una herramienta.
    await _aplicar_empresa_del_token(payload, user)

    token_data = {
        "user_id": user["id"],
        "company_id": user["company_id"],
        "role": user["role"]
    }
    
    new_access_token = create_access_token(token_data)
    new_refresh_token = create_refresh_token(token_data)
    
    user_response = {
        "id": user["id"],
        "company_id": user["company_id"],
        "name": user["name"],
        "email": user.get("email"),
        "dni": user.get("dni"),
        "role": user["role"]
    }
    
    return TokenResponse(
        access_token=new_access_token,
        refresh_token=new_refresh_token,
        user=user_response
    )

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    # Strip sensitive fields before returning
    safe_user = {k: v for k, v in current_user.items() if k not in ("password_hash", "pin_hash", "_id")}
    return serialize_doc(safe_user)

# ============== COMPANY ROUTES ==============
@api_router.get("/companies")
async def get_companies(current_user: dict = Depends(require_roles("superadmin", "owner", "admin"))):
    # Superadmin sees ALL companies
    if current_user["role"] == "superadmin":
        async with db_pg.tx_global("superadmin: listar todas las empresas") as conn:
            filas = await conn.fetch("select * from companies order by name limit 100")
        return db_pg.rows_to_api(filas)
    # Others see only their own company
    company = await _empresa_pg(current_user["company_id"])
    return [company] if company else []

@api_router.get("/company")
async def get_current_company(current_user: dict = Depends(get_current_user)):
    company = await _empresa_pg(current_user["company_id"])
    return serialize_doc(company)

@api_router.post("/companies/{company_id}/switch")
async def switch_company(
    request: Request,
    company_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Superadmin can switch to any company context to manage it"""
    if current_user["role"] != "superadmin":
        raise HTTPException(status_code=403, detail="Solo superadmin puede cambiar de empresa")

    # Desde el subdominio de una empresa no: el token que sale de aca es de la
    # empresa destino, y la peticion siguiente a este mismo host lo rechazaria
    # por no coincidir (_exigir_host_de_la_empresa). Seria cambiar de empresa
    # para quedarse fuera. La consola del superadmin vive en el host de la
    # marca, que es donde esto funciona.
    if tenant_host.slug_desde_host(request.headers.get("host")):
        raise HTTPException(
            status_code=409,
            detail="Cambia de empresa desde la consola, no desde el subdominio "
                   "de un cliente",
        )

    # La empresa destino es, por definicion, distinta de la actual: hay que
    # mirarla desde fuera del aislamiento.
    async with db_pg.tx_global("superadmin: cambiar de contexto de empresa") as conn:
        company = db_pg.to_api(await conn.fetchrow(
            "select * from companies where id = $1", db_pg.as_uuid(company_id)
        ))
    if not company:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")

    # Que un superadmin entre en los datos de un cliente tiene que dejar
    # rastro, y el rastro va DENTRO de la empresa visitada: es su registro de
    # auditoria el que debe poder responder "quien de la plataforma entro aca y
    # cuando". Si no falla el switch por esto: un fallo al escribir el log no
    # es motivo para dejar sin herramienta a quien esta atendiendo una
    # incidencia.
    try:
        await create_audit_log(
            company_id=company_id,
            user_id=current_user["id"],
            user_name=current_user.get("name") or "superadmin",
            action="entrar_a_empresa",
            entity_type="company",
            entity_id=company_id,
            details={"empresa": company.get("name"), "slug": company.get("slug")},
        )
    except Exception:
        logger.exception("no se pudo registrar la entrada del superadmin a %s", company_id)

    # current_user es la FILA de users, o sea que la clave es "id". Con
    # ["user_id"] esto levantaba KeyError y el endpoint respondia 500: el
    # switch no es que no surtiera efecto, es que nunca llegaba a completarse.
    token_data = {
        "user_id": current_user["id"],
        "company_id": company_id,
        "role": "superadmin"
    }
    access_token = create_access_token(token_data)
    refresh_token = create_refresh_token(token_data)

    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "company": serialize_doc(company),
        "message": f"Cambiado a empresa: {company.get('name', company_id)}"
    }


@api_router.post("/auth/salir-de-empresa")
async def salir_de_empresa(current_user: dict = Depends(get_current_user)):
    """Devuelve al superadmin a su propia empresa.

    Existe como endpoint propio y no como "haz switch a tu empresa" porque el
    camino de vuelta tiene que estar siempre disponible: si la empresa en la
    que se entro se borra o se rompe, buscarse a uno mismo en un listado que
    quiza ya no carga no es un plan.
    """
    if current_user.get("role") != "superadmin":
        raise HTTPException(status_code=403, detail="Solo superadmin")

    propia = current_user.get("company_id_propio") or current_user["company_id"]
    token_data = {
        "user_id": current_user["id"],
        "company_id": propia,
        "role": "superadmin",
    }
    return {
        "access_token": create_access_token(token_data),
        "refresh_token": create_refresh_token(token_data),
        "company_id": propia,
        "message": "De vuelta en tu empresa",
    }

# ============== USER ROUTES ==============
@api_router.get("/users")
async def get_users(
    role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if role and role not in [r.value for r in UserRole]:
        return []
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(role, "role = $?::user_role", role)
    async with db_pg.tx(current_user) as conn:
        filas = await conn.fetch(
            "select * from users where " + f.where + " order by name limit 1000",
            *f.values,
        )
    # Los hashes no salen nunca de la base hacia la API.
    return [_sin_secretos(u) for u in db_pg.rows_to_api(filas)]

@api_router.get("/users/{user_id}")
async def get_user(user_id: str, current_user: dict = Depends(get_current_user)):
    user = await _usuario_pg(current_user["company_id"], user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return _sin_secretos(user)

@api_router.post("/users")
async def create_user(request: CreateUserRequest, current_user: dict = Depends(require_roles("owner", "admin"))):
    
    # Check if email/dni already exists within same company
    user = User(
        company_id=current_user["company_id"],
        email=request.email,
        dni=request.dni,
        name=request.name,
        role=request.role,
        license_number=request.license_number,
        license_expiry=request.license_expiry,
        phone=request.phone,
        created_by=current_user["id"]
    )
    
    if request.password:
        user.password_hash = hash_password(request.password)
    if request.pin:
        user.pin_hash = hash_password(request.pin)
    
    # Comprobar duplicados y dar de alta en la MISMA transaccion: separados,
    # dos altas simultaneas con el mismo email pasaban las dos comprobaciones.
    cid = current_user["company_id"]
    async with db_pg.tx(current_user) as conn:
        if request.email:
            if await conn.fetchval(
                "select 1 from users where email = $1 and company_id = $2",
                request.email, db_pg.as_uuid(cid),
            ):
                raise HTTPException(status_code=400, detail="Email ya registrado en esta empresa")
        if request.dni:
            if await conn.fetchval(
                "select 1 from users where dni = $1 and company_id = $2",
                request.dni, db_pg.as_uuid(cid),
            ):
                raise HTTPException(status_code=400, detail="DNI ya registrado en esta empresa")

        sql, values = db_pg.build_insert("users", USER_COLS, _modelo_a_fila(user.model_dump()))
        await conn.execute(sql, *values)
    return {"id": user.id, "message": "Usuario creado exitosamente"}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["owner", "admin"] and current_user["id"] != user_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    # Remove sensitive fields from update
    request.pop("password_hash", None)
    request.pop("pin_hash", None)
    request.pop("id", None)
    request.pop("company_id", None)
    
    request["updated_at"] = datetime.now(timezone.utc)
    request["id"] = user_id
    request["company_id"] = current_user["company_id"]

    async with db_pg.tx(current_user) as conn:
        if not await conn.fetchval(
            "select 1 from users where id = $1 and company_id = $2",
            db_pg.as_uuid(user_id), db_pg.as_uuid(current_user["company_id"]),
        ):
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        # USER_COLS hace de lista blanca: lo que venga en el body y no sea una
        # columna declarada se ignora, igual que antes lo ignoraba el modelo.
        sql, values = db_pg.build_update("users", USER_COLS, request, ["id", "company_id"])
        if sql:
            await conn.execute(sql, *values)

    return {"message": "Usuario actualizado"}

@api_router.post("/users/{user_id}/reset-pin")
async def reset_user_pin(user_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    
    new_pin = request.get("pin")
    if not new_pin or len(new_pin) != 6 or not new_pin.isdigit():
        raise HTTPException(status_code=400, detail="PIN debe ser de 6 dígitos")
    
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update users set pin_hash = $1, force_password_change = true, "
            "failed_attempts = 0, locked_until = null, updated_at = now() "
            "where id = $2 and company_id = $3",
            hash_password(new_pin),
            db_pg.as_uuid(user_id),
            db_pg.as_uuid(current_user["company_id"]),
        )
    
    return {"message": "PIN reseteado. El usuario deberá cambiarlo en su próximo inicio de sesión."}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(require_roles("owner", "admin"))):
    
    # Cannot delete owner users or self
    target_user = await _usuario_pg(current_user["company_id"], user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    if target_user.get("role") == "owner":
        raise HTTPException(status_code=400, detail="No se puede eliminar al propietario")
    
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="No puede eliminarse a sí mismo")
    
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "delete from users where id = $1 and company_id = $2",
            db_pg.as_uuid(user_id), db_pg.as_uuid(current_user["company_id"]),
        )

    return {"message": "Usuario eliminado"}

# ============== MULTI-TENANT / COMPANY ROUTES ==============
@api_router.get("/companies/{company_id}")
async def get_company(company_id: str, current_user: dict = Depends(get_current_user)):
    """Get company details"""
    if current_user["role"] != "owner" and current_user["company_id"] != company_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    company = await _empresa_pg(company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return serialize_doc(company)

@api_router.post("/companies")
async def create_company(request: dict = Body(...), current_user: dict = Depends(require_roles("superadmin", "owner"))):
    """Create new company (superadmin/owner only)"""
    
    company_id = str(uuid.uuid4())
    company = {
        "id": company_id,
        "name": request["name"],
        "ruc": request.get("ruc", ""),
        "address": request.get("address"),
        "phone": request.get("phone"),
        "email": request.get("email"),
        "config": request.get("config", {}),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    # Subdominio: el que pida quien crea la empresa, o uno derivado del nombre.
    # Fuera de la transaccion de abajo a proposito: esa esta acotada a la
    # empresa que se crea, y RLS le esconderia justo las demas empresas, que
    # son contra las que hay que comprobar la colision.
    if request.get("slug"):
        company["slug"] = _slug_pedido(request["slug"])
        if not await tenant_host.slug_esta_libre(company["slug"]):
            raise HTTPException(status_code=409, detail="Esa direccion ya esta ocupada")
    else:
        company["slug"] = await tenant_host.slug_libre(
            company["name"], company.get("ruc") or ""
        )

    # El contexto se fija en la empresa que se esta creando: la politica RLS
    # exige id = empresa_actual para insertar en companies, asi que esto entra
    # sin necesidad de saltarse el aislamiento.
    async with db_pg.tx({"company_id": company_id}) as conn:
        sql, values = db_pg.build_insert("companies", COMPANY_COLS, company)
        await conn.execute(sql, *values)

        # Create default admin user for the company
        if request.get("admin_email") and request.get("admin_password"):
            admin_user = {
                "id": str(uuid.uuid4()),
                "company_id": company_id,
                "email": request["admin_email"],
                "name": request.get("admin_name", "Administrador"),
                "role": "admin",
                "password_hash": hash_password(request["admin_password"]),
                "is_active": True,
                "failed_attempts": 0,
                "created_at": datetime.now(timezone.utc),
            }
            sql, values = db_pg.build_insert("users", USER_COLS, admin_user)
            await conn.execute(sql, *values)

    return {"id": company_id, "message": "Empresa creada exitosamente"}

@api_router.put("/companies/{company_id}")
async def update_company(company_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Update company (superadmin/owner or admin of that company)"""
    if current_user["role"] == "superadmin":
        pass  # superadmin can edit any company
    elif current_user["role"] in ["owner", "admin"] and current_user["company_id"] == company_id:
        pass  # owner/admin can edit their own company
    else:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    # Cambiar de subdominio no es editar un campo: la direccion anterior se
    # queda muerta y con ella los marcadores de los usuarios y la PWA que
    # tengan instalada con ese origen grabado. Hace falta una tabla de alias y
    # una redireccion (ver la nota al pie de la migracion 016), y hasta que
    # existan es mejor no poder que poder a medias.
    if "slug" in request:
        raise HTTPException(
            status_code=400,
            detail="La direccion web de la empresa no se puede cambiar todavia",
        )

    request.pop("id", None)
    request["updated_at"] = datetime.now(timezone.utc)
    request["id"] = company_id

    # El contexto va a la empresa EDITADA, no a la del usuario: un superadmin
    # puede editar cualquiera, y asi la politica RLS lo permite sin apagar el
    # aislamiento para el resto de la peticion.
    async with db_pg.tx({"company_id": company_id}) as conn:
        if not await conn.fetchval(
            "select 1 from companies where id = $1", db_pg.as_uuid(company_id)
        ):
            raise HTTPException(status_code=404, detail="Empresa no encontrada")
        sql, values = db_pg.build_update("companies", COMPANY_COLS, request, ["id"])
        if sql:
            await conn.execute(sql, *values)

    return {"message": "Empresa actualizada"}

@api_router.delete("/companies/{company_id}")
async def delete_company(company_id: str, current_user: dict = Depends(require_roles("superadmin", "owner"))):
    """Delete company and all its data (superadmin/owner only)"""
    
    # En Postgres hay que quitar TODA fila que apunte a la empresa antes que la
    # empresa misma: el corte 004 devolvio las FKs y ahora la base las exige.
    #
    # El orden correcto depende de esas FKs. En vez de mantener a mano una lista
    # ordenada que se desincroniza en cuanto alguien agrega una tabla, se
    # intentan todas y se reintentan las que fallan por clave foranea: en cada
    # vuelta caen las hojas y en la siguiente sus padres. Si una vuelta entera
    # no logra avanzar, hay un ciclo y se corta en vez de dejar la empresa a
    # medio eliminar.
    async with db_pg.tx({"company_id": company_id}) as conn:
        # El corte 013 trajo el unico CICLO de FKs del esquema:
        # work_orders.issue_id -> issues y issues.work_order_id -> work_orders.
        # El bucle de mas abajo no puede deshacerlo solo -- ninguna de las dos
        # llega a borrarse nunca y termina abortando con "dependencias sin
        # resolver". Se corta una punta antes de empezar: las dos columnas son
        # nulables y la empresa se va entera igual.
        await conn.execute(
            "update work_orders set issue_id = null where company_id = $1",
            db_pg.as_uuid(company_id),
        )

        pendientes = [r["t"] for r in await conn.fetch(
            "select table_name as t from information_schema.columns "
            "where table_schema = 'public' and column_name = 'company_id'"
        )]
        while pendientes:
            quedan = []
            for tabla in pendientes:
                try:
                    async with conn.transaction():  # savepoint: el fallo no aborta todo
                        await conn.execute(
                            'delete from "' + tabla + '" where company_id = $1',
                            db_pg.as_uuid(company_id),
                        )
                except db_pg.ForeignKeyViolationError:
                    quedan.append(tabla)
            if len(quedan) == len(pendientes):
                raise HTTPException(
                    status_code=500,
                    detail="No se pudo eliminar la empresa: dependencias sin resolver en "
                           + ", ".join(quedan),
                )
            pendientes = quedan

        await conn.execute(
            "delete from companies where id = $1", db_pg.as_uuid(company_id)
        )

    # Ya no hay nada que barrer en Mongo: con el corte 013 las 50 tablas viven
    # en Postgres y el bucle de arriba se las lleva todas. La lista de
    # colecciones sobrevivientes desaparece junto con la ultima de ellas.

    return {"message": "Empresa y todos sus datos eliminados"}

@api_router.get("/companies/{company_id}/stats")
async def get_company_stats(company_id: str, current_user: dict = Depends(get_current_user)):
    """Get company statistics (owner or admin)"""
    if current_user["role"] != "owner" and current_user["company_id"] != company_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    async with db_pg.tx({"company_id": company_id}) as conn:
        total_ots, ots_abiertas = await conn.fetchrow(
            "select count(*), count(*) filter (where status <> 'completada') "
            "from work_orders where company_id = $1",
            db_pg.as_uuid(company_id),
        )

    stats = {
        "users": await _contar_usuarios(company_id),
        "vehicles": await _contar_vehiculos(company_id),
        "trips": await _contar_viajes(company_id),
        "active_trips": await _contar_viajes(company_id, "en_curso"),
        "work_orders": total_ots,
        "open_work_orders": ots_abiertas,
    }
    return stats

# ============== VEHICLE ROUTES ==============
@api_router.get("/vehicles")
async def get_vehicles(
    vehicle_type: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if vehicle_type and vehicle_type not in [t.value for t in VehicleType]:
        return []
    if status and status not in [e.value for e in VehicleStatus]:
        return []
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(vehicle_type, "vehicle_type = $?::vehicle_type", vehicle_type)
    f.si(status, "status = $?::vehicle_status", status)
    async with db_pg.tx(current_user) as conn:
        filas = await conn.fetch(
            "select * from vehicles where " + f.where + " order by plate limit 1000",
            *f.values,
        )
    return db_pg.rows_to_api(filas)

@api_router.get("/vehicles/{vehicle_id}")
async def get_vehicle(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    vehicle = await _vehiculo_pg(current_user["company_id"], vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    return serialize_doc(vehicle)

@api_router.post("/vehicles")
async def create_vehicle(request: CreateVehicleRequest, current_user: dict = Depends(require_roles("owner", "admin", "flota"))):
    await _exigir_cupo_de_vehiculos(current_user)

    # Check if plate already exists
    async with db_pg.tx(current_user) as conn:
        existing = await conn.fetchval(
            "select 1 from vehicles where plate = $1 and company_id = $2",
            request.plate.upper(), db_pg.as_uuid(current_user["company_id"]),
        )
    if existing:
        raise HTTPException(status_code=400, detail="Placa ya registrada")
    
    vehicle = Vehicle(
        company_id=current_user["company_id"],
        plate=request.plate.upper(),
        vehicle_type=request.vehicle_type,
        brand=request.brand,
        model=request.model,
        year=request.year,
        vin=request.vin,
        color=request.color,
        fuel_capacity=request.fuel_capacity,
        tire_config=request.tire_config,
        axle_config=request.axle_config,
        created_by=current_user["id"]
    )
    
    doc = vehicle.model_dump()
    for key, value in doc.items():
        if isinstance(value, datetime):
            doc[key] = value.isoformat()
    
    sql, values = db_pg.build_insert("vehicles", VEHICLE_COLS, _modelo_a_fila(vehicle.model_dump()))
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": vehicle.id, "message": "Vehículo creado exitosamente"}

@api_router.put("/vehicles/{vehicle_id}")
async def update_vehicle(vehicle_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "flota"))):
    request.pop("id", None)
    request.pop("company_id", None)
    request.pop("axle_config_history", None)  # managed server-side
    request["updated_at"] = datetime.now(timezone.utc).isoformat()

    if "plate" in request:
        request["plate"] = request["plate"].upper()

    historia_nueva = None

    # Track axle_config changes in history
    if "axle_config" in request:
        current = await _vehiculo_pg(current_user["company_id"], vehicle_id)
        if current is not None and current.get("axle_config") != request["axle_config"]:
            historia_nueva = {
                "axle_config": request["axle_config"],
                "date": datetime.now(timezone.utc).isoformat(),
                "changed_by": current_user["id"],
            }

    existe = await _actualizar_vehiculo(current_user["company_id"], vehicle_id, request)
    if not existe:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")

    # El historial de configuracion de ejes era un $push de Mongo; en jsonb es
    # concatenar al array, con coalesce por si todavia es null.
    if historia_nueva is not None:
        async with db_pg.tx(current_user) as conn:
            await conn.execute(
                "update vehicles set axle_config_history = "
                "coalesce(axle_config_history, '[]'::jsonb) || $1::jsonb "
                "where id = $2 and company_id = $3",
                [historia_nueva],
                db_pg.as_uuid(vehicle_id),
                db_pg.as_uuid(current_user["company_id"]),
            )

    return {"message": "Vehículo actualizado"}

@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, current_user: dict = Depends(require_roles("owner", "admin"))):
    async with db_pg.tx(current_user) as conn:
        borrado = await conn.fetchval(
            "delete from vehicles where id = $1 and company_id = $2 returning id",
            db_pg.as_uuid(vehicle_id), db_pg.as_uuid(current_user["company_id"]),
        )

    if not borrado:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    
    return {"message": "Vehículo eliminado"}

# ============== TABLAS EN POSTGRES: VIAJES ==============
# trips, couplings, units y routes ya cortaron (migracion 007). Ninguna FK
# suya sale hacia Mongo: todas apuntan a companies, users, vehicles o entre
# ellas mismas, y esas ya estaban en Postgres.
#
# checklist_id y settlement_id van sin FK a proposito: checklists y settlements
# siguen en Mongo, asi que la fila destino no existe aca.

TRIP_COLS = {
    "id": "uuid", "company_id": "uuid", "trip_number": "text",
    "tracto_id": "uuid", "carreta_id": "uuid", "driver_id": "uuid",
    "route_id": "uuid", "client_name": "text", "cargo_description": "text",
    "cargo_weight": "float", "status": "enum:trip_status",
    "is_round_trip": "bool", "scheduled_date": "ts",
    "start_date": "ts", "end_date": "ts", "km_start": "int", "km_end": "int",
    "total_advance": "float", "total_expenses": "float",
    "checklist_id": "uuid", "checklist_approved": "bool",
    "checklist_result": "text", "settlement_id": "uuid",
    "settlement_status": "text",
    "viatico_budget": "float", "viatico_days": "int", "viatico_daily": "float",
    "notes": "text", "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}

COUPLING_COLS = {
    "id": "uuid", "company_id": "uuid", "tracto_id": "uuid",
    "carreta_id": "uuid", "trip_id": "uuid",
    "start_date": "ts", "end_date": "ts", "created_by": "uuid",
}

UNIT_COLS = {
    "id": "uuid", "company_id": "uuid", "tracto_id": "uuid",
    "carreta_id": "uuid", "driver_id": "uuid", "status": "text",
    "epp_items": "json", "active": "bool",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}

ROUTE_COLS = {
    "id": "uuid", "company_id": "uuid", "name": "text", "origin": "text",
    "destination": "text", "distance_km": "float",
    "estimated_hours": "float", "toll_cost": "float", "created_at": "ts",
}


# ============== TABLA EN POSTGRES: VEHICULOS ==============
# vehicles ya corto (db/migrations/006_corte_vehicles.sql). Es la tabla mas
# referenciada del modelo operativo, pero sus cuatro FKs apuntan a companies,
# users y proveedores, que ya estaban en Postgres: no tuvo ninguna saliente.
#
# Las tablas que la referencian (tires, trips, couplings, units, checklists,
# work_orders...) tambien cruzaron ya; las ultimas, en el corte 013.

VEHICLE_COLS = {
    "id": "uuid", "company_id": "uuid", "plate": "text",
    "vehicle_type": "enum:vehicle_type", "brand": "text", "model": "text",
    "year": "int", "vin": "text", "color": "text",
    "status": "enum:vehicle_status", "odometer": "int",
    "fuel_capacity": "float", "tire_config": "text",
    "axle_config": "json", "axle_config_history": "json",
    "assigned_driver_id": "uuid", "photo_url": "text",
    "proveedor_id": "uuid", "viatico_fijo": "float",
    "last_maintenance_km": "int", "last_maintenance_date": "ts",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}


# ============== VEHICLE DRIVER ASSIGNMENT ==============
@api_router.post("/vehicles/{vehicle_id}/assign-driver")
async def assign_driver_to_vehicle(vehicle_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "operaciones"))):
    driver_id = request.get("driver_id")

    # If assigning (not unassigning), validate driver exists
    if driver_id:
        driver = await _usuario_pg(current_user["company_id"], driver_id)
        if driver and driver.get("role") != "chofer":
            driver = None
        if not driver:
            raise HTTPException(status_code=404, detail="Chofer no encontrado")

        # Remove this driver from any other vehicle
        async with db_pg.tx(current_user) as conn:
            await conn.execute(
                "update vehicles set assigned_driver_id = null, updated_at = now() "
                "where company_id = $1 and assigned_driver_id = $2",
                db_pg.as_uuid(current_user["company_id"]), db_pg.as_uuid(driver_id),
            )

    existe = await _actualizar_vehiculo(
        current_user["company_id"], vehicle_id, {"assigned_driver_id": driver_id}
    )

    if not existe:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")

    return {"message": "Chofer asignado" if driver_id else "Chofer desasignado"}

# ============== VEHICLE EQUIPMENT (EPP) ==============
@api_router.get("/vehicles/{vehicle_id}/equipment")
async def get_vehicle_equipment(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        doc = db_pg.to_api(await conn.fetchrow(
            "select * from vehicle_equipment where vehicle_id = $1 and company_id = $2",
            db_pg.as_uuid(vehicle_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    if not doc:
        # Return default EPP items
        return {
            "vehicle_id": vehicle_id,
            "items": [
                {"name": "botiquin", "label": "Botiquín", "quantity": 0, "condition": "pendiente", "expiry_date": None},
                {"name": "extintor", "label": "Extintor", "quantity": 0, "condition": "pendiente", "expiry_date": None},
                {"name": "chaleco", "label": "Chalecos Reflectivos", "quantity": 0, "condition": "pendiente", "expiry_date": None},
                {"name": "casco", "label": "Cascos", "quantity": 0, "condition": "pendiente", "expiry_date": None},
                {"name": "guantes", "label": "Guantes", "quantity": 0, "condition": "pendiente", "expiry_date": None},
                {"name": "lentes", "label": "Lentes de Seguridad", "quantity": 0, "condition": "pendiente", "expiry_date": None},
                {"name": "conos", "label": "Conos/Triángulos", "quantity": 0, "condition": "pendiente", "expiry_date": None},
                {"name": "linterna", "label": "Linterna", "quantity": 0, "condition": "pendiente", "expiry_date": None},
            ]
        }
    return serialize_doc(doc)

@api_router.put("/vehicles/{vehicle_id}/equipment")
async def update_vehicle_equipment(vehicle_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "flota", "almacen"))):
    items = request.get("items", [])

    # Buscar y guardar en la misma transaccion: dos guardados simultaneos no
    # pueden pasar los dos por el "no existe" y dejar dos filas de EPP para el
    # mismo vehiculo.
    async with db_pg.tx(current_user) as conn:
        existing = await conn.fetchval(
            "select id from vehicle_equipment "
            "where vehicle_id = $1 and company_id = $2 for update",
            db_pg.as_uuid(vehicle_id), db_pg.as_uuid(current_user["company_id"]),
        )
        if existing:
            await conn.execute(
                "update vehicle_equipment set items = $1, updated_at = now(), "
                "updated_by = $2 where id = $3",
                items, db_pg.as_uuid(current_user["id"]), existing,
            )
        else:
            equipment = VehicleEquipment(
                company_id=current_user["company_id"],
                vehicle_id=vehicle_id,
                items=items,
                updated_by=current_user["id"]
            )
            sql, values = db_pg.build_insert(
                "vehicle_equipment", VEHICLE_EQUIPMENT_COLS,
                _modelo_a_fila(equipment.model_dump()),
            )
            await conn.execute(sql, *values)

    return {"message": "Equipamiento actualizado"}

# ============== VIÁTICOS BUDGET ==============
@api_router.post("/trips/{trip_id}/viatico-budget")
async def set_viatico_budget(trip_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "operaciones", "contabilidad"))):

    budget = float(request.get("budget", 0))
    if budget < 0:
        raise HTTPException(status_code=400, detail="El presupuesto debe ser positivo")
    days = int(request.get("days", 1))
    if days < 1:
        days = 1
    daily = round(budget / days, 2)

    # _actualizar_viaje pone updated_at y devuelve False si el viaje no existe,
    # que es lo que antes decia matched_count == 0.
    actualizado = await _actualizar_viaje(
        current_user["company_id"], trip_id,
        {"viatico_budget": budget, "viatico_days": days, "viatico_daily": daily},
    )
    if not actualizado:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")

    return {"message": "Presupuesto de viáticos asignado", "daily": daily}

# ============== COUPLING ROUTES ==============
@api_router.post("/couplings")
async def create_coupling(request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    coupling = CouplingHistory(
        company_id=current_user["company_id"],
        tracto_id=request["tracto_id"],
        carreta_id=request["carreta_id"],
        trip_id=request.get("trip_id"),
        created_by=current_user["id"]
    )
    
    doc = coupling.model_dump()
    doc["start_date"] = doc["start_date"].isoformat()
    
    sql, values = db_pg.build_insert("couplings", COUPLING_COLS, doc)
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": coupling.id, "message": "Enganche registrado"}

@api_router.get("/couplings")
async def get_couplings(
    vehicle_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # Filtros repite el mismo $n en las dos ramas del or: un solo parametro
    # comparado contra las dos columnas, que es lo que hacia el $or de Mongo.
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(vehicle_id, "(tracto_id = $? or carreta_id = $?)", db_pg.as_uuid(vehicle_id))
    async with db_pg.tx(current_user) as conn:
        couplings = db_pg.rows_to_api(await conn.fetch(
            "select * from couplings where " + f.where
            + " order by start_date desc nulls last limit 100",
            *f.values,
        ))
    return couplings

@api_router.put("/couplings/{coupling_id}")
async def update_coupling(coupling_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Actualiza un enganche. Se usa para DESACOPLAR (enviar end_date)."""
    async with db_pg.tx(current_user) as conn:
        coupling = db_pg.to_api(await conn.fetchrow(
            "select * from couplings where id = $1 and company_id = $2",
            db_pg.as_uuid(coupling_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    if not coupling:
        raise HTTPException(status_code=404, detail="Enganche no encontrado")

    update_data = {}
    if "end_date" in request:
        end_date = request["end_date"]
        if isinstance(end_date, datetime):
            end_date = end_date.isoformat()
        update_data["end_date"] = end_date or datetime.now(timezone.utc).isoformat()
    if "trip_id" in request:
        update_data["trip_id"] = request["trip_id"]
    if not update_data:
        update_data["end_date"] = datetime.now(timezone.utc).isoformat()

    update_data["id"] = coupling_id
    update_data["company_id"] = current_user["company_id"]
    sql, values = db_pg.build_update("couplings", COUPLING_COLS, update_data, ["id", "company_id"])
    async with db_pg.tx(current_user) as conn:
        if sql:
            await conn.execute(sql, *values)
    return {"id": coupling_id, "message": "Enganche actualizado"}

# ============== DOCUMENT TYPE ROUTES ==============
@api_router.get("/document-types")
async def get_document_types(
    applies_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(applies_to, "applies_to = $?", applies_to)
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from document_types where " + f.where
            + " order by name limit 100", *f.values
        ))

@api_router.post("/document-types")
async def create_document_type(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    doc_type = DocumentType(
        company_id=current_user["company_id"],
        name=request["name"],
        applies_to=request["applies_to"],
        is_critical=request.get("is_critical", False),
        requires_expiry=request.get("requires_expiry", True),
        alert_days=request.get("alert_days", [60, 30, 15, 7, 3, 1]),
        block_rule=request.get("block_rule", BlockRule.SOLO_ALERTA)
    )
    
    sql, values = db_pg.build_insert(
        "document_types", DOCUMENT_TYPE_COLS, _modelo_a_fila(doc_type.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": doc_type.id, "message": "Tipo de documento creado"}

# ============== DOCUMENT ROUTES ==============
@api_router.get("/documents")
async def get_documents(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(entity_type, "entity_type = $?", entity_type)
    f.si(entity_id, "entity_id = $?", db_pg.as_uuid(entity_id))
    # status::text: el valor lo elige el cliente y uno que no exista en el enum
    # devuelve lista vacia, como hacia Mongo, en vez de un 500 por el cast.
    f.si(status, "status::text = $?", status)
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from documents where " + f.where
            + " order by expiry_date desc nulls last limit 1000", *f.values
        ))

@api_router.post("/documents")
async def create_document(request: CreateDocumentRequest, current_user: dict = Depends(get_current_user)):
    document = Document(
        company_id=current_user["company_id"],
        document_type_id=request.document_type_id,
        entity_type=request.entity_type,
        entity_id=request.entity_id,
        number=request.number,
        issue_date=request.issue_date,
        expiry_date=request.expiry_date,
        notes=request.notes,
        created_by=current_user["id"]
    )
    
    # Determine initial status based on expiry
    if request.expiry_date:
        days_until = (request.expiry_date - datetime.now(timezone.utc)).days
        if days_until < 0:
            document.status = DocumentStatus.VENCIDO
        elif days_until <= 30:
            document.status = DocumentStatus.POR_VENCER
        else:
            document.status = DocumentStatus.VIGENTE
    
    sql, values = db_pg.build_insert(
        "documents", DOCUMENT_COLS, _modelo_a_fila(document.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": document.id, "message": "Documento creado"}

@api_router.put("/documents/{document_id}")
async def update_document(document_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    request.pop("id", None)
    request.pop("company_id", None)
    datos = dict(request)
    datos["updated_at"] = datetime.now(timezone.utc)
    datos["id"] = document_id
    datos["company_id"] = current_user["company_id"]

    # DOCUMENT_COLS actua de lista blanca: lo que no sea una columna declarada
    # se ignora, igual que en el resto de los cortes.
    sql, values = db_pg.build_update(
        "documents", DOCUMENT_COLS, datos, ["id", "company_id"]
    )
    async with db_pg.tx(current_user) as conn:
        existia = await conn.fetchval(sql + " returning id", *values)
    if existia is None:
        raise HTTPException(status_code=404, detail="Documento no encontrado")

    return {"message": "Documento actualizado"}

@api_router.post("/documents/{document_id}/approve")
async def approve_document(document_id: str, current_user: dict = Depends(require_roles("owner", "admin", "flota"))):
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update documents set status = $1::document_status, approved_by = $2, "
            "approved_at = now(), updated_at = now() "
            "where id = $3 and company_id = $4",
            DocumentStatus.APROBADO.value,
            db_pg.as_uuid(current_user["id"]),
            db_pg.as_uuid(document_id),
            db_pg.as_uuid(current_user["company_id"]),
        )

    return {"message": "Documento aprobado"}

@api_router.get("/documents/matrix")
async def get_documents_matrix(
    entity_type: str = "vehicle",
    current_user: dict = Depends(get_current_user)
):
    """Get document matrix for vehicles or drivers"""
    # Map entity_type to applies_to field
    applies_to_map = {"vehicle": "vehiculo", "user": "chofer"}
    applies_to = applies_to_map.get(entity_type, entity_type)
    
    # Get document types
    async with db_pg.tx(current_user) as conn:
        doc_types = db_pg.rows_to_api(await conn.fetch(
            "select * from document_types where company_id = $1 and applies_to = $2 "
            "order by name limit 100",
            db_pg.as_uuid(current_user["company_id"]), applies_to,
        ))
    
    # Get entities
    if entity_type == "vehicle":
        async with db_pg.tx(current_user) as conn:
            entities = db_pg.rows_to_api(await conn.fetch(
                "select * from vehicles where company_id = $1 order by plate limit 1000",
                db_pg.as_uuid(current_user["company_id"]),
            ))
    else:
        async with db_pg.tx(current_user) as conn:
            filas = await conn.fetch(
                "select * from users where company_id = $1 "
                "and role = 'chofer'::user_role limit 1000",
                db_pg.as_uuid(current_user["company_id"]),
            )
        entities = [_sin_secretos(u) for u in db_pg.rows_to_api(filas)]
    
    # Get all documents
    async with db_pg.tx(current_user) as conn:
        documents = db_pg.rows_to_api(await conn.fetch(
            "select * from documents where company_id = $1 and entity_type = $2 "
            "limit 10000",
            db_pg.as_uuid(current_user["company_id"]), entity_type,
        ))
    
    # Build matrix
    doc_map = {}
    for doc in documents:
        key = f"{doc['entity_id']}_{doc['document_type_id']}"
        if key not in doc_map or doc.get("expiry_date", "") > doc_map[key].get("expiry_date", ""):
            doc_map[key] = doc
    
    matrix = []
    for entity in entities:
        row = {
            "entity": serialize_doc(entity),
            "documents": {},
            "no_aplica": {}
        }
        for doc_type in doc_types:
            key = f"{entity['id']}_{doc_type['id']}"
            na = _revision_tecnica_no_aplica(doc_type, entity, entity_type)
            cell = serialize_doc(doc_map.get(key))
            if na:
                row["no_aplica"][doc_type["id"]] = True
                if cell is None:
                    cell = {"no_aplica": True}
                else:
                    cell["no_aplica"] = True
            row["documents"][doc_type["id"]] = cell
        matrix.append(row)
    
    return {
        "document_types": [serialize_doc(dt) for dt in doc_types],
        "matrix": matrix
    }

# ============== ALERT ROUTES ==============
@api_router.get("/alerts")
async def get_alerts(
    resolved: Optional[bool] = None,
    severity: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    if resolved is not None:
        f.agregar("resolved = $?", bool(resolved))
    f.si(severity, "severity = $?", severity)
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from alerts where " + f.where
            + " order by created_at desc nulls last limit 500", *f.values
        ))

@api_router.post("/alerts/{alert_id}/resolve")
async def resolve_alert(alert_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update alerts set resolved = true, is_read = true "
            "where id = $1 and company_id = $2",
            db_pg.as_uuid(alert_id), db_pg.as_uuid(current_user["company_id"]),
        )
    return {"message": "Alerta resuelta"}

# ============== OPERATIONAL BLOCKS ROUTES ==============
@api_router.get("/blocks")
async def get_blocks(
    is_active: Optional[bool] = True,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    if is_active is not None:
        f.agregar("is_active = $?", bool(is_active))
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from blocks where " + f.where
            + " order by created_at desc nulls last limit 500", *f.values
        ))

@api_router.post("/blocks/{block_id}/resolve")
async def resolve_block(block_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "flota", "operaciones"))):
    
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update blocks set is_active = false, resolved_at = now(), "
            "resolved_by = $1 where id = $2 and company_id = $3",
            db_pg.as_uuid(current_user["id"]),
            db_pg.as_uuid(block_id),
            db_pg.as_uuid(current_user["company_id"]),
        )
    return {"message": "Bloqueo resuelto"}

# ============== ROUTE ROUTES ==============
@api_router.get("/routes")
async def get_routes(current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        routes = db_pg.rows_to_api(await conn.fetch(
            "select * from routes where company_id = $1 limit 500",
            db_pg.as_uuid(current_user["company_id"]),
        ))
    return routes

@api_router.post("/routes")
async def create_route(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "operaciones"))):
    
    route = Route(
        company_id=current_user["company_id"],
        name=request["name"],
        origin=request["origin"],
        destination=request["destination"],
        distance_km=request["distance_km"],
        estimated_hours=request.get("estimated_hours", 0),
        toll_cost=request.get("toll_cost", 0)
    )
    
    doc = route.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    
    sql, values = db_pg.build_insert("routes", ROUTE_COLS, doc)
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": route.id, "message": "Ruta creada"}

# ============== TRIP ROUTES ==============
@api_router.get("/trips")
async def get_trips(
    status: Optional[str] = None,
    driver_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # El chofer ve solo los suyos: pisa el driver_id que haya pedido, igual que
    # hacia el query de Mongo al reasignar la clave.
    if current_user["role"] == "chofer":
        driver_id = current_user["id"]

    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    # status::text y no un cast a trip_status: el valor lo elige el cliente, y
    # uno que no exista en el enum haria fallar el cast con un 500. Comparando
    # como texto un status desconocido devuelve lista vacia, igual que Mongo.
    f.si(status, "status::text = $?", status)
    f.si(driver_id, "driver_id = $?", db_pg.as_uuid(driver_id))
    async with db_pg.tx(current_user) as conn:
        trips = db_pg.rows_to_api(await conn.fetch(
            "select * from trips where " + f.where
            + " order by scheduled_date desc nulls last limit 1000",
            *f.values,
        ))
    return trips

@api_router.get("/trips/{trip_id}")
async def get_trip(trip_id: str, current_user: dict = Depends(get_current_user)):
    trip = await _viaje_pg(current_user["company_id"], trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")
    return serialize_doc(trip)

@api_router.post("/trips")
async def create_trip(request: CreateTripRequest, current_user: dict = Depends(require_roles("owner", "admin", "operaciones"))):
    
    # Validate all blocks using helper function
    validation = await validate_trip_can_be_assigned(
        current_user["company_id"],
        request.tracto_id,
        request.carreta_id,
        request.driver_id
    )
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail="; ".join(validation["errors"]))
    
    # Generate trip number
    count = await _contar_viajes(current_user["company_id"])
    trip_number = f"VJ-{count + 1:05d}"
    
    trip = Trip(
        company_id=current_user["company_id"],
        trip_number=trip_number,
        tracto_id=request.tracto_id,
        carreta_id=request.carreta_id,
        driver_id=request.driver_id,
        route_id=request.route_id,
        client_name=request.client_name,
        cargo_description=request.cargo_description,
        cargo_weight=request.cargo_weight,
        scheduled_date=request.scheduled_date,
        is_round_trip=request.is_round_trip,
        notes=request.notes,
        created_by=current_user["id"]
    )
    
    sql, values = db_pg.build_insert("trips", TRIP_COLS, _modelo_a_fila(trip.model_dump()))
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    
    # Create coupling record if carreta assigned
    if request.carreta_id:
        coupling = CouplingHistory(
            company_id=current_user["company_id"],
            tracto_id=request.tracto_id,
            carreta_id=request.carreta_id,
            trip_id=trip.id,
            created_by=current_user["id"]
        )
        coupling_doc = coupling.model_dump()
        coupling_doc["start_date"] = coupling_doc["start_date"].isoformat()
        sql, values = db_pg.build_insert("couplings", COUPLING_COLS, coupling_doc)
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)

    # Notificar (push + notification) al chofer asignado
    if request.driver_id:
        await notify_users(
            current_user["company_id"],
            "Nuevo viaje asignado",
            f"Se te asignó el viaje {trip_number}"
            + (f" - {request.client_name}" if request.client_name else ""),
            "info",
            user_id=request.driver_id,
            entity_type="trip",
            entity_id=trip.id,
        )

    return {"id": trip.id, "message": "Viaje creado"}

@api_router.put("/trips/{trip_id}")
async def update_trip(trip_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    request.pop("id", None)
    request.pop("company_id", None)

    # updated_at lo pone _actualizar_viaje; devuelve False si el viaje no
    # existe, que es lo que antes decia matched_count == 0.
    if not await _actualizar_viaje(current_user["company_id"], trip_id, request):
        raise HTTPException(status_code=404, detail="Viaje no encontrado")
    
    return {"message": "Viaje actualizado"}

@api_router.delete("/trips/{trip_id}")
async def delete_trip(trip_id: str, current_user: dict = Depends(require_roles("owner", "admin"))):
    
    # Check if trip is in progress
    trip = await _viaje_pg(current_user["company_id"], trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")
    
    if trip.get("status") == "en_curso":
        raise HTTPException(status_code=400, detail="No se puede eliminar un viaje en curso")
    
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "delete from trips where id = $1 and company_id = $2",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        )
    
    return {"message": "Viaje eliminado"}

@api_router.post("/trips/{trip_id}/start")
async def start_trip(trip_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    trip = await _viaje_pg(current_user["company_id"], trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")
    
    if trip["status"] not in ["programado", "checklist_pendiente"]:
        raise HTTPException(status_code=400, detail="El viaje no está en estado programado")
    
    # Validate trip can start (checklist, blocks, etc)
    validation = await validate_trip_can_start(current_user["company_id"], trip_id, trip)
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail="; ".join(validation["errors"]))

    # Kilómetro de inicio REAL. Debe ser >= odómetro del tracto (o >= 0).
    vehicle = await _vehiculo_pg(current_user["company_id"], trip["tracto_id"])
    veh_odo = (vehicle.get("odometer", 0) or 0) if vehicle else 0
    km_start = request.get("km_start")
    if km_start is None:
        km_start = veh_odo
    try:
        km_start = int(km_start)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="km_start inválido")
    if km_start < 0:
        raise HTTPException(status_code=400, detail="El kilometraje inicial no puede ser negativo")
    if veh_odo > 0 and km_start < veh_odo:
        raise HTTPException(
            status_code=400,
            detail=f"El kilometraje inicial ({km_start}) no puede ser menor al odómetro del vehículo ({veh_odo})"
        )

    await _actualizar_viaje(current_user["company_id"], trip_id, {
        "status": TripStatus.EN_CURSO.value,
        "start_date": datetime.now(timezone.utc),
        "km_start": km_start,
    })
    
    # Update vehicle status
    cid = current_user["company_id"]
    await _actualizar_vehiculo(cid, trip["tracto_id"], {"status": VehicleStatus.EN_VIAJE.value})
    if trip.get("carreta_id"):
        await _actualizar_vehiculo(
            cid, trip["carreta_id"], {"status": VehicleStatus.EN_VIAJE.value}
        )
    
    # Audit log
    await create_audit_log(
        current_user["company_id"],
        current_user["id"],
        current_user["name"],
        "start_trip",
        "trip",
        trip_id,
        {"km_start": km_start}
    )

    return {"message": "Viaje iniciado"}

@api_router.post("/trips/{trip_id}/complete")
async def complete_trip(trip_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    trip = await _viaje_pg(current_user["company_id"], trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")

    # Kilómetro final REAL. Validar km_end >= km_start.
    km_start = trip.get("km_start") or 0
    km_end = request.get("km_end")
    trip_set = {
        "status": TripStatus.COMPLETADO.value,
        "end_date": datetime.now(timezone.utc).isoformat(),
    }
    if km_end is not None:
        try:
            km_end = int(km_end)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="km_end inválido")
        if km_end < km_start:
            raise HTTPException(
                status_code=400,
                detail=f"El kilometraje final ({km_end}) no puede ser menor al inicial ({km_start})"
            )
        trip_set["km_end"] = km_end

    await _actualizar_viaje(current_user["company_id"], trip_id, trip_set)

    # Update vehicle status
    cid = current_user["company_id"]
    await _actualizar_vehiculo(cid, trip["tracto_id"], {"status": VehicleStatus.DISPONIBLE.value})
    if trip.get("carreta_id"):
        await _actualizar_vehiculo(
            cid, trip["carreta_id"], {"status": VehicleStatus.DISPONIBLE.value}
        )

    # CLAVE: actualizar odómetro del TRACTO y de la CARRETA acoplada con el km final.
    # Alimenta llantas y mantenimiento + dispara checks (alertas/notificaciones).
    if km_end is not None:
        await apply_odometer_update(current_user["company_id"], trip["tracto_id"], km_end, current_user["id"])
        if trip.get("carreta_id"):
            await apply_odometer_update(current_user["company_id"], trip["carreta_id"], km_end, current_user["id"])

    # Close coupling
    if trip.get("carreta_id"):
        # Cierra el enganche abierto de este viaje (el que no tiene fin).
        async with db_pg.tx(current_user) as conn:
            await conn.execute(
                "update couplings set end_date = now() "
                "where trip_id = $1 and company_id = $2 and end_date is null",
                db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
            )

    return {"message": "Viaje completado"}

@api_router.get("/trips/{trip_id}/viatico-status")
async def get_trip_viatico_status(trip_id: str, current_user: dict = Depends(get_current_user)):
    """Estado de viáticos del viaje: presupuesto, gastado, remanente y bandera de alerta."""
    trip = await _viaje_pg(current_user["company_id"], trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")
    config = await _company_config(current_user["company_id"])
    per_trip = config.get("viatico_por_viaje", DEFAULT_VIATICO_POR_VIAJE)
    budget = trip.get("viatico_budget")
    if budget is None:
        budget = per_trip
    spent = trip.get("total_expenses", 0) or 0
    remaining = budget - spent
    return {
        "budget": budget,
        "spent": spent,
        "remaining": remaining,
        "per_trip": per_trip,
        "alert": remaining < per_trip,
    }

# ============== TABLAS EN POSTGRES: CIERRE DEL VIAJE ==============
# trip_advances, trip_expenses y settlements cortaron con la migracion 008.
# Van juntas y con trips porque el detalle y el total del viaje tienen que
# escribirse en la misma transaccion: mientras el detalle vivio en Mongo y el
# total en Postgres, un fallo a mitad dejaba el viaje con un total que no
# cuadraba con sus gastos.

TRIP_ADVANCE_COLS = {
    "id": "uuid", "company_id": "uuid", "trip_id": "uuid",
    "amount": "float", "payment_method": "text",
    "delivered_date": "ts", "delivered_by": "uuid",
    "notes": "text", "created_at": "ts",
}

TRIP_EXPENSE_COLS = {
    "id": "uuid", "company_id": "uuid", "trip_id": "uuid",
    "category": "enum:expense_category", "description": "text",
    "amount": "float", "provider": "text", "ruc": "text",
    "has_igv": "bool", "receipt_url": "text",
    "expense_date": "ts", "created_at": "ts", "created_by": "uuid",
}

SETTLEMENT_COLS = {
    "id": "uuid", "company_id": "uuid", "trip_id": "uuid",
    "total_advances": "float", "total_expenses": "float",
    "deductions": "float", "deduction_notes": "text",
    "balance": "float", "balance_type": "enum:balance_type",
    "status": "enum:settlement_status",
    "reviewed_by": "uuid", "reviewed_at": "ts",
    "closed_by": "uuid", "closed_at": "ts",
    "notes": "text", "created_at": "ts", "updated_at": "ts",
}

# Las categorias validas se derivan del enum y no se repiten a mano, para que
# agregar una sola valga para el modelo y para esta validacion a la vez.
_CATEGORIAS_GASTO = {c.value for c in ExpenseCategory}


async def _sumar_del_viaje(conn, tabla, trip_id, company_id) -> float:
    """Suma los importes de una tabla de detalle para un viaje.

    Suma en la base en vez de traerse las filas y sumarlas en Python, que es lo
    que hacia la version Mongo. El filtro por empresa va explicito: la consulta
    vieja buscaba solo por trip_id, sin tenant.
    """
    total = await conn.fetchval(
        "select coalesce(sum(amount), 0) from " + tabla
        + " where trip_id = $1 and company_id = $2",
        db_pg.as_uuid(trip_id), db_pg.as_uuid(company_id),
    )
    return float(total or 0)


# ============== TRIP ADVANCE/EXPENSE ROUTES ==============
@api_router.get("/trips/{trip_id}/advances")
async def get_trip_advances(trip_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        advances = db_pg.rows_to_api(await conn.fetch(
            "select * from trip_advances where trip_id = $1 and company_id = $2 "
            "order by delivered_date desc nulls last limit 100",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    return advances

@api_router.post("/trips/{trip_id}/advances")
async def create_trip_advance(trip_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))):
    
    advance = TripAdvance(
        company_id=current_user["company_id"],
        trip_id=trip_id,
        amount=request["amount"],
        payment_method=request.get("payment_method", "efectivo"),
        delivered_by=current_user["id"],
        notes=request.get("notes")
    )
    
    # El anticipo y el total del viaje se escriben en la MISMA transaccion: o
    # entran los dos o ninguno. Mientras el detalle vivio en Mongo y el total en
    # Postgres esto no se podia garantizar, y un fallo a mitad dejaba el viaje
    # con un total que no cuadraba con sus anticipos.
    #
    # La suma va en la propia base (total + $1) y no leyendo-y-reescribiendo
    # desde Python, que abriria una carrera entre dos anticipos simultaneos.
    sql, values = db_pg.build_insert(
        "trip_advances", TRIP_ADVANCE_COLS, _modelo_a_fila(advance.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
        await conn.execute(
            "update trips set total_advance = total_advance + $1, updated_at = now() "
            "where id = $2 and company_id = $3",
            _to_float(request["amount"]),
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        )
    
    return {"id": advance.id, "message": "Anticipo registrado"}

@api_router.get("/trips/{trip_id}/expenses")
async def get_trip_expenses(trip_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        expenses = db_pg.rows_to_api(await conn.fetch(
            "select * from trip_expenses where trip_id = $1 and company_id = $2 "
            "order by expense_date desc nulls last limit 500",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    return expenses

@api_router.post("/trips/{trip_id}/expenses")
async def create_trip_expense(trip_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    # category va a una columna enum: una que no exista hace fallar el insert
    # con un 500 opaco. Se valida antes para responder 400 diciendo cual es.
    if request.get("category") not in _CATEGORIAS_GASTO:
        raise HTTPException(
            status_code=400,
            detail="Categoría de gasto inválida: %r. Válidas: %s"
                   % (request.get("category"), ", ".join(sorted(_CATEGORIAS_GASTO))),
        )
    expense = TripExpense(
        company_id=current_user["company_id"],
        trip_id=trip_id,
        category=request["category"],
        description=request.get("description"),
        amount=request["amount"],
        provider=request.get("provider"),
        ruc=request.get("ruc"),
        has_igv=request.get("has_igv", False),
        receipt_url=request.get("receipt_url"),
        created_by=current_user["id"]
    )
    
    # Gasto y total del viaje, en la misma transaccion (ver create_trip_advance).
    sql, values = db_pg.build_insert(
        "trip_expenses", TRIP_EXPENSE_COLS, _modelo_a_fila(expense.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
        await conn.execute(
            "update trips set total_expenses = total_expenses + $1, updated_at = now() "
            "where id = $2 and company_id = $3",
            _to_float(request["amount"]),
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        )

    # Alerta de viáticos bajos si corresponde
    await check_viatico_alert(current_user["company_id"], trip_id)

    return {"id": expense.id, "message": "Gasto registrado"}

# ============== CHECKLIST ROUTES ==============
# ============== TABLAS EN POSTGRES: CHECKLIST PRE-VIAJE ==============
# checklist_templates, checklists y checklist_runs cortaron con la migracion
# 010.
#
# Hay DOS modelos para la misma cosa -Checklist y ChecklistRun, cada uno con
# sus endpoints- y trips.checklist_id recibe ids de los dos. Por eso esa
# columna se queda sin FK: es polimorfica de facto. La deuda es anterior a la
# migracion; unificarlos toca el frontend y va aparte (ver la 010).


def _fila_con_ubicacion(doc: dict) -> dict:
    """El dict location {lat, lng} -> las columnas location_lat y location_lng.

    Cuatro tablas guardan la ubicacion en dos columnas separadas (checklists,
    checklist_runs, fuel_loads e issues), mientras que los modelos la llevan
    como un unico dict, que es la forma que manda y espera el frontend.

    Sin esta traduccion el dict no coincide con ninguna columna declarada y
    build_insert lo descarta por su lista blanca: la ubicacion se perderia sin
    ningun error, que es la peor forma de perder un dato.
    """
    fila = dict(doc)
    ubicacion = fila.pop("location", None)
    if isinstance(ubicacion, dict):
        fila["location_lat"] = ubicacion.get("lat")
        fila["location_lng"] = ubicacion.get("lng")
    return fila


def _api_con_ubicacion(fila: dict) -> dict:
    """Lo inverso: las dos columnas vuelven a ser un dict location {lat, lng},
    que es la forma que el frontend ya sabe leer."""
    if not fila:
        return fila
    salida = dict(fila)
    lat = salida.pop("location_lat", None)
    lng = salida.pop("location_lng", None)
    salida["location"] = (
        {"lat": lat, "lng": lng} if lat is not None or lng is not None else None
    )
    return salida


CHECKLIST_TEMPLATE_COLS = {
    "id": "uuid", "company_id": "uuid", "name": "text",
    "vehicle_type": "text", "items": "json", "is_active": "bool",
    "created_at": "ts", "created_by": "uuid",
}

CHECKLIST_COLS = {
    "id": "uuid", "company_id": "uuid", "trip_id": "uuid",
    "vehicle_id": "uuid", "driver_id": "uuid",
    "items": "json", "tire_checks": "json",
    "result": "enum:checklist_result", "signature_url": "text",
    "location_lat": "float", "location_lng": "float",
    "completed_at": "ts", "created_at": "ts",
}

CHECKLIST_RUN_COLS = {
    "id": "uuid", "company_id": "uuid", "template_id": "uuid",
    "trip_id": "uuid", "tracto_id": "uuid", "carreta_id": "uuid",
    "driver_id": "uuid", "responses": "json", "tire_checks": "json",
    "result": "enum:checklist_result", "signature_url": "text",
    "location_lat": "float", "location_lng": "float",
    "started_at": "ts", "completed_at": "ts",
    "photos": "text[]", "created_by": "uuid",
}


@api_router.get("/checklists")
async def get_checklists(
    trip_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(trip_id, "trip_id = $?", db_pg.as_uuid(trip_id))
    async with db_pg.tx(current_user) as conn:
        filas = db_pg.rows_to_api(await conn.fetch(
            "select * from checklists where " + f.where
            + " order by created_at desc nulls last limit 500", *f.values
        ))
    return [_api_con_ubicacion(c) for c in filas]

@api_router.post("/checklists")
async def create_checklist(request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    checklist = Checklist(
        company_id=current_user["company_id"],
        trip_id=request["trip_id"],
        vehicle_id=request["vehicle_id"],
        driver_id=current_user["id"],
        items=request.get("items", []),
        tire_checks=request.get("tire_checks", []),
        result=request.get("result", "pending"),
        location=request.get("location")
    )
    
    sql, values = db_pg.build_insert(
        "checklists", CHECKLIST_COLS,
        _fila_con_ubicacion(_modelo_a_fila(checklist.model_dump())),
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    
    # If result is OK, approve checklist on trip
    if request.get("result") == "ok":
        await _actualizar_viaje(
            current_user["company_id"], request["trip_id"],
            {"checklist_id": checklist.id, "checklist_approved": True},
        )
    
    return {"id": checklist.id, "message": "Checklist creado"}

# ============== TABLAS EN POSTGRES: COMBUSTIBLE ==============
# fuel_vouchers y fuel_loads cortaron con la migracion 011. Van juntas porque
# una carga descuenta del vale que la autorizo.
#
# fuel_loads guarda la ubicacion en location_lat/location_lng, igual que los
# checklists: se traduce con los mismos helpers (_fila_con_ubicacion y
# _api_con_ubicacion).

FUEL_VOUCHER_COLS = {
    "id": "uuid", "company_id": "uuid", "vehicle_id": "uuid", "trip_id": "uuid",
    "voucher_number": "text", "provider": "text",
    "limit_amount": "float", "limit_liters": "float",
    "valid_from": "ts", "valid_until": "ts", "is_used": "bool",
    "approved_by": "uuid", "voucher_photo_url": "text",
    "invoice_photo_url": "text", "invoice_number": "text",
    "created_at": "ts",
}

FUEL_LOAD_COLS = {
    "id": "uuid", "company_id": "uuid", "vehicle_id": "uuid",
    "voucher_id": "uuid", "trip_id": "uuid",
    "voucher_number": "text", "invoice_number": "text",
    "liters": "float", "price_per_liter": "float", "total_amount": "float",
    "odometer": "int", "provider": "text", "receipt_url": "text",
    "voucher_photo_url": "text", "invoice_photo_url": "text",
    "location_lat": "float", "location_lng": "float",
    "load_date": "ts", "created_at": "ts", "created_by": "uuid",
}


# ============== FUEL ROUTES ==============
@api_router.get("/fuel/vouchers")
async def get_fuel_vouchers(
    vehicle_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(vehicle_id, "vehicle_id = $?", db_pg.as_uuid(vehicle_id))
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from fuel_vouchers where " + f.where
            + " order by valid_from desc nulls last limit 500", *f.values
        ))

@api_router.post("/fuel/vouchers")
async def create_fuel_voucher(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "operaciones"))):
    
    voucher = FuelVoucher(
        company_id=current_user["company_id"],
        vehicle_id=request["vehicle_id"],
        trip_id=request.get("trip_id"),
        voucher_number=request["voucher_number"],
        provider=request["provider"],
        limit_amount=request.get("limit_amount"),
        limit_liters=request.get("limit_liters"),
        valid_from=datetime.fromisoformat(request["valid_from"]) if isinstance(request["valid_from"], str) else request["valid_from"],
        valid_until=datetime.fromisoformat(request["valid_until"]) if isinstance(request["valid_until"], str) else request["valid_until"],
        approved_by=current_user["id"]
    )
    
    sql, values = db_pg.build_insert(
        "fuel_vouchers", FUEL_VOUCHER_COLS, _modelo_a_fila(voucher.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": voucher.id, "message": "Vale creado"}

@api_router.get("/fuel/loads")
async def get_fuel_loads(
    vehicle_id: Optional[str] = None,
    trip_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(vehicle_id, "vehicle_id = $?", db_pg.as_uuid(vehicle_id))
    f.si(trip_id, "trip_id = $?", db_pg.as_uuid(trip_id))
    async with db_pg.tx(current_user) as conn:
        filas = db_pg.rows_to_api(await conn.fetch(
            "select * from fuel_loads where " + f.where
            + " order by load_date desc nulls last limit 1000", *f.values
        ))
    return [_api_con_ubicacion(l) for l in filas]

@api_router.post("/fuel/loads")
async def create_fuel_load(request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    load = FuelLoad(
        company_id=current_user["company_id"],
        vehicle_id=request["vehicle_id"],
        voucher_id=request.get("voucher_id"),
        trip_id=request.get("trip_id"),
        liters=request["liters"],
        price_per_liter=request["price_per_liter"],
        total_amount=request["liters"] * request["price_per_liter"],
        odometer=request["odometer"],
        provider=request["provider"],
        location=request.get("location"),
        created_by=current_user["id"]
    )
    
    # La carga y el vale que consume se escriben en la misma transaccion: un
    # vale no puede quedar marcado como usado por una carga que no entro, ni
    # al reves.
    sql, values = db_pg.build_insert(
        "fuel_loads", FUEL_LOAD_COLS,
        _fila_con_ubicacion(_modelo_a_fila(load.model_dump())),
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)

        # Mark voucher as used if provided
        if request.get("voucher_id"):
            await conn.execute(
                "update fuel_vouchers set is_used = true "
                "where id = $1 and company_id = $2",
                db_pg.as_uuid(request["voucher_id"]),
                db_pg.as_uuid(current_user["company_id"]),
            )

    # Si la carga está ligada a un viaje, el diésel DESCUENTA del saldo de viáticos:
    # se registra como gasto categoría "combustible" y se incrementa total_expenses.
    trip_id = request.get("trip_id")
    if trip_id:
        fuel_expense = TripExpense(
            company_id=current_user["company_id"],
            trip_id=trip_id,
            category="combustible",
            description=f"Combustible {load.liters} gl - {load.provider}",
            amount=load.total_amount,
            provider=load.provider,
            ruc=request.get("ruc"),
            receipt_url=request.get("receipt_url") or request.get("invoice_photo_url"),
            created_by=current_user["id"]
        )
        sql, values = db_pg.build_insert(
            "trip_expenses", TRIP_EXPENSE_COLS,
            _modelo_a_fila(fuel_expense.model_dump()),
        )
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)
            await conn.execute(
                "update trips set total_expenses = total_expenses + $1, updated_at = now() "
                "where id = $2 and company_id = $3",
                _to_float(load.total_amount),
                db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
            )
        await check_viatico_alert(current_user["company_id"], trip_id)

    # Actualizar odómetro del vehículo (max) y disparar checks de mantenimiento/llanta
    await apply_odometer_update(
        current_user["company_id"], request["vehicle_id"], request["odometer"], current_user["id"]
    )

    return {"id": load.id, "message": "Cargue registrado"}

@api_router.put("/fuel/vouchers/{voucher_id}")
async def update_fuel_voucher(voucher_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    """Update a fuel voucher - Admin only"""
    async with db_pg.tx(current_user) as conn:
        voucher = db_pg.to_api(await conn.fetchrow(
            "select * from fuel_vouchers where id = $1 and company_id = $2",
            db_pg.as_uuid(voucher_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    if not voucher:
        raise HTTPException(status_code=404, detail="Vale no encontrado")
    
    update_data = {}
    allowed_fields = ["voucher_number", "provider", "limit_amount", "limit_liters", "valid_from", "valid_until", "is_used", "photo_url"]
    for field in allowed_fields:
        if field in request:
            if field in ["valid_from", "valid_until"] and isinstance(request[field], str):
                update_data[field] = request[field]
            else:
                update_data[field] = request[field]
    
    # "photo_url" esta en allowed_fields pero no es una columna (las que hay son
    # voucher_photo_url e invoice_photo_url). FUEL_VOUCHER_COLS actua de lista
    # blanca y lo descarta, igual que hacia falta en los cortes anteriores.
    update_data["id"] = voucher_id
    update_data["company_id"] = current_user["company_id"]
    sql, values = db_pg.build_update(
        "fuel_vouchers", FUEL_VOUCHER_COLS, update_data, ["id", "company_id"]
    )
    if sql:
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)

    return {"message": "Vale actualizado"}

@api_router.delete("/fuel/vouchers/{voucher_id}")
async def delete_fuel_voucher(voucher_id: str, current_user: dict = Depends(require_roles("owner", "admin"))):
    """Delete a fuel voucher - Admin only"""
    async with db_pg.tx(current_user) as conn:
        borrado = await conn.fetchval(
            "delete from fuel_vouchers where id = $1 and company_id = $2 returning id",
            db_pg.as_uuid(voucher_id), db_pg.as_uuid(current_user["company_id"]),
        )
    if borrado is None:
        raise HTTPException(status_code=404, detail="Vale no encontrado")
    
    return {"message": "Vale eliminado"}

@api_router.put("/fuel/loads/{load_id}")
async def update_fuel_load(load_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    """Update a fuel load - Admin only"""
    async with db_pg.tx(current_user) as conn:
        load = db_pg.to_api(await conn.fetchrow(
            "select * from fuel_loads where id = $1 and company_id = $2",
            db_pg.as_uuid(load_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    if not load:
        raise HTTPException(status_code=404, detail="Carga no encontrada")
    
    update_data = {}
    allowed_fields = ["liters", "price_per_liter", "odometer", "provider", "receipt_url", "photo_url"]
    for field in allowed_fields:
        if field in request:
            update_data[field] = request[field]
    
    # Recalculate total if liters or price changed
    if "liters" in update_data or "price_per_liter" in update_data:
        liters = update_data.get("liters", load.get("liters", 0))
        price = update_data.get("price_per_liter", load.get("price_per_liter", 0))
        update_data["total_amount"] = liters * price
    
    update_data["id"] = load_id
    update_data["company_id"] = current_user["company_id"]
    sql, values = db_pg.build_update(
        "fuel_loads", FUEL_LOAD_COLS, update_data, ["id", "company_id"]
    )
    if sql:
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)

    return {"message": "Carga actualizada"}

@api_router.delete("/fuel/loads/{load_id}")
async def delete_fuel_load(load_id: str, current_user: dict = Depends(require_roles("owner", "admin"))):
    """Delete a fuel load - Admin only"""
    async with db_pg.tx(current_user) as conn:
        borrada = await conn.fetchval(
            "delete from fuel_loads where id = $1 and company_id = $2 returning id",
            db_pg.as_uuid(load_id), db_pg.as_uuid(current_user["company_id"]),
        )
    if borrada is None:
        raise HTTPException(status_code=404, detail="Carga no encontrada")
    
    return {"message": "Carga eliminada"}

# ============== TABLAS EN POSTGRES: LLANTAS Y MANTENIMIENTO ==============
# Corte 013, el ultimo. Diez tablas que forman un solo componente conectado:
# el ciclo work_orders <-> issues y todo lo que cuelga de el (ver la migracion
# 013 para por que no se podia partir en cortes mas chicos).

TIRE_COLS = {
    "id": "uuid", "company_id": "uuid", "serial": "text", "brand": "text",
    "model": "text", "dimension": "text", "position_type": "text",
    "purchase_cost": "float", "purchase_date": "ts", "supplier": "text",
    "status": "enum:tire_status", "life_number": "int",
    "initial_depth": "float", "last_depth": "float",
    "band_brand": "text", "band_model": "text",
    "scrap_reason": "text", "scrap_date": "ts", "scrap_odometer": "int",
    "current_vehicle_id": "uuid", "current_position": "text",
    "total_km": "int", "created_at": "ts", "updated_at": "ts",
}

TIRE_MOUNT_COLS = {
    "id": "uuid", "company_id": "uuid", "tire_id": "uuid", "vehicle_id": "uuid",
    "position_code": "text", "mount_date": "ts", "mount_odometer": "int",
    "unmount_date": "ts", "unmount_odometer": "int", "reason": "text",
    "created_by": "uuid",
}

TIRE_INSPECTION_COLS = {
    "id": "uuid", "company_id": "uuid", "tire_id": "uuid", "vehicle_id": "uuid",
    "position_code": "text", "depths": "float[]", "pressure": "float",
    "irregular_wear": "bool", "wear_type": "text", "photos": "text[]",
    "odometer": "int", "notes": "text", "inspection_date": "ts",
    "created_by": "uuid",
}

TIRE_LIFE_EVENT_COLS = {
    "id": "uuid", "company_id": "uuid", "tire_id": "uuid", "life_number": "int",
    "event_type": "text", "cost": "float", "supplier": "text", "notes": "text",
    "odometer": "int", "event_date": "ts", "created_by": "uuid",
}

TIRE_ROTATION_COLS = {
    "id": "uuid", "company_id": "uuid", "vehicle_id": "uuid", "changes": "json",
    "reason": "text", "odometer": "int", "rotation_date": "ts",
    "created_by": "uuid",
}

ALIGNMENT_RECORD_COLS = {
    "id": "uuid", "company_id": "uuid", "vehicle_id": "uuid", "axle": "text",
    "workshop": "text", "cost": "float", "notes": "text",
    "alignment_date": "ts", "created_by": "uuid",
}

MAINTENANCE_PLAN_COLS = {
    "id": "uuid", "company_id": "uuid", "name": "text",
    "vehicle_type": "enum:vehicle_type", "component": "text",
    "interval_km": "int", "interval_days": "int", "interval_hours": "int",
    "tasks": "text[]", "created_at": "ts",
}

WORK_ORDER_COLS = {
    "id": "uuid", "company_id": "uuid", "order_number": "text",
    "vehicle_id": "uuid", "order_type": "text",
    "priority": "enum:work_order_priority", "status": "enum:work_order_status",
    "description": "text", "maintenance_plan_id": "uuid", "issue_id": "uuid",
    "items": "json", "labor_cost": "float", "parts_cost": "float",
    "total_cost": "float", "workshop": "text", "technician": "text",
    "scheduled_date": "ts", "start_date": "ts", "end_date": "ts",
    "odometer_at_service": "int", "notes": "text",
    "created_at": "ts", "updated_at": "ts",
    "created_by": "uuid", "closed_by": "uuid",
}

DOWNTIME_RECORD_COLS = {
    "id": "uuid", "company_id": "uuid", "vehicle_id": "uuid",
    "work_order_id": "uuid", "reason": "text", "start_time": "ts",
    "end_time": "ts", "duration_hours": "float", "created_by": "uuid",
}

ISSUE_COLS = {
    "id": "uuid", "company_id": "uuid", "issue_number": "text",
    "trip_id": "uuid", "vehicle_id": "uuid", "driver_id": "uuid",
    "checklist_id": "uuid", "tire_id": "uuid",
    "issue_type": "enum:issue_type", "severity": "enum:issue_severity",
    "status": "enum:issue_status", "title": "text", "description": "text",
    "location_lat": "float", "location_lng": "float", "photos": "text[]",
    "cost": "float", "responsible": "text", "resolution": "text",
    "work_order_id": "uuid", "resolved_by": "uuid", "resolved_at": "ts",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}


async def _llanta_pg(company_id: str, tire_id: str):
    """Una llanta por id, o None. Se repite en casi todos los endpoints del
    modulo, siempre para validar antes de escribir."""
    async with db_pg.tx({"company_id": company_id}) as conn:
        return db_pg.to_api(await conn.fetchrow(
            "select * from tires where id = $1 and company_id = $2",
            db_pg.as_uuid(tire_id), db_pg.as_uuid(company_id),
        ))


async def _actualizar_llanta(company_id: str, tire_id: str, datos: dict):
    """UPDATE de tires por id + empresa. Devuelve si toco alguna fila.

    El filtro por empresa va explicito aunque RLS ya lo garantice: es la regla
    del modulo (ver db_pg.py). En Mongo varios de estos UPDATE se hacian solo
    por id, sin empresa -- lo que dejaba a una empresa modificar la llanta de
    otra si adivinaba el uuid.
    """
    fila = dict(datos)
    fila["id"] = tire_id
    fila["company_id"] = company_id
    fila.setdefault("updated_at", datetime.now(timezone.utc))
    sql, values = db_pg.build_update("tires", TIRE_COLS, fila, ["id", "company_id"])
    if not sql:
        return False
    async with db_pg.tx({"company_id": company_id}) as conn:
        return (await conn.execute(sql, *values)) != "UPDATE 0"


async def _orden_pg(company_id: str, order_id: str):
    """Una orden de trabajo por id, o None."""
    async with db_pg.tx({"company_id": company_id}) as conn:
        return db_pg.to_api(await conn.fetchrow(
            "select * from work_orders where id = $1 and company_id = $2",
            db_pg.as_uuid(order_id), db_pg.as_uuid(company_id),
        ))


async def _ultima_inspeccion_pg(company_id: str, tire_id: str):
    """La inspeccion mas reciente de una llanta, o None."""
    async with db_pg.tx({"company_id": company_id}) as conn:
        return db_pg.to_api(await conn.fetchrow(
            "select * from tire_inspections where tire_id = $1 and company_id = $2 "
            "order by inspection_date desc limit 1",
            db_pg.as_uuid(tire_id), db_pg.as_uuid(company_id),
        ))


# ============== TIRE ROUTES ==============
@api_router.get("/tires")
async def get_tires(
    vehicle_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(vehicle_id, "current_vehicle_id = $?", db_pg.as_uuid(vehicle_id))
    # status::text y no $?::tire_status: el valor llega de la query string, y
    # uno que no exista en el enum debe devolver lista vacia, no un 500.
    f.si(status, "status::text = $?", status)
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from tires where " + f.where + " order by serial limit 1000",
            *f.values
        ))

@api_router.get("/tires/{tire_id}")
async def get_tire(tire_id: str, current_user: dict = Depends(get_current_user)):
    tire = await _llanta_pg(current_user["company_id"], tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")
    return tire

@api_router.post("/tires")
async def create_tire(request: CreateTireRequest, current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))):
    
    tire = Tire(
        company_id=current_user["company_id"],
        serial=request.serial,
        brand=request.brand,
        model=request.model,
        dimension=request.dimension,
        position_type=request.position_type or "toda_posicion",
        purchase_cost=request.purchase_cost,
        purchase_date=request.purchase_date,
        supplier=request.supplier,
        initial_depth=request.initial_depth
    )
    
    sql, values = db_pg.build_insert(
        "tires", TIRE_COLS, _modelo_a_fila(tire.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": tire.id, "message": "Llanta creada"}

@api_router.put("/tires/{tire_id}")
async def update_tire(tire_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("superadmin", "owner", "admin", "mantenimiento", "flota"))):
    """Update tire details (admin/flota/mantenimiento)"""
    tire = await _llanta_pg(current_user["company_id"], tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")

    # Allowed editable fields
    allowed = ["serial", "brand", "model", "dimension", "position_type", "purchase_cost", "purchase_date",
               "supplier", "status", "life_number", "current_position", "total_km", "initial_depth"]
    update_data = {field: request[field] for field in allowed if field in request}
    if update_data:
        await _actualizar_llanta(current_user["company_id"], tire_id, update_data)
    return {"message": "Llanta actualizada"}

@api_router.delete("/tires/{tire_id}")
async def delete_tire(tire_id: str, current_user: dict = Depends(require_roles("superadmin", "owner", "admin", "flota"))):
    """Delete a tire (only if not mounted)"""
    tire = await _llanta_pg(current_user["company_id"], tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")
    if tire.get("current_vehicle_id"):
        raise HTTPException(status_code=400, detail="No se puede eliminar una llanta montada. Desmonte primero.")

    # El DELETE puede rebotar contra las FKs de su historial (montajes,
    # inspecciones, eventos de vida, incidentes). En Mongo se borraba siempre y
    # ese historial quedaba huerfano apuntando a una llanta inexistente.
    try:
        async with db_pg.tx(current_user) as conn:
            await conn.execute(
                "delete from tires where id = $1 and company_id = $2",
                db_pg.as_uuid(tire_id), db_pg.as_uuid(current_user["company_id"]),
            )
    except db_pg.ForeignKeyViolationError:
        raise HTTPException(
            status_code=400,
            detail="La llanta tiene historial registrado (montajes, inspecciones o incidentes) y no se puede eliminar. Dela de baja con /scrap.",
        )
    return {"message": "Llanta eliminada"}

# ============== TIRE / AXLE HELPERS ==============
def _axle_num_from_position(position_code: Optional[str]) -> Optional[int]:
    """Deriva el número de eje (1-based) desde un position_code tipo 'EJE2-IZQ-EXT'."""
    if not position_code:
        return None
    m = re.search(r'(\d+)', str(position_code))
    return int(m.group(1)) if m else None

def _axle_for_position(vehicle: Optional[dict], position_code: str) -> Optional[dict]:
    """Devuelve el dict del eje de axle_config que corresponde a un position_code, o None."""
    if not vehicle:
        return None
    axle_config = vehicle.get("axle_config")
    if not axle_config:
        return None
    num = _axle_num_from_position(position_code)
    if not num or num < 1 or num > len(axle_config):
        return None
    return axle_config[num - 1]

def _tire_axle_compatible(position_type: Optional[str], axle_type: Optional[str]) -> bool:
    """toda_posicion/mixto en cualquier eje; direccional solo en direccional; traccion solo en traccion."""
    pt = (position_type or "toda_posicion").lower()
    at = (axle_type or "").lower()
    if pt in ("toda_posicion", "mixto"):
        return True
    if pt == "direccional":
        return at == "direccional"
    if pt == "traccion":
        return at == "traccion"
    return True

@api_router.post("/tires/mount")
async def mount_tire(request: MountTireRequest, current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))):

    # Check if tire exists and is available
    tire = await _llanta_pg(current_user["company_id"], request.tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")
    if tire.get("current_vehicle_id"):
        raise HTTPException(status_code=400, detail="Llanta ya está montada en otro vehículo")

    # Check if position is available
    # El filtro por empresa es nuevo: en Mongo esta consulta no lo llevaba, asi
    # que una llanta de OTRA empresa montada en esa posicion bloqueaba el
    # montaje con un "Posición ya ocupada" imposible de entender.
    async with db_pg.tx(current_user) as conn:
        existing = await conn.fetchrow(
            "select id from tires where company_id = $1 and current_vehicle_id = $2 "
            "and current_position = $3",
            db_pg.as_uuid(current_user["company_id"]),
            db_pg.as_uuid(request.vehicle_id),
            request.position_code,
        )
    if existing:
        raise HTTPException(status_code=400, detail="Posición ya ocupada")

    # Validate tire/axle compatibility (only if the vehicle declares axle_config)
    vehicle = await _vehiculo_pg(current_user["company_id"], request.vehicle_id)
    axle = _axle_for_position(vehicle, request.position_code)
    if axle and axle.get("type"):
        if not _tire_axle_compatible(tire.get("position_type"), axle.get("type")):
            raise HTTPException(status_code=400, detail="Tipo de llanta incompatible con el eje")

    # Create mount record
    mount = TireMount(
        company_id=current_user["company_id"],
        tire_id=request.tire_id,
        vehicle_id=request.vehicle_id,
        position_code=request.position_code,
        mount_odometer=request.mount_odometer,
        created_by=current_user["id"]
    )
    
    # El registro de montaje y el estado de la llanta van en la MISMA
    # transaccion: una llanta marcada en uso sin su montaje (o al reves) deja
    # el historial mintiendo y no hay forma de reconstruirlo.
    sql, values = db_pg.build_insert(
        "tire_mounts", TIRE_MOUNT_COLS, _modelo_a_fila(mount.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
        await conn.execute(
            "update tires set current_vehicle_id = $1, current_position = $2, "
            "status = $3::tire_status, updated_at = now() "
            "where id = $4 and company_id = $5",
            db_pg.as_uuid(request.vehicle_id),
            request.position_code,
            TireStatus.EN_USO.value,
            db_pg.as_uuid(request.tire_id),
            db_pg.as_uuid(current_user["company_id"]),
        )

    return {"id": mount.id, "message": "Llanta montada"}

@api_router.post("/tires/{tire_id}/unmount")
async def unmount_tire(tire_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))):
    
    tire = await _llanta_pg(current_user["company_id"], tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")

    odometro = db_pg.as_int(request.get("odometer"), 0)
    async with db_pg.tx(current_user) as conn:
        # Cierra el montaje abierto y devuelve de una vez su odometro de
        # montaje. En Mongo eran dos consultas: se cerraba el montaje y despues
        # se releia "el mas reciente" para calcular los km -- que ya era el que
        # se acababa de cerrar, pero nada lo garantizaba.
        cerrado = await conn.fetchrow(
            "update tire_mounts set unmount_date = now(), unmount_odometer = $1, "
            "reason = $2 "
            "where id = (select id from tire_mounts "
            "            where tire_id = $3 and company_id = $4 and unmount_date is null "
            "            order by mount_date desc limit 1) "
            "returning mount_odometer",
            odometro,
            request.get("reason"),
            db_pg.as_uuid(tire_id),
            db_pg.as_uuid(current_user["company_id"]),
        )
        km_traveled = max(0, odometro - (cerrado["mount_odometer"] or 0)) if cerrado else 0

        await conn.execute(
            "update tires set current_vehicle_id = null, current_position = null, "
            "status = $1::tire_status, total_km = total_km + $2, updated_at = now() "
            "where id = $3 and company_id = $4",
            request.get("new_status") or TireStatus.NUEVO.value,
            km_traveled,
            db_pg.as_uuid(tire_id),
            db_pg.as_uuid(current_user["company_id"]),
        )

    return {"message": "Llanta desmontada", "km_traveled": km_traveled}

@api_router.get("/tires/vehicle/{vehicle_id}")
async def get_vehicle_tires(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    """Get all tires mounted on a vehicle with their positions"""
    async with db_pg.tx(current_user) as conn:
        tires = db_pg.rows_to_api(await conn.fetch(
            "select * from tires where company_id = $1 and current_vehicle_id = $2 "
            "order by current_position nulls last limit 20",
            db_pg.as_uuid(current_user["company_id"]), db_pg.as_uuid(vehicle_id),
        ))
        # La ultima inspeccion y el montaje vigente de CADA llanta, en dos
        # consultas en vez de dos por llanta. "distinct on (tire_id)" con el
        # order by correcto se queda con la primera fila de cada grupo, que es
        # exactamente el sort+limit 1 que Mongo hacia dentro del bucle.
        ids = [db_pg.as_uuid(t["id"]) for t in tires]
        ultimas = {
            str(r["tire_id"]): db_pg.to_api(r)
            for r in await conn.fetch(
                "select distinct on (tire_id) * from tire_inspections "
                "where company_id = $1 and tire_id = any($2::uuid[]) "
                "order by tire_id, inspection_date desc",
                db_pg.as_uuid(current_user["company_id"]), ids,
            )
        }
        montajes = {
            str(r["tire_id"]): db_pg.to_api(r)
            for r in await conn.fetch(
                "select distinct on (tire_id) * from tire_mounts "
                "where company_id = $1 and tire_id = any($2::uuid[]) "
                "and vehicle_id = $3 and unmount_date is null "
                "order by tire_id, mount_date desc",
                db_pg.as_uuid(current_user["company_id"]), ids,
                db_pg.as_uuid(vehicle_id),
            )
        }

    vehicle = await _vehiculo_pg(current_user["company_id"], vehicle_id)
    vehicle_odometer = vehicle.get("odometer") if vehicle else None
    config = await _company_config(current_user["company_id"])

    # Get latest inspection for each tire + computed fields
    result = []
    for tire in tires:
        tire_data = dict(tire)
        inspection = ultimas.get(tire["id"])
        tire_data["last_inspection"] = inspection

        # Active mount record (montada = current_vehicle_id set)
        mount = montajes.get(tire["id"])
        mount_odometer = mount.get("mount_odometer") if mount else None

        # km_recorridos
        km = None
        if vehicle_odometer is not None and mount_odometer is not None:
            diff = vehicle_odometer - mount_odometer
            km = diff if diff >= 0 else None
        tire_data["km_recorridos"] = km

        # cod_vida (VN, R1, R2...)
        life = tire.get("life_number", 1) or 1
        tire_data["cod_vida"] = "VN" if life <= 1 else f"R{life - 1}"

        # costo de compra
        costo = tire.get("purchase_cost")

        # cost_per_km
        cost_per_km = None
        if costo and km and km > 0:
            cost_per_km = round(costo / km, 4)
        tire_data["cost_per_km"] = cost_per_km

        # cost_per_mm (profundidad inicial - profundidad actual)
        cost_per_mm = None
        initial_depth = tire.get("initial_depth")
        if costo and initial_depth is not None and inspection and inspection.get("depths"):
            current_depth = min(inspection["depths"])
            worn = initial_depth - current_depth
            if worn > 0:
                cost_per_mm = round(costo / worn, 4)
        tire_data["cost_per_mm"] = cost_per_mm

        # Proyección de vida: km_remaining / needs_review por llanta
        projection = await compute_tire_projection(
            current_user["company_id"], tire, vehicle, config,
            km_recorridos=km, latest_inspection=inspection
        )
        tire_data["km_remaining"] = projection["km_remaining"]
        tire_data["needs_review"] = projection["needs_review"]
        tire_data["wear_rate_mm_per_km"] = projection["wear_rate_mm_per_km"]
        tire_data["estimated_change_date"] = projection["estimated_change_date"]

        result.append(tire_data)

    return result

@api_router.post("/tires/inspect")
async def create_tire_inspection(request: CreateInspectionRequest, current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota", "chofer"))):
    
    inspection = TireInspection(
        company_id=current_user["company_id"],
        tire_id=request.tire_id,
        vehicle_id=request.vehicle_id,
        position_code=request.position_code,
        depths=request.depths,
        pressure=request.pressure,
        irregular_wear=request.irregular_wear,
        wear_type=request.wear_type,
        odometer=request.odometer,
        notes=request.notes,
        created_by=current_user["id"]
    )
    
    sql, values = db_pg.build_insert(
        "tire_inspections", TIRE_INSPECTION_COLS,
        _modelo_a_fila(inspection.model_dump()),
    )

    # Check for alerts
    min_depth = min(request.depths) if request.depths else 0

    # La inspeccion y el last_depth que deja en la llanta van en la misma
    # transaccion: son el mismo hecho medido, y si se separan la proyeccion de
    # vida puede leer una profundidad que no corresponde a ninguna inspeccion.
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)

        # Esta inspección es la más reciente -> actualizar last_depth de la llanta
        if request.depths:
            await conn.execute(
                "update tires set last_depth = $1, updated_at = now() "
                "where id = $2 and company_id = $3",
                min_depth,
                db_pg.as_uuid(request.tire_id),
                db_pg.as_uuid(current_user["company_id"]),
            )

        # El serial va en el texto de las alertas de mas abajo. Si la llanta no
        # existe se usa un dict vacio: antes eso reventaba con AttributeError.
        tire = db_pg.to_api(await conn.fetchrow(
            "select * from tires where id = $1 and company_id = $2",
            db_pg.as_uuid(request.tire_id),
            db_pg.as_uuid(current_user["company_id"]),
        )) or {}

    # Umbrales desde configuración de empresa (defaults 3/5)
    config = await _company_config(current_user["company_id"])
    critical_depth = config.get("tire_critical_depth", DEFAULT_TIRE_CRITICAL_DEPTH)
    warning_depth = config.get("tire_warning_depth", DEFAULT_TIRE_WARNING_DEPTH)

    if min_depth < critical_depth:
        alert = Alert(
            company_id=current_user["company_id"],
            alert_type="tire_critical",
            entity_type="tire",
            entity_id=request.tire_id,
            message=f"Llanta {tire.get('serial', '')} con profundidad crítica: {min_depth}mm",
            severity="critical"
        )
        await _insertar_alerta(current_user["company_id"], alert)
    elif min_depth < warning_depth:
        alert = Alert(
            company_id=current_user["company_id"],
            alert_type="tire_warning",
            entity_type="tire",
            entity_id=request.tire_id,
            message=f"Llanta {tire.get('serial', '')} con profundidad baja: {min_depth}mm",
            severity="warning"
        )
        await _insertar_alerta(current_user["company_id"], alert)

    if request.irregular_wear:
        alert = Alert(
            company_id=current_user["company_id"],
            alert_type="tire_irregular_wear",
            entity_type="tire",
            entity_id=request.tire_id,
            message=f"Llanta {tire.get('serial', '')} con desgaste irregular. Se recomienda alineación.",
            severity="warning"
        )
        await _insertar_alerta(current_user["company_id"], alert)

    return {"id": inspection.id, "message": "Inspección registrada"}

@api_router.get("/tires/{tire_id}/inspections")
async def get_tire_inspections(tire_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from tire_inspections where tire_id = $1 and company_id = $2 "
            "order by inspection_date desc limit 100",
            db_pg.as_uuid(tire_id), db_pg.as_uuid(current_user["company_id"]),
        ))

@api_router.get("/tires/{tire_id}/projection")
async def get_tire_projection(tire_id: str, current_user: dict = Depends(get_current_user)):
    """Proyección de vida de la llanta: tasa de desgaste, km restantes y fecha estimada de cambio."""
    tire = await _llanta_pg(current_user["company_id"], tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")
    vehicle = None
    if tire.get("current_vehicle_id"):
        vehicle = await _vehiculo_pg(current_user["company_id"], tire["current_vehicle_id"])
    latest = await _ultima_inspeccion_pg(current_user["company_id"], tire_id)
    return await compute_tire_projection(
        current_user["company_id"], tire, vehicle, latest_inspection=latest
    )

# ============== MAINTENANCE ROUTES ==============
@api_router.get("/maintenance/plans")
async def get_maintenance_plans(current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from maintenance_plans where company_id = $1 "
            "order by name limit 100",
            db_pg.as_uuid(current_user["company_id"]),
        ))

@api_router.get("/vehicles/{vehicle_id}/maintenance-status")
async def get_vehicle_maintenance_status(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    """Faltan X km para el próximo servicio del vehículo (plan matricial o por tipo)."""
    vehicle = await _vehiculo_pg(current_user["company_id"], vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    return await compute_maintenance_status(current_user["company_id"], vehicle)

@api_router.post("/maintenance/plans")
async def create_maintenance_plan(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento"))):
    
    plan = MaintenancePlan(
        company_id=current_user["company_id"],
        name=request["name"],
        vehicle_type=request["vehicle_type"],
        component=request["component"],
        interval_km=request.get("interval_km"),
        interval_days=request.get("interval_days"),
        interval_hours=request.get("interval_hours"),
        tasks=request.get("tasks", [])
    )
    
    sql, values = db_pg.build_insert(
        "maintenance_plans", MAINTENANCE_PLAN_COLS, _modelo_a_fila(plan.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": plan.id, "message": "Plan creado"}

# ============== MATRIX MAINTENANCE PLANS (E MAX 540 style) ==============
@api_router.get("/maintenance/matrix-plans")
async def list_matrix_plans(current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        filas = db_pg.rows_to_api(await conn.fetch(
            "select * from maintenance_matrix_plans where company_id = $1 "
            "order by name limit 100",
            db_pg.as_uuid(current_user["company_id"]),
        ))
        return [await _plan_matriz_a_api(conn, p) for p in filas]

@api_router.get("/maintenance/matrix-plans/{plan_id}")
async def get_matrix_plan(plan_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        plan = db_pg.to_api(await conn.fetchrow(
            "select * from maintenance_matrix_plans where id = $1 and company_id = $2",
            db_pg.as_uuid(plan_id), db_pg.as_uuid(current_user["company_id"]),
        ))
        if not plan:
            raise HTTPException(status_code=404, detail="Plan no encontrado")
        return await _plan_matriz_a_api(conn, plan)

@api_router.post("/maintenance/matrix-plans")
async def create_matrix_plan(request: dict = Body(...), current_user: dict = Depends(require_roles("superadmin", "owner", "admin", "mantenimiento"))):

    plan = MaintenanceMatrixPlan(
        company_id=current_user["company_id"],
        name=request["name"],
        vehicle_model=request.get("vehicle_model"),
        applies_to_vehicle_ids=request.get("applies_to_vehicle_ids", []),
        intervals=request.get("intervals", []),
        sections=request.get("sections", []),
        notes=request.get("notes"),
        created_by=current_user.get("user_id") or current_user.get("id"),
    )
    sql, values = db_pg.build_insert(
        "maintenance_matrix_plans", MATRIX_PLAN_COLS, _modelo_a_fila(plan.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
        await _guardar_vehiculos_del_plan(
            conn, plan.id, current_user["company_id"], plan.applies_to_vehicle_ids
        )
    return {"id": plan.id, "message": "Plan creado"}

@api_router.put("/maintenance/matrix-plans/{plan_id}")
async def update_matrix_plan(plan_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("superadmin", "owner", "admin", "mantenimiento"))):

    allowed = ["name", "vehicle_model", "applies_to_vehicle_ids", "intervals", "sections", "notes"]
    update_data = {k: request[k] for k in allowed if k in request}
    update_data["updated_at"] = datetime.now(timezone.utc)
    update_data["id"] = plan_id
    update_data["company_id"] = current_user["company_id"]

    async with db_pg.tx(current_user) as conn:
        existe = await conn.fetchval(
            "select id from maintenance_matrix_plans where id = $1 and company_id = $2",
            db_pg.as_uuid(plan_id), db_pg.as_uuid(current_user["company_id"]),
        )
        if not existe:
            raise HTTPException(status_code=404, detail="Plan no encontrado")
        # applies_to_vehicle_ids no es columna: lo descarta la lista blanca y se
        # guarda aparte, en la tabla puente.
        sql, values = db_pg.build_update(
            "maintenance_matrix_plans", MATRIX_PLAN_COLS, update_data, ["id", "company_id"]
        )
        if sql:
            await conn.execute(sql, *values)
        if "applies_to_vehicle_ids" in request:
            await _guardar_vehiculos_del_plan(
                conn, plan_id, current_user["company_id"],
                request["applies_to_vehicle_ids"],
            )
    return {"message": "Plan actualizado"}

@api_router.delete("/maintenance/matrix-plans/{plan_id}")
async def delete_matrix_plan(plan_id: str, current_user: dict = Depends(require_roles("superadmin", "owner", "admin"))):
    # La tabla puente cae sola: su FK al plan es on delete cascade.
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "delete from maintenance_matrix_plans where id = $1 and company_id = $2",
            db_pg.as_uuid(plan_id), db_pg.as_uuid(current_user["company_id"]),
        )
    return {"message": "Plan eliminado"}

@api_router.post("/maintenance/matrix-plans/import-excel")
async def import_matrix_plan_excel(file: UploadFile = File(...), current_user: dict = Depends(require_roles("superadmin", "owner", "admin", "mantenimiento"))):
    """Import a maintenance plan from Excel file (E MAX 540 format)"""

    try:
        import pandas as pd
        import io
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), header=None)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer Excel: {str(e)}")

    # Parse intervals from row 7 (codes), row 8 (hours), row 9 (km)
    # Cols 5..17 typically hold intervals
    intervals = []
    try:
        codes_row = df.iloc[7]
        hours_row = df.iloc[8]
        km_row = df.iloc[9]
        for col in range(5, 18):
            code = codes_row[col]
            hrs = hours_row[col]
            km = km_row[col]
            if pd.notna(code) and pd.notna(hrs):
                intervals.append({
                    "code": str(code).strip(),
                    "hours": int(hrs) if pd.notna(hrs) else None,
                    "km": int(km) * 1000 if pd.notna(km) else None,
                })
    except Exception:
        pass

    # Parse sections and tasks (rows 10+)
    sections = []
    current_section = None
    section_pattern = re.compile(r"^[A-Z]$")

    for idx in range(10, len(df)):
        row = df.iloc[idx]
        code_or_n = row[1] if pd.notna(row[1]) else None
        desc = row[2] if pd.notna(row[2]) else None

        if code_or_n is None and desc is None:
            continue

        code_str = str(code_or_n).strip() if code_or_n is not None else ""

        # Section header (single uppercase letter A-H)
        if section_pattern.match(code_str):
            if current_section:
                sections.append(current_section)
            current_section = {
                "code": code_str,
                "name": str(desc).strip() if desc else "",
                "tasks": []
            }
            continue

        # Task row
        if current_section is not None and desc is not None:
            actions = {}
            for i, interval in enumerate(intervals):
                col_idx = 5 + i
                if col_idx < len(row):
                    val = row[col_idx]
                    if pd.notna(val) and str(val).strip():
                        actions[interval["code"] + "_" + str(i)] = str(val).strip()

            task = {
                "n": str(code_or_n).strip() if code_or_n else "",
                "description": str(desc).strip(),
                "component_type": str(row[3]).strip() if pd.notna(row[3]) else None,
                "quantity": float(row[4]) if pd.notna(row[4]) else None,
                "actions": actions,
            }
            current_section["tasks"].append(task)

    if current_section:
        sections.append(current_section)

    # Get plan name from filename
    plan_name = file.filename.replace(".xlsx", "").replace(".xls", "").replace("Plan de mantenimiento ", "").strip()

    plan = MaintenanceMatrixPlan(
        company_id=current_user["company_id"],
        name=plan_name,
        intervals=intervals,
        sections=sections,
        created_by=current_user.get("user_id") or current_user.get("id"),
    )
    sql, values = db_pg.build_insert(
        "maintenance_matrix_plans", MATRIX_PLAN_COLS, _modelo_a_fila(plan.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
        await _guardar_vehiculos_del_plan(
            conn, plan.id, current_user["company_id"], plan.applies_to_vehicle_ids
        )

    return {
        "id": plan.id,
        "name": plan_name,
        "intervals_count": len(intervals),
        "sections_count": len(sections),
        "tasks_count": sum(len(s["tasks"]) for s in sections),
        "message": "Plan importado exitosamente"
    }

@api_router.get("/maintenance/work-orders")
async def get_work_orders(
    vehicle_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(vehicle_id, "vehicle_id = $?", db_pg.as_uuid(vehicle_id))
    f.si(status, "status::text = $?", status)
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from work_orders where " + f.where
            + " order by created_at desc limit 500", *f.values
        ))

@api_router.post("/maintenance/work-orders")
async def create_work_order(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))):
    
    # Generate order number
    async with db_pg.tx(current_user) as conn:
        count = await conn.fetchval(
            "select count(*) from work_orders where company_id = $1",
            db_pg.as_uuid(current_user["company_id"]),
        )
    order_number = f"OT-{count + 1:05d}"
    
    order = WorkOrder(
        company_id=current_user["company_id"],
        vehicle_id=request["vehicle_id"],
        order_number=order_number,
        order_type=request["order_type"],
        priority=request.get("priority", "normal"),
        description=request["description"],
        items=request.get("items", []),
        workshop=request.get("workshop"),
        created_by=current_user["id"]
    )
    
    sql, values = db_pg.build_insert(
        "work_orders", WORK_ORDER_COLS, _modelo_a_fila(order.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    
    # If critical priority, create block
    if request.get("priority") == "critica":
        block = OperationalBlock(
            company_id=current_user["company_id"],
            entity_type="vehicle",
            entity_id=request["vehicle_id"],
            reason=f"OT Crítica: {request['description'][:50]}",
            block_type="bloquea_asignacion"
        )
        sql, values = db_pg.build_insert(
            "blocks", BLOCK_COLS, _modelo_a_fila(block.model_dump())
        )
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)
        
        # Update vehicle status
        await _actualizar_vehiculo(
            current_user["company_id"], request["vehicle_id"],
            {"status": VehicleStatus.EN_MANTENIMIENTO.value},
        )

    return {"id": order.id, "order_number": order_number, "message": "Orden de trabajo creada"}

@api_router.put("/maintenance/work-orders/{order_id}")
async def update_work_order(order_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    request.pop("id", None)
    request.pop("company_id", None)
    request["updated_at"] = datetime.now(timezone.utc)

    # If completing order, set vehicle to available
    if request.get("status") == "completada":
        order = await _orden_pg(current_user["company_id"], order_id)
        if order:
            await _actualizar_vehiculo(
                current_user["company_id"], order["vehicle_id"],
                {"status": VehicleStatus.DISPONIBLE.value},
            )
            # Resolve any blocks
            async with db_pg.tx(current_user) as conn:
                await conn.execute(
                    "update blocks set is_active = false, resolved_at = now(), "
                    "resolved_by = $1 where company_id = $2 and entity_id = $3 "
                    "and is_active",
                    db_pg.as_uuid(current_user["id"]),
                    db_pg.as_uuid(current_user["company_id"]),
                    db_pg.as_uuid(order["vehicle_id"]),
                )
    
    datos = dict(request)
    datos["id"] = order_id
    datos["company_id"] = current_user["company_id"]
    sql, values = db_pg.build_update(
        "work_orders", WORK_ORDER_COLS, datos, ["id", "company_id"]
    )
    async with db_pg.tx(current_user) as conn:
        # Sin campos que tocar igual hay que comprobar que la orden existe: el
        # endpoint respondia 404 y el frontend cuenta con eso.
        if sql is None:
            existe = await conn.fetchval(
                "select 1 from work_orders where id = $1 and company_id = $2",
                db_pg.as_uuid(order_id), db_pg.as_uuid(current_user["company_id"]),
            )
            if not existe:
                raise HTTPException(status_code=404, detail="Orden no encontrada")
        elif (await conn.execute(sql, *values)) == "UPDATE 0":
            raise HTTPException(status_code=404, detail="Orden no encontrada")

    return {"message": "Orden actualizada"}

@api_router.delete("/maintenance/work-orders/{order_id}")
async def delete_work_order(order_id: str, current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento"))):
    order = await _orden_pg(current_user["company_id"], order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    if order.get("status") != "abierta":
        raise HTTPException(status_code=400, detail="Solo se pueden eliminar órdenes abiertas")

    # Una OT abierta puede tener ya consumos de repuestos, indisponibilidad o
    # un incidente que la referencia. En Mongo el borrado se hacia igual y esas
    # filas quedaban apuntando a una orden inexistente.
    try:
        async with db_pg.tx(current_user) as conn:
            await conn.execute(
                "delete from work_orders where id = $1 and company_id = $2",
                db_pg.as_uuid(order_id), db_pg.as_uuid(current_user["company_id"]),
            )
    except db_pg.ForeignKeyViolationError:
        raise HTTPException(
            status_code=400,
            detail="La orden tiene movimientos asociados (consumos, indisponibilidad o incidentes) y no se puede eliminar. Cancélela en su lugar.",
        )

    return {"message": "Orden eliminada"}

# ============== ISSUE ROUTES ==============
@api_router.get("/issues")
async def get_issues(
    status: Optional[str] = None,
    issue_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    # ::text en los dos: los valores llegan de la query string y uno que no
    # exista en el enum debe devolver lista vacia, no un 500.
    f.si(status, "status::text = $?", status)
    f.si(issue_type, "issue_type::text = $?", issue_type)
    async with db_pg.tx(current_user) as conn:
        filas = db_pg.rows_to_api(await conn.fetch(
            "select * from issues where " + f.where
            + " order by created_at desc limit 500", *f.values
        ))
    return [_api_con_ubicacion(i) for i in filas]

@api_router.post("/issues")
async def create_issue(request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    # Chofer sólo puede registrar incidentes a su propio nombre
    driver_id = request.get("driver_id")
    if current_user["role"] == "chofer":
        driver_id = current_user["id"]

    # Número de incidente
    async with db_pg.tx(current_user) as conn:
        count = await conn.fetchval(
            "select count(*) from issues where company_id = $1",
            db_pg.as_uuid(current_user["company_id"]),
        )
    issue_number = f"INC-{count + 1:05d}"

    issue = Issue(
        company_id=current_user["company_id"],
        issue_number=issue_number,
        trip_id=request.get("trip_id"),
        vehicle_id=request.get("vehicle_id"),
        driver_id=driver_id,
        checklist_id=request.get("checklist_id"),
        tire_id=request.get("tire_id"),
        issue_type=request.get("issue_type", "otro"),
        severity=request.get("severity", "media"),
        title=request.get("title", ""),
        description=request["description"],
        location=request.get("location"),
        photos=request.get("photos", []),
        cost=request.get("cost", 0),
        responsible=request.get("responsible"),
        created_by=current_user["id"]
    )

    # La ubicacion viaja como {lat, lng} y en la tabla son dos columnas: sin
    # _fila_con_ubicacion la lista blanca de build_insert la descartaria en
    # silencio (misma traduccion que checklists y fuel_loads).
    sql, values = db_pg.build_insert(
        "issues", ISSUE_COLS,
        _fila_con_ubicacion(_modelo_a_fila(issue.model_dump())),
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)

    # Notificar al admin de incidentes críticos/altos
    if issue.severity in ("critica", "alta"):
        await notify_users(
            current_user["company_id"],
            "Incidente reportado",
            f"[{issue.severity.upper()}] {issue.title or issue.issue_type}: {issue.description[:120]}",
            "alert",
            target_role="admin",
            entity_type="issue",
            entity_id=issue.id,
        )

    return {"id": issue.id, "message": "Incidente registrado"}

@api_router.put("/issues/{issue_id}")
async def update_issue(issue_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    request.pop("id", None)
    request.pop("company_id", None)
    request["updated_at"] = datetime.now(timezone.utc)

    datos = _fila_con_ubicacion(request)
    datos["id"] = issue_id
    datos["company_id"] = current_user["company_id"]
    sql, values = db_pg.build_update("issues", ISSUE_COLS, datos, ["id", "company_id"])
    async with db_pg.tx(current_user) as conn:
        if sql is None:
            existe = await conn.fetchval(
                "select 1 from issues where id = $1 and company_id = $2",
                db_pg.as_uuid(issue_id), db_pg.as_uuid(current_user["company_id"]),
            )
            if not existe:
                raise HTTPException(status_code=404, detail="Incidente no encontrado")
        elif (await conn.execute(sql, *values)) == "UPDATE 0":
            raise HTTPException(status_code=404, detail="Incidente no encontrado")

    return {"message": "Incidente actualizado"}

# ============== CHECKLIST TEMPLATE ROUTES ==============
@api_router.get("/checklist-templates")
async def get_checklist_templates(
    vehicle_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.crudo("is_active")
    f.si(vehicle_type, "vehicle_type = $?", vehicle_type)
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from checklist_templates where " + f.where
            + " order by name limit 100", *f.values
        ))

@api_router.post("/checklist-templates")
async def create_checklist_template(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    
    template = ChecklistTemplate(
        company_id=current_user["company_id"],
        name=request["name"],
        vehicle_type=request.get("vehicle_type"),
        items=request.get("items", []),
        created_by=current_user["id"]
    )
    
    sql, values = db_pg.build_insert(
        "checklist_templates", CHECKLIST_TEMPLATE_COLS,
        _modelo_a_fila(template.model_dump()),
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": template.id, "message": "Plantilla creada"}

@api_router.put("/checklist-templates/{template_id}")
async def update_checklist_template(template_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    
    request.pop("id", None)
    request.pop("company_id", None)

    datos = dict(request)
    datos["id"] = template_id
    datos["company_id"] = current_user["company_id"]
    sql, values = db_pg.build_update(
        "checklist_templates", CHECKLIST_TEMPLATE_COLS, datos, ["id", "company_id"]
    )
    if sql:
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)
    return {"message": "Plantilla actualizada"}

# ============== CHECKLIST RUN ROUTES ==============
@api_router.get("/checklists/trip/{trip_id}")
async def get_checklist_by_trip(trip_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        fila = db_pg.to_api(await conn.fetchrow(
            "select * from checklist_runs where trip_id = $1 and company_id = $2 "
            "order by started_at desc nulls last limit 1",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    return _api_con_ubicacion(fila) if fila else None

@api_router.post("/checklists/start")
async def start_checklist(request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Start a new checklist for a trip"""
    trip_id = request["trip_id"]
    
    # Get trip
    trip = await _viaje_pg(current_user["company_id"], trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")
    
    # Check if checklist already exists
    # Con el filtro por empresa, que la consulta de Mongo no tenia.
    async with db_pg.tx(current_user) as conn:
        existing = db_pg.to_api(await conn.fetchrow(
            "select * from checklist_runs where trip_id = $1 and company_id = $2 "
            "order by started_at desc nulls last limit 1",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    if existing and existing.get("result") != "pending":
        raise HTTPException(status_code=400, detail="Ya existe un checklist completado para este viaje")
    
    # Get default template
    template_id = request.get("template_id")
    if not template_id:
        async with db_pg.tx(current_user) as conn:
            hallada = await conn.fetchval(
                "select id from checklist_templates "
                "where company_id = $1 and is_active limit 1",
                db_pg.as_uuid(current_user["company_id"]),
            )
        template_id = str(hallada) if hallada else None

    if not template_id:
        raise HTTPException(status_code=400, detail="No hay plantilla de checklist disponible")
    
    checklist = ChecklistRun(
        company_id=current_user["company_id"],
        template_id=template_id,
        trip_id=trip_id,
        tracto_id=trip["tracto_id"],
        carreta_id=trip.get("carreta_id"),
        driver_id=current_user["id"],
        location=request.get("location"),
        created_by=current_user["id"]
    )
    
    sql, values = db_pg.build_insert(
        "checklist_runs", CHECKLIST_RUN_COLS,
        _fila_con_ubicacion(_modelo_a_fila(checklist.model_dump())),
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)

    # Update trip
    await _actualizar_viaje(
        current_user["company_id"], trip_id,
        {"checklist_id": checklist.id, "status": "checklist_pendiente"},
    )
    
    return {"id": checklist.id, "message": "Checklist iniciado"}

@api_router.post("/checklists/{checklist_id}/submit")
async def submit_checklist(checklist_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Submit a completed checklist"""
    async with db_pg.tx(current_user) as conn:
        checklist = db_pg.to_api(await conn.fetchrow(
            "select * from checklist_runs where id = $1 and company_id = $2",
            db_pg.as_uuid(checklist_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist no encontrado")
    
    responses = request.get("responses", [])
    tire_checks = request.get("tire_checks", [])
    signature_url = request.get("signature_url")
    photos = request.get("photos", [])
    location = request.get("location")
    
    # Calculate result based on responses
    result = "ok"
    critical_items = []
    observed_items = []
    
    for resp in responses:
        if resp.get("severity") == "critico" and resp.get("value") == False:
            result = "critico"
            critical_items.append(resp.get("label", "Item"))
        elif resp.get("severity") == "observado" and resp.get("value") == False:
            if result != "critico":
                result = "observado"
            observed_items.append(resp.get("label", "Item"))
    
    # Check tire conditions
    for tire in tire_checks:
        if tire.get("condition") == "critico":
            result = "critico"
            critical_items.append(f"Llanta {tire.get('position', '')}")
        elif tire.get("condition") == "mal" and result != "critico":
            result = "observado"
            observed_items.append(f"Llanta {tire.get('position', '')}")
    
    # Update checklist
    datos = _fila_con_ubicacion({
        "responses": responses,
        "tire_checks": tire_checks,
        "signature_url": signature_url,
        "photos": photos,
        "location": location,
        "result": result,
        "completed_at": datetime.now(timezone.utc),
        "id": checklist_id,
        "company_id": current_user["company_id"],
    })
    sql, values = db_pg.build_update(
        "checklist_runs", CHECKLIST_RUN_COLS, datos, ["id", "company_id"]
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    
    # Update trip
    await _actualizar_viaje(current_user["company_id"], checklist["trip_id"], {
        "checklist_result": result,
        "status": "programado" if result != "critico" else "checklist_pendiente",
    })
    
    # Create issue if critical
    if result == "critico":
        async with db_pg.tx(current_user) as conn:
            issue_count = await conn.fetchval(
                "select count(*) from issues where company_id = $1",
                db_pg.as_uuid(current_user["company_id"]),
            )
        issue = Issue(
            company_id=current_user["company_id"],
            issue_number=f"ISS-{issue_count + 1:05d}",
            trip_id=checklist["trip_id"],
            vehicle_id=checklist["tracto_id"],
            driver_id=checklist["driver_id"],
            checklist_id=checklist_id,
            issue_type="checklist_critico",
            severity="alta",
            title="Checklist Pre-Viaje Crítico",
            description=f"Items críticos: {', '.join(critical_items)}",
            photos=photos,
            created_by=current_user["id"]
        )
        sql, values = db_pg.build_insert(
            "issues", ISSUE_COLS,
            _fila_con_ubicacion(_modelo_a_fila(issue.model_dump())),
        )
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)

    # Audit log
    await create_audit_log(
        current_user["company_id"],
        current_user["id"],
        current_user["name"],
        "submit_checklist",
        "checklist",
        checklist_id,
        {"result": result, "trip_id": checklist["trip_id"]}
    )
    
    return {
        "message": "Checklist enviado",
        "result": result,
        "critical_items": critical_items,
        "observed_items": observed_items
    }

@api_router.post("/trip/{trip_id}/checklist")
async def submit_trip_checklist(trip_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Submit a complete checklist for a trip (combines start and submit)"""
    # Get trip
    trip = await _viaje_pg(current_user["company_id"], trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")
    
    # Check if checklist already exists and is completed
    # Con el filtro por empresa, que la consulta de Mongo no tenia.
    async with db_pg.tx(current_user) as conn:
        existing = db_pg.to_api(await conn.fetchrow(
            "select * from checklist_runs where trip_id = $1 and company_id = $2 "
            "order by started_at desc nulls last limit 1",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    if existing and existing.get("result") not in [None, "pending"]:
        raise HTTPException(status_code=400, detail="Ya existe un checklist completado para este viaje")
    
    # Get or create checklist
    if existing:
        checklist_id = existing["id"]
    else:
        # Get default template
        async with db_pg.tx(current_user) as conn:
            hallada = await conn.fetchval(
                "select id from checklist_templates "
                "where company_id = $1 and is_active limit 1",
                db_pg.as_uuid(current_user["company_id"]),
            )
        if not hallada:
            # template_id es NOT NULL con FK a checklist_templates: el "default"
            # que se usaba antes no es un id real y aca reventaria el insert.
            raise HTTPException(
                status_code=400, detail="No hay plantilla de checklist disponible"
            )
        template_id = str(hallada)

        checklist = ChecklistRun(
            company_id=current_user["company_id"],
            template_id=template_id,
            trip_id=trip_id,
            tracto_id=trip["tracto_id"],
            carreta_id=trip.get("carreta_id"),
            driver_id=current_user["id"],
            location=request.get("location"),
            created_by=current_user["id"]
        )
        
        sql, values = db_pg.build_insert(
            "checklist_runs", CHECKLIST_RUN_COLS,
            _fila_con_ubicacion(_modelo_a_fila(checklist.model_dump())),
        )
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)
        checklist_id = checklist.id
    
    # Process responses
    responses = request.get("responses", [])
    tire_checks = request.get("tire_checks", [])
    result = request.get("result", "ok")
    
    # Count critical and observed items
    critical_items = [r for r in responses if r.get("is_critical") and r.get("status") == "critico"]
    observed_items = [r for r in responses if r.get("status") in ["observado", "critico"]]
    
    # Update checklist
    # km_start y notes NO son columnas de checklist_runs y se quedan fuera:
    # CHECKLIST_RUN_COLS actua de lista blanca. El km inicial no se pierde,
    # porque el bloque de abajo lo guarda en el viaje, que es donde se usa.
    datos = {
        "responses": responses,
        "tire_checks": tire_checks,
        "signature_url": request.get("signature_url"),
        "result": result,
        "completed_at": datetime.now(timezone.utc),
        "id": checklist_id,
        "company_id": current_user["company_id"],
    }
    sql, values = db_pg.build_update(
        "checklist_runs", CHECKLIST_RUN_COLS, datos, ["id", "company_id"]
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    
    # Update trip
    await _actualizar_viaje(current_user["company_id"], trip_id, {
        "checklist_id": checklist_id,
        "checklist_result": result,
        "km_start": request.get("km_start"),
        "status": "programado" if result != "critico" else "checklist_pendiente",
    })
    
    # If critical, create an issue
    if result == "critico":
        issue = Issue(
            company_id=current_user["company_id"],
            trip_id=trip_id,
            vehicle_id=trip["tracto_id"],
            driver_id=trip["driver_id"],
            checklist_id=checklist_id,
            issue_type="checklist_critico",
            severity="alta",
            description=f"Checklist crítico: {len(critical_items)} items críticos detectados",
            created_by=current_user["id"]
        )
        sql, values = db_pg.build_insert(
            "issues", ISSUE_COLS,
            _fila_con_ubicacion(_modelo_a_fila(issue.model_dump())),
        )
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)

    return {
        "message": "Checklist enviado",
        "result": result,
        "critical_items": len(critical_items),
        "observed_items": len(observed_items)
    }

# ============== SETTLEMENT ROUTES ==============
@api_router.get("/settlements")
async def get_settlements(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    # status::text y no un cast al enum: el valor lo elige el cliente, y uno que
    # no exista devuelve lista vacia -como hacia Mongo- en vez de un 500.
    f.si(status, "status::text = $?", status)
    async with db_pg.tx(current_user) as conn:
        settlements = db_pg.rows_to_api(await conn.fetch(
            "select * from settlements where " + f.where
            + " order by created_at desc nulls last limit 500",
            *f.values,
        ))
    return settlements

@api_router.get("/trips/{trip_id}/settlement")
async def get_trip_settlement(trip_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        return db_pg.to_api(await conn.fetchrow(
            "select * from settlements where trip_id = $1 and company_id = $2",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        ))

@api_router.post("/trips/{trip_id}/settlement")
async def create_or_update_settlement(trip_id: str, request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Create or update settlement for a trip"""
    trip = await _viaje_pg(current_user["company_id"], trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")
    
    cid = current_user["company_id"]
    deductions = _to_float(request.get("deductions", 0))

    # Todo el calculo y las dos escrituras van en UNA transaccion: los totales
    # que se guardan son los que habia al leerlos, y el settlement_id del viaje
    # no puede quedar apuntando a una liquidacion que no llego a insertarse.
    async with db_pg.tx(current_user) as conn:
        total_advances = await _sumar_del_viaje(conn, "trip_advances", trip_id, cid)
        total_expenses = await _sumar_del_viaje(conn, "trip_expenses", trip_id, cid)

        balance = total_advances - total_expenses - deductions
        balance_type = "favor_empresa" if balance >= 0 else "favor_chofer"

        existing = db_pg.to_api(await conn.fetchrow(
            "select * from settlements where trip_id = $1 and company_id = $2",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(cid),
        ))

        if existing:
            if existing.get("status") == "cerrado":
                raise HTTPException(status_code=400, detail="Liquidación ya cerrada")
            datos = {
                "total_advances": total_advances,
                "total_expenses": total_expenses,
                "deductions": deductions,
                "deduction_notes": request.get("deduction_notes"),
                "balance": abs(balance),
                "balance_type": balance_type,
                "notes": request.get("notes"),
                "updated_at": datetime.now(timezone.utc),
                "id": existing["id"],
                "company_id": cid,
            }
            sql, values = db_pg.build_update(
                "settlements", SETTLEMENT_COLS, datos, ["id", "company_id"]
            )
            await conn.execute(sql, *values)
            return {"id": existing["id"], "message": "Liquidación actualizada"}

        settlement = TripSettlement(
            company_id=cid,
            trip_id=trip_id,
            total_advances=total_advances,
            total_expenses=total_expenses,
            deductions=deductions,
            deduction_notes=request.get("deduction_notes"),
            balance=abs(balance),
            balance_type=balance_type,
            notes=request.get("notes")
        )
        sql, values = db_pg.build_insert(
            "settlements", SETTLEMENT_COLS, _modelo_a_fila(settlement.model_dump())
        )
        await conn.execute(sql, *values)
        await conn.execute(
            "update trips set settlement_id = $1, settlement_status = 'pendiente', "
            "updated_at = now() where id = $2 and company_id = $3",
            db_pg.as_uuid(settlement.id),
            db_pg.as_uuid(trip_id), db_pg.as_uuid(cid),
        )
        return {"id": settlement.id, "message": "Liquidación creada"}

@api_router.post("/settlements/{settlement_id}/close")
async def close_settlement(settlement_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))):
    """Close a settlement"""
    
    cid = current_user["company_id"]
    # Leer, cerrar y marcar el viaje, todo en la misma transaccion. El for
    # update sobre la liquidacion evita que dos cierres simultaneos pasen los
    # dos por el chequeo de "ya cerrada".
    async with db_pg.tx(current_user) as conn:
        settlement = db_pg.to_api(await conn.fetchrow(
            "select * from settlements where id = $1 and company_id = $2 for update",
            db_pg.as_uuid(settlement_id), db_pg.as_uuid(cid),
        ))
        if not settlement:
            raise HTTPException(status_code=404, detail="Liquidación no encontrada")

        if settlement.get("status") == "cerrado":
            raise HTTPException(status_code=400, detail="Liquidación ya cerrada")

        datos = {
            "status": "cerrado",
            "closed_by": current_user["id"],
            "closed_at": datetime.now(timezone.utc),
            "notes": request.get("notes", settlement.get("notes")),
            "updated_at": datetime.now(timezone.utc),
            "id": settlement_id,
            "company_id": cid,
        }
        sql, values = db_pg.build_update(
            "settlements", SETTLEMENT_COLS, datos, ["id", "company_id"]
        )
        await conn.execute(sql, *values)
        await conn.execute(
            "update trips set settlement_status = 'cerrado', updated_at = now() "
            "where id = $1 and company_id = $2",
            db_pg.as_uuid(settlement["trip_id"]), db_pg.as_uuid(cid),
        )
    
    # Audit log
    await create_audit_log(
        current_user["company_id"],
        current_user["id"],
        current_user["name"],
        "close_settlement",
        "settlement",
        settlement_id,
        {"trip_id": settlement["trip_id"], "balance": settlement.get("balance")}
    )
    
    return {"message": "Liquidación cerrada"}

# ============== INVENTORY ROUTES ==============
@api_router.get("/inventory/items")
async def get_inventory_items(
    category: Optional[str] = None,
    low_stock: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.crudo("is_active")
    f.si(category, "category = $?", category)
    # El filtro de stock bajo lo resuelve la base. Antes se traian los 1000
    # articulos al proceso para descartar casi todos en Python.
    if low_stock:
        f.crudo("current_stock <= min_stock")

    async with db_pg.tx(current_user) as conn:
        filas = await conn.fetch(
            "select * from inventory_items where " + f.where
            + " order by name limit 1000",
            *f.values,
        )
    return db_pg.rows_to_api(filas)

@api_router.post("/inventory/items")
async def create_inventory_item(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "almacen"))):
    
    item = InventoryItem(
        company_id=current_user["company_id"],
        code=request["code"],
        name=request["name"],
        description=request.get("description"),
        category=request["category"],
        unit=request.get("unit", "unidad"),
        min_stock=request.get("min_stock", 0),
        max_stock=request.get("max_stock"),
        unit_cost=request.get("unit_cost", 0),
        location=request.get("location")
    )
    
    sql, values = db_pg.build_insert(
        "inventory_items", INVENTORY_ITEM_COLS, _modelo_a_fila(item.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": item.id, "message": "Item creado"}

@api_router.post("/inventory/moves")
async def create_stock_move(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "almacen", "mantenimiento"))):
    """Create a stock movement (entry/exit/adjustment)"""
    
    item_id = request["item_id"]
    move_type = request["move_type"]
    quantity = request["quantity"]
    
    if move_type not in [t.value for t in StockMoveType]:
        raise HTTPException(status_code=400, detail="Tipo de movimiento invalido")

    # Leer el stock, validarlo, registrar el movimiento y actualizar el saldo
    # ocurren ahora en UNA transaccion, con FOR UPDATE sobre el articulo.
    #
    # Antes eran tres operaciones sueltas: dos salidas simultaneas del mismo
    # articulo leian el mismo stock, las dos pasaban la validacion y el saldo
    # terminaba en negativo. El bloqueo de fila hace que la segunda espere y
    # lea el stock ya actualizado.
    async with db_pg.tx(current_user) as conn:
        item = db_pg.to_api(await conn.fetchrow(
            "select * from inventory_items where id = $1 and company_id = $2 for update",
            db_pg.as_uuid(item_id), db_pg.as_uuid(current_user["company_id"]),
        ))
        if not item:
            raise HTTPException(status_code=404, detail="Item no encontrado")

        current_stock = item.get("current_stock", 0)

        if move_type in ["salida", "consumo_ot"]:
            if current_stock < quantity:
                raise HTTPException(status_code=400, detail="Stock insuficiente")
            new_stock = current_stock - quantity
        elif move_type == "entrada":
            new_stock = current_stock + quantity
        else:  # ajuste
            new_stock = quantity

        move = StockMove(
            company_id=current_user["company_id"],
            item_id=item_id,
            move_type=move_type,
            quantity=quantity,
            unit_cost=request.get("unit_cost", item.get("unit_cost", 0)),
            total_cost=quantity * request.get("unit_cost", item.get("unit_cost", 0)),
            reference_type=request.get("reference_type"),
            reference_id=request.get("reference_id"),
            work_order_id=request.get("work_order_id"),
            notes=request.get("notes"),
            created_by=current_user["id"]
        )
        sql, values = db_pg.build_insert(
            "stock_moves", STOCK_MOVE_COLS, _modelo_a_fila(move.model_dump())
        )
        await conn.execute(sql, *values)

        await conn.execute(
            "update inventory_items set current_stock = $1, updated_at = now() "
            "where id = $2 and company_id = $3",
            new_stock,
            db_pg.as_uuid(item_id),
            db_pg.as_uuid(current_user["company_id"]),
        )

    # La alerta queda FUERA de la transaccion y sigue en Mongo (alerts no ha
    # cortado). Si fallara, el movimiento ya esta registrado, que es el mismo
    # comportamiento de antes.
    # Check for low stock alert
    if new_stock <= item.get("min_stock", 0):
        alert = Alert(
            company_id=current_user["company_id"],
            alert_type="low_stock",
            entity_type="inventory",
            entity_id=item_id,
            message=f"Stock bajo: {item.get('name', '')} - {new_stock} {item.get('unit', 'unidades')}",
            severity="warning"
        )
        await _insertar_alerta(current_user["company_id"], alert)

    return {"id": move.id, "new_stock": new_stock, "message": "Movimiento registrado"}

@api_router.get("/inventory/kardex/{item_id}")
async def get_kardex(item_id: str, current_user: dict = Depends(get_current_user)):
    """Get stock movement history for an item"""
    async with db_pg.tx(current_user) as conn:
        filas = await conn.fetch(
            "select * from stock_moves where item_id = $1 and company_id = $2 "
            "order by move_date desc limit 500",
            db_pg.as_uuid(item_id), db_pg.as_uuid(current_user["company_id"]),
        )
    return db_pg.rows_to_api(filas)

# ============== TABLAS EN POSTGRES: INVENTARIO ==============
# Las cuatro tablas del modulo ya cortaron
# (db/migrations/005_corte_inventario.sql). alerts cruzo en el 012 y
# work_orders en el 013, asi que el consumo de repuestos de una OT y el cierre
# de la orden ya caben en una sola transaccion.

SUPPLIER_COLS = {
    "id": "uuid", "company_id": "uuid", "name": "text", "ruc": "text",
    "address": "text", "phone": "text", "email": "text",
    "contact_person": "text", "category": "text", "is_active": "bool",
    "created_at": "ts",
}

INVENTORY_ITEM_COLS = {
    "id": "uuid", "company_id": "uuid", "code": "text", "name": "text",
    "description": "text", "category": "text", "unit": "text",
    "min_stock": "int", "max_stock": "int", "current_stock": "int",
    "unit_cost": "float", "location": "text", "is_active": "bool",
    "created_at": "ts", "updated_at": "ts",
}

STOCK_MOVE_COLS = {
    "id": "uuid", "company_id": "uuid", "item_id": "uuid",
    "move_type": "enum:stock_move_type", "quantity": "int",
    "unit_cost": "float", "total_cost": "float",
    "reference_type": "text", "reference_id": "uuid",
    "work_order_id": "uuid", "notes": "text",
    "move_date": "ts", "created_by": "uuid",
}

PURCHASE_ORDER_COLS = {
    "id": "uuid", "company_id": "uuid", "order_number": "text",
    "supplier_id": "uuid", "status": "text", "items": "json",
    "subtotal": "float", "tax": "float", "total": "float", "notes": "text",
    "approved_by": "uuid", "approved_at": "ts",
    "received_by": "uuid", "received_at": "ts",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}


# ============== SUPPLIER ROUTES ==============
@api_router.get("/suppliers")
async def get_suppliers(current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        filas = await conn.fetch(
            "select * from suppliers where company_id = $1 and is_active "
            "order by name limit 500",
            db_pg.as_uuid(current_user["company_id"]),
        )
    return db_pg.rows_to_api(filas)

@api_router.post("/suppliers")
async def create_supplier(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "almacen"))):
    
    supplier = Supplier(
        company_id=current_user["company_id"],
        name=request["name"],
        ruc=request.get("ruc"),
        address=request.get("address"),
        phone=request.get("phone"),
        email=request.get("email"),
        contact_person=request.get("contact_person"),
        category=request.get("category")
    )
    
    sql, values = db_pg.build_insert(
        "suppliers", SUPPLIER_COLS, _modelo_a_fila(supplier.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": supplier.id, "message": "Proveedor creado"}

# ============== PURCHASE ORDER ROUTES ==============
@api_router.get("/purchase-orders")
async def get_purchase_orders(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(status, "status = $?", status)
    async with db_pg.tx(current_user) as conn:
        filas = await conn.fetch(
            "select * from purchase_orders where " + f.where
            + " order by created_at desc limit 500",
            *f.values,
        )
    return db_pg.rows_to_api(filas)

@api_router.post("/purchase-orders")
async def create_purchase_order(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "almacen"))):
    
    async with db_pg.tx(current_user) as conn:
        count = await conn.fetchval(
            "select count(*) from purchase_orders where company_id = $1",
            db_pg.as_uuid(current_user["company_id"]),
        )
    # Mismo correlativo y misma carrera que tenia con Mongo: dos altas
    # simultaneas pueden calcular el mismo numero. No se cambia aca para no
    # mezclar el cambio de base con un cambio de comportamiento.
    order_number = f"OC-{count + 1:05d}"
    
    items = request.get("items", [])
    subtotal = sum(i.get("quantity", 0) * i.get("unit_price", 0) for i in items)
    tax = subtotal * 0.18  # IGV Peru
    total = subtotal + tax
    
    order = PurchaseOrder(
        company_id=current_user["company_id"],
        order_number=order_number,
        supplier_id=request["supplier_id"],
        items=items,
        subtotal=subtotal,
        tax=tax,
        total=total,
        notes=request.get("notes"),
        created_by=current_user["id"]
    )
    
    sql, values = db_pg.build_insert(
        "purchase_orders", PURCHASE_ORDER_COLS, _modelo_a_fila(order.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
    return {"id": order.id, "order_number": order_number, "message": "Orden de compra creada"}

@api_router.post("/purchase-orders/{order_id}/receive")
async def receive_purchase_order(order_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "almacen"))):
    """Receive a purchase order and update inventory"""
    
    # Toda la recepcion en UNA transaccion. Antes, si fallaba a mitad de los
    # articulos, quedaban unos con stock sumado y otros no, y la orden seguia
    # figurando como no recibida: al reintentar se sumaba dos veces.
    async with db_pg.tx(current_user) as conn:
        order = db_pg.to_api(await conn.fetchrow(
            "select * from purchase_orders where id = $1 and company_id = $2 for update",
            db_pg.as_uuid(order_id), db_pg.as_uuid(current_user["company_id"]),
        ))
        if not order:
            raise HTTPException(status_code=404, detail="Orden no encontrada")

        if order.get("status") == "recibido":
            raise HTTPException(status_code=400, detail="Orden ya recibida")

        # Create stock entries for each item
        for item in order.get("items", []):
            if item.get("item_id"):
                move = StockMove(
                    company_id=current_user["company_id"],
                    item_id=item["item_id"],
                    move_type="entrada",
                    quantity=item.get("quantity", 0),
                    unit_cost=item.get("unit_price", 0),
                    total_cost=item.get("quantity", 0) * item.get("unit_price", 0),
                    reference_type="purchase_order",
                    reference_id=order_id,
                    notes=f"Recepción OC {order.get('order_number', '')}",
                    created_by=current_user["id"]
                )
                sql, values = db_pg.build_insert(
                    "stock_moves", STOCK_MOVE_COLS, _modelo_a_fila(move.model_dump())
                )
                await conn.execute(sql, *values)

                # El saldo se suma en la propia base: leerlo y reescribirlo
                # desde Python volveria a abrir la carrera que evita el
                # movimiento de stock normal.
                await conn.execute(
                    "update inventory_items set current_stock = current_stock + $1, "
                    "updated_at = now() where id = $2 and company_id = $3",
                    item.get("quantity", 0),
                    db_pg.as_uuid(item["item_id"]),
                    db_pg.as_uuid(current_user["company_id"]),
                )

        await conn.execute(
            "update purchase_orders set status = $1, received_by = $2, "
            "received_at = now(), updated_at = now() "
            "where id = $3 and company_id = $4",
            "recibido",
            db_pg.as_uuid(current_user["id"]),
            db_pg.as_uuid(order_id),
            db_pg.as_uuid(current_user["company_id"]),
        )

    return {"message": "Orden recibida e inventario actualizado"}

# ============== EXTENDED WORK ORDER ROUTES ==============
@api_router.post("/maintenance/work-orders/{order_id}/start")
async def start_work_order(order_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento"))):
    """Start a work order"""
    
    order = await _orden_pg(current_user["company_id"], order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    # Create downtime record
    downtime = DowntimeRecord(
        company_id=current_user["company_id"],
        vehicle_id=order["vehicle_id"],
        work_order_id=order_id,
        reason=order.get("description", "Mantenimiento"),
        created_by=current_user["id"]
    )
    sql, values = db_pg.build_insert(
        "downtime_records", DOWNTIME_RECORD_COLS, _modelo_a_fila(downtime.model_dump())
    )

    # La orden pasa a en_proceso y nace su registro de indisponibilidad en la
    # misma transaccion: si solo entrara uno de los dos, el vehiculo quedaria
    # parado sin nada que lo explique, o al reves.
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update work_orders set status = 'en_proceso', start_date = now(), "
            "technician = $1, odometer_at_service = $2, updated_at = now() "
            "where id = $3 and company_id = $4",
            request.get("technician"),
            db_pg.as_int(request.get("odometer")),
            db_pg.as_uuid(order_id), db_pg.as_uuid(current_user["company_id"]),
        )
        await conn.execute(sql, *values)
    
    # Update vehicle status
    await _actualizar_vehiculo(
        current_user["company_id"], order["vehicle_id"], {"status": "en_mantenimiento"}
    )
    
    return {"message": "Orden iniciada"}

@api_router.post("/maintenance/work-orders/{order_id}/complete")
async def complete_work_order(order_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento"))):
    """Complete a work order"""
    
    order = await _orden_pg(current_user["company_id"], order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    labor_cost = request.get("labor_cost", 0)
    parts_cost = request.get("parts_cost", 0)
    total_cost = labor_cost + parts_cost

    consumos = [
        i for i in request.get("consumed_items", [])
        if i.get("item_id") and i.get("quantity")
    ]

    # Cerrar la OT, descontar los repuestos y cerrar la indisponibilidad son
    # ahora UNA sola transaccion. Hasta el corte 013 no podia serlo: la orden
    # vivia en Mongo y el inventario en Postgres, asi que un fallo entre medio
    # dejaba stock descontado contra una orden que seguia abierta.
    cid = db_pg.as_uuid(current_user["company_id"])
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update work_orders set status = 'completada', end_date = now(), "
            "labor_cost = $1, parts_cost = $2, total_cost = $3, items = $4, "
            "notes = $5, closed_by = $6, updated_at = now() "
            "where id = $7 and company_id = $8",
            db_pg.as_float(labor_cost, 0.0),
            db_pg.as_float(parts_cost, 0.0),
            db_pg.as_float(total_cost, 0.0),
            request.get("items", order.get("items", [])),
            request.get("notes"),
            db_pg.as_uuid(current_user["id"]),
            db_pg.as_uuid(order_id), cid,
        )

        # Consume parts from inventory
        for item in consumos:
            move = StockMove(
                company_id=current_user["company_id"],
                item_id=item["item_id"],
                move_type="consumo_ot",
                quantity=item["quantity"],
                work_order_id=order_id,
                notes=f"Consumo OT {order.get('order_number', '')}",
                created_by=current_user["id"]
            )
            sql, values = db_pg.build_insert(
                "stock_moves", STOCK_MOVE_COLS, _modelo_a_fila(move.model_dump())
            )
            await conn.execute(sql, *values)

            await conn.execute(
                "update inventory_items set current_stock = current_stock - $1, "
                "updated_at = now() where id = $2 and company_id = $3",
                item["quantity"],
                db_pg.as_uuid(item["item_id"]),
                cid,
            )

        # Close downtime
        # duration_hours se calcula al cerrar en vez de quedarse en 0: la
        # columna existia desde el principio y nadie la llenaba.
        await conn.execute(
            "update downtime_records set end_time = now(), "
            "duration_hours = extract(epoch from (now() - start_time)) / 3600.0 "
            "where work_order_id = $1 and company_id = $2 and end_time is null",
            db_pg.as_uuid(order_id), cid,
        )
    
    # Update vehicle
    # last_maintenance_km y last_maintenance_date son columnas nuevas de la
    # migracion 006: sin ellas este guardado se perdia y check_maintenance_due
    # leia siempre 0.
    await _actualizar_vehiculo(
        current_user["company_id"], order["vehicle_id"],
        {
            "status": "disponible",
            "last_maintenance_km": request.get("odometer", 0),
            "last_maintenance_date": datetime.now(timezone.utc),
        },
    )
    
    # Resolve any blocks
    #
    # Esta consulta buscaba blocks por work_order_id, que NO es una columna de
    # blocks (la tienen stock_moves y downtime_records, no esta). En Mongo eso
    # no coincidia nunca: era un no-op silencioso y este cierre jamas resolvio
    # un bloqueo. En SQL habria sido un error por columna inexistente.
    #
    # Se corrige a lo que evidentemente queria hacer, que ademas es lo que ya
    # hace el otro camino de cierre en update_work_order: soltar los bloqueos
    # del vehiculo de la orden. Se agrega tambien el filtro por empresa, que la
    # consulta vieja no tenia.
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update blocks set is_active = false, resolved_at = now(), "
            "resolved_by = $1 where company_id = $2 and entity_id = $3 and is_active",
            db_pg.as_uuid(current_user["id"]),
            db_pg.as_uuid(current_user["company_id"]),
            db_pg.as_uuid(order["vehicle_id"]),
        )
    
    # Audit log
    await create_audit_log(
        current_user["company_id"],
        current_user["id"],
        current_user["name"],
        "complete_work_order",
        "work_order",
        order_id,
        {"total_cost": total_cost, "vehicle_id": order["vehicle_id"]}
    )
    
    return {"message": "Orden completada", "total_cost": total_cost}

# ============== TIRE EXTENDED ROUTES ==============
@api_router.post("/tires/rotate")
async def rotate_tires(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))):
    """Rotate tires on a vehicle"""
    
    vehicle_id = request["vehicle_id"]
    changes = request.get("changes", [])  # [{from_position, to_position, tire_id}]
    odometer = request.get("odometer", 0)
    
    # Create rotation record
    rotation = TireRotation(
        company_id=current_user["company_id"],
        vehicle_id=vehicle_id,
        changes=changes,
        reason=request.get("reason"),
        odometer=odometer,
        created_by=current_user["id"]
    )
    sql, values = db_pg.build_insert(
        "tire_rotations", TIRE_ROTATION_COLS, _modelo_a_fila(rotation.model_dump())
    )

    # Validar, mover y registrar, todo en UNA transaccion. En Mongo eran N
    # validaciones, N updates y un insert sueltos: si algo fallaba en el medio
    # quedaba media rotacion aplicada y ningun registro de que habia pasado.
    # El filtro por empresa en la validacion tambien es nuevo.
    async with db_pg.tx(current_user) as conn:
        # Validate all tires exist and are mounted
        for change in changes:
            montada = await conn.fetchrow(
                "select id from tires where id = $1 and company_id = $2 "
                "and current_vehicle_id = $3",
                db_pg.as_uuid(change["tire_id"]),
                db_pg.as_uuid(current_user["company_id"]),
                db_pg.as_uuid(vehicle_id),
            )
            if not montada:
                raise HTTPException(status_code=400, detail=f"Llanta {change['tire_id']} no está montada en este vehículo")

        # Perform rotation
        for change in changes:
            await conn.execute(
                "update tires set current_position = $1, updated_at = now() "
                "where id = $2 and company_id = $3",
                change["to_position"],
                db_pg.as_uuid(change["tire_id"]),
                db_pg.as_uuid(current_user["company_id"]),
            )

        await conn.execute(sql, *values)

    return {"id": rotation.id, "message": "Rotación realizada"}

@api_router.post("/tires/align")
async def record_alignment(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))):
    """Record an alignment service"""
    
    alignment = AlignmentRecord(
        company_id=current_user["company_id"],
        vehicle_id=request["vehicle_id"],
        axle=request["axle"],
        workshop=request.get("workshop"),
        cost=request.get("cost", 0),
        notes=request.get("notes"),
        created_by=current_user["id"]
    )
    sql, values = db_pg.build_insert(
        "alignment_records", ALIGNMENT_RECORD_COLS,
        _modelo_a_fila(alignment.model_dump()),
    )

    # Resolve any alignment alerts
    # resolved_at no es una columna de alerts (si la tienen blocks): se omite.
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
        await conn.execute(
            "update alerts set resolved = true where company_id = $1 "
            "and entity_id = $2 "
            "and alert_type = any(array['tire_irregular_wear','axle_misalignment']) "
            "and not resolved",
            db_pg.as_uuid(current_user["company_id"]),
            db_pg.as_uuid(request["vehicle_id"]),
        )
    
    return {"id": alignment.id, "message": "Alineación registrada"}

@api_router.get("/tires/reports/required")
async def get_tires_required_report(current_user: dict = Depends(get_current_user)):
    """Get report of tires that need replacement/retreading"""
    company_id = current_user["company_id"]
    
    # Get company config for thresholds
    company = await _empresa_pg(company_id)
    config = company.get("config", {}) if company else {}
    critical_depth = config.get("tire_critical_depth", 3)
    warning_depth = config.get("tire_warning_depth", 5)
    
    # Get all tires in use with their latest inspection
    async with db_pg.tx(current_user) as conn:
        tires = db_pg.rows_to_api(await conn.fetch(
            "select * from tires where company_id = $1 and status = 'en_uso' "
            "order by serial limit 1000",
            db_pg.as_uuid(company_id),
        ))
    
    replace_needed = []
    retread_needed = []
    
    for tire in tires:
        last_depth = tire.get("last_depth")
        if last_depth is not None:
            if last_depth <= critical_depth:
                if tire.get("life_number", 1) < 3:  # Can still retread
                    retread_needed.append(serialize_doc(tire))
                else:
                    replace_needed.append(serialize_doc(tire))
            elif last_depth <= warning_depth:
                retread_needed.append(serialize_doc(tire))
    
    # Group by dimension
    replace_by_dim = {}
    for t in replace_needed:
        dim = t.get("dimension", "Unknown")
        if dim not in replace_by_dim:
            replace_by_dim[dim] = []
        replace_by_dim[dim].append(t)
    
    retread_by_dim = {}
    for t in retread_needed:
        dim = t.get("dimension", "Unknown")
        if dim not in retread_by_dim:
            retread_by_dim[dim] = []
        retread_by_dim[dim].append(t)
    
    return {
        "replace_needed": replace_by_dim,
        "retread_needed": retread_by_dim,
        "total_replace": len(replace_needed),
        "total_retread": len(retread_needed)
    }

@api_router.get("/tires/{tire_id}/history")
async def get_tire_history(tire_id: str, current_user: dict = Depends(get_current_user)):
    """Get complete history of a tire"""
    tire = await _llanta_pg(current_user["company_id"], tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")

    # Las tres consultas de historial llevan ahora filtro por empresa, que en
    # Mongo no tenian: bastaba el uuid de una llanta ajena para leerlo entero.
    cid = db_pg.as_uuid(current_user["company_id"])
    tid = db_pg.as_uuid(tire_id)
    async with db_pg.tx(current_user) as conn:
        mounts = db_pg.rows_to_api(await conn.fetch(
            "select * from tire_mounts where tire_id = $1 and company_id = $2 "
            "order by mount_date desc limit 100", tid, cid,
        ))
        inspections = db_pg.rows_to_api(await conn.fetch(
            "select * from tire_inspections where tire_id = $1 and company_id = $2 "
            "order by inspection_date desc limit 100", tid, cid,
        ))
        life_events = db_pg.rows_to_api(await conn.fetch(
            "select * from tire_life_events where tire_id = $1 and company_id = $2 "
            "order by event_date desc limit 100", tid, cid,
        ))

    return {
        "tire": tire,
        "mounts": mounts,
        "inspections": inspections,
        "life_events": life_events
    }

@api_router.get("/tires/vehicle/{vehicle_id}/diagnostics")
async def get_vehicle_tire_diagnostics(
    vehicle_id: str,
    max_depth_diff: float = 1.5,
    current_user: dict = Depends(get_current_user)
):
    """Motor de diagnóstico por eje: diferencias de profundidad y desgaste irregular."""
    company_id = current_user["company_id"]
    vehicle = await _vehiculo_pg(company_id, vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")

    axle_config = vehicle.get("axle_config") or []
    async with db_pg.tx(current_user) as conn:
        tires = db_pg.rows_to_api(await conn.fetch(
            "select * from tires where company_id = $1 and current_vehicle_id = $2 "
            "limit 50",
            db_pg.as_uuid(company_id), db_pg.as_uuid(vehicle_id),
        ))
        # Igual que en /tires/vehicle/{id}: una sola consulta para la ultima
        # inspeccion de todas las llantas, en vez de una por llanta.
        ultimas = {
            str(r["tire_id"]): db_pg.to_api(r)
            for r in await conn.fetch(
                "select distinct on (tire_id) * from tire_inspections "
                "where company_id = $1 and tire_id = any($2::uuid[]) "
                "order by tire_id, inspection_date desc",
                db_pg.as_uuid(company_id),
                [db_pg.as_uuid(t["id"]) for t in tires],
            )
        }

    config = await _company_config(company_id)
    critical_depth = config.get("tire_critical_depth", DEFAULT_TIRE_CRITICAL_DEPTH)
    axle_groups: Dict[str, List[Dict[str, Any]]] = {}
    suggestions: List[Dict[str, Any]] = []

    for tire in tires:
        position = tire.get("current_position") or ""
        num = _axle_num_from_position(position)
        # Nombre del eje: usa axle_config si existe, si no "EJE{n}"
        if num and 1 <= num <= len(axle_config) and axle_config[num - 1].get("name"):
            axle_name = axle_config[num - 1]["name"]
        elif num:
            axle_name = f"EJE{num}"
        else:
            axle_name = "SIN_EJE"

        inspection = ultimas.get(tire["id"])
        depths = inspection.get("depths") if inspection else None
        min_depth = min(depths) if depths else None
        irregular = bool(inspection.get("irregular_wear")) if inspection else False

        axle_groups.setdefault(axle_name, []).append({
            "tire_id": tire["id"], "position": position,
            "min_depth": min_depth, "irregular": irregular
        })

        if irregular:
            suggestions.append({
                "tire_id": tire["id"], "position": position, "action": "revisar_alineacion",
                "description": f"Desgaste irregular detectado en posición {position}. Revisar alineación.",
                "severity": "warning"
            })
        if min_depth is not None and min_depth <= critical_depth:
            suggestions.append({
                "tire_id": tire["id"], "position": position, "action": "reemplazar",
                "description": f"Profundidad crítica ({min_depth}mm) en posición {position}.",
                "severity": "critical"
            })

    axle_issues: List[Dict[str, Any]] = []
    for axle_name, items in axle_groups.items():
        measured = [i for i in items if i["min_depth"] is not None]
        if len(measured) >= 2:
            depths_vals = [i["min_depth"] for i in measured]
            diff = max(depths_vals) - min(depths_vals)
            if diff > max_depth_diff:
                axle_issues.append({
                    "axle": axle_name, "type": "diferencia_profundidad",
                    "description": f"Diferencia de profundidad de {round(diff, 2)}mm en {axle_name} (umbral {max_depth_diff}mm)."
                })
                lowest = min(measured, key=lambda i: i["min_depth"])
                suggestions.append({
                    "tire_id": lowest["tire_id"], "position": lowest["position"], "action": "rotar",
                    "description": f"Rotar llanta en {lowest['position']} por desbalance de profundidad en {axle_name}.",
                    "severity": "critical" if lowest["min_depth"] <= critical_depth else "warning"
                })
        if any(i["irregular"] for i in items):
            axle_issues.append({
                "axle": axle_name, "type": "desalineado",
                "description": f"Desgaste irregular en {axle_name}. Posible desalineación."
            })

    return {"axle_issues": axle_issues, "suggestions": suggestions}

@api_router.put("/tires/inspections/{inspection_id}")
async def update_tire_inspection(
    inspection_id: str,
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))
):
    """Modifica una inspección existente y recalcula last_depth si es la más reciente."""
    company_id = current_user["company_id"]
    cid = db_pg.as_uuid(company_id)
    async with db_pg.tx(current_user) as conn:
        inspection = db_pg.to_api(await conn.fetchrow(
            "select * from tire_inspections where id = $1 and company_id = $2",
            db_pg.as_uuid(inspection_id), cid,
        ))
        if not inspection:
            raise HTTPException(status_code=404, detail="Inspección no encontrada")

        allowed = ["depths", "pressure", "irregular_wear", "wear_type", "notes"]
        update_data = {k: request[k] for k in allowed if k in request}
        if update_data:
            datos = dict(update_data)
            datos["id"] = inspection_id
            datos["company_id"] = company_id
            sql, values = db_pg.build_update(
                "tire_inspections", TIRE_INSPECTION_COLS, datos, ["id", "company_id"]
            )
            if sql:
                await conn.execute(sql, *values)

        # Recalcular last_depth de la llanta si esta inspección es la más reciente
        latest = await conn.fetchrow(
            "select id from tire_inspections where tire_id = $1 and company_id = $2 "
            "order by inspection_date desc limit 1",
            db_pg.as_uuid(inspection["tire_id"]), cid,
        )
        if latest and str(latest["id"]) == inspection_id:
            depths = update_data.get("depths", inspection.get("depths") or [])
            if depths:
                await conn.execute(
                    "update tires set last_depth = $1, updated_at = now() "
                    "where id = $2 and company_id = $3",
                    min(db_pg.as_float(d, 0.0) for d in depths),
                    db_pg.as_uuid(inspection["tire_id"]), cid,
                )

    return {"message": "Inspección actualizada"}

@api_router.post("/tires/{tire_id}/retread")
async def retread_tire(
    tire_id: str,
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))
):
    """Reencauche: incrementa life_number, reinicia baseline de profundidad, deja la llanta en almacén."""
    company_id = current_user["company_id"]
    tire = await _llanta_pg(company_id, tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")
    if tire.get("current_vehicle_id"):
        raise HTTPException(status_code=400, detail="Desmonte la llanta antes de reencaucharla")

    new_life = (tire.get("life_number", 1) or 1) + 1
    new_baseline = request.get("new_depth")  # profundidad inicial del reencauche (opcional)

    event = TireLifeEvent(
        company_id=company_id, tire_id=tire_id, life_number=new_life,
        event_type="reencauche", cost=request.get("cost", 0) or 0,
        supplier=request.get("band_brand"),
        notes=f"Reencauche R{new_life - 1} banda={request.get('band_brand')} {request.get('band_model') or ''}".strip(),
        created_by=current_user["id"]
    )
    doc = _modelo_a_fila(event.model_dump())
    if request.get("date"):
        doc["event_date"] = request["date"]
    doc["odometer"] = request.get("odometer")
    sql, values = db_pg.build_insert("tire_life_events", TIRE_LIFE_EVENT_COLS, doc)

    # La llanta y su evento de vida, juntos: si el evento no queda escrito, el
    # life_number sube igual y la vida R2 no tiene de donde reconstruirse.
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update tires set life_number = $1, initial_depth = $2, last_depth = $2, "
            "band_brand = $3, band_model = $4, status = $5::tire_status, "
            "updated_at = now() where id = $6 and company_id = $7",
            new_life,
            db_pg.as_float(new_baseline),
            request.get("band_brand"),
            request.get("band_model"),
            TireStatus.ALMACEN.value,
            db_pg.as_uuid(tire_id), db_pg.as_uuid(company_id),
        )
        await conn.execute(sql, *values)

    return {"id": event.id, "message": "Reencauche registrado", "life_number": new_life}

@api_router.post("/tires/{tire_id}/regroove")
async def regroove_tire(
    tire_id: str,
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))
):
    """Reesculturado/regrabado: registra el evento sin cambiar life_number."""
    company_id = current_user["company_id"]
    tire = await _llanta_pg(company_id, tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")

    event = TireLifeEvent(
        company_id=company_id, tire_id=tire_id,
        life_number=tire.get("life_number", 1) or 1,
        event_type="regroove", cost=request.get("cost", 0) or 0,
        notes=request.get("notes"),
        created_by=current_user["id"]
    )
    doc = _modelo_a_fila(event.model_dump())
    if request.get("date"):
        doc["event_date"] = request["date"]
    sql, values = db_pg.build_insert("tire_life_events", TIRE_LIFE_EVENT_COLS, doc)
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)

    return {"id": event.id, "message": "Reesculturado registrado"}

@api_router.post("/tires/{tire_id}/scrap")
async def scrap_tire(
    tire_id: str,
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "mantenimiento", "flota"))
):
    """Baja / fin de vida de la llanta."""
    company_id = current_user["company_id"]
    tire = await _llanta_pg(company_id, tire_id)
    if not tire:
        raise HTTPException(status_code=404, detail="Llanta no encontrada")

    event = TireLifeEvent(
        company_id=company_id, tire_id=tire_id,
        life_number=tire.get("life_number", 1) or 1,
        event_type="baja", notes=request.get("reason"),
        created_by=current_user["id"]
    )
    doc = _modelo_a_fila(event.model_dump())
    if request.get("date"):
        doc["event_date"] = request["date"]
    doc["odometer"] = request.get("odometer")
    sql, values = db_pg.build_insert("tire_life_events", TIRE_LIFE_EVENT_COLS, doc)

    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update tires set status = $1::tire_status, current_vehicle_id = null, "
            "current_position = null, scrap_reason = $2, "
            "scrap_date = coalesce($3, now()), scrap_odometer = $4, "
            "updated_at = now() where id = $5 and company_id = $6",
            TireStatus.BAJA.value,
            request.get("reason"),
            db_pg.as_ts(request.get("date")),
            db_pg.as_int(request.get("odometer")),
            db_pg.as_uuid(tire_id), db_pg.as_uuid(company_id),
        )
        await conn.execute(sql, *values)

    return {"message": "Llanta dada de baja"}

@api_router.get("/tires/reports/scrap-pile")
async def get_scrap_pile_report(current_user: dict = Depends(get_current_user)):
    """Listado de llantas dadas de baja con análisis por motivo y marca."""
    company_id = current_user["company_id"]
    async with db_pg.tx(current_user) as conn:
        tires = db_pg.rows_to_api(await conn.fetch(
            "select * from tires where company_id = $1 and status = $2::tire_status "
            "order by scrap_date desc nulls last limit 1000",
            db_pg.as_uuid(company_id), TireStatus.BAJA.value,
        ))

    by_reason: Dict[str, Dict[str, Any]] = {}
    by_brand: Dict[str, Dict[str, Any]] = {}

    for tire in tires:
        km = tire.get("total_km") or 0
        reason = tire.get("scrap_reason") or "sin_motivo"
        brand = tire.get("brand") or "Desconocida"

        r = by_reason.setdefault(reason, {"count": 0, "km_total": 0, "km_count": 0})
        r["count"] += 1
        if km > 0:
            r["km_total"] += km
            r["km_count"] += 1

        b = by_brand.setdefault(brand, {"count": 0, "km_total": 0, "km_count": 0})
        b["count"] += 1
        if km > 0:
            b["km_total"] += km
            b["km_count"] += 1

    def _finish(groups):
        out = {}
        for key, v in groups.items():
            out[key] = {
                "count": v["count"],
                "avg_km": round(v["km_total"] / v["km_count"], 1) if v["km_count"] else None
            }
        return out

    return {
        "tires": [serialize_doc(t) for t in tires],
        "total": len(tires),
        "by_reason": _finish(by_reason),
        "by_brand": _finish(by_brand)
    }

# ============== AUDIT LOG ROUTES ==============
@api_router.get("/audit-logs")
async def get_audit_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    action: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(require_roles("owner", "admin"))
):
    query = {"company_id": current_user["company_id"]}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    if action:
        query["action"] = action
    
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(entity_type, "entity_type = $?", entity_type)
    f.si(entity_id, "entity_id = $?", db_pg.as_uuid(entity_id))
    f.si(action, "action = $?", action)
    # El limite se interpola como entero, no como parametro: es un int de la
    # firma del endpoint, nunca texto del cliente.
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from audit_logs where " + f.where
            + " order by created_at desc nulls last limit " + str(int(limit)),
            *f.values
        ))

# ============== FUEL EXTENDED ROUTES ==============
@api_router.get("/fuel/conciliation")
async def get_fuel_conciliation(
    vehicle_id: Optional[str] = None,
    trip_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get fuel conciliation report (vouchers vs actual loads)"""
    query = {"company_id": current_user["company_id"]}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    if trip_id:
        query["trip_id"] = trip_id
    
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(vehicle_id, "vehicle_id = $?", db_pg.as_uuid(vehicle_id))
    f.si(trip_id, "trip_id = $?", db_pg.as_uuid(trip_id))
    async with db_pg.tx(current_user) as conn:
        vouchers = db_pg.rows_to_api(await conn.fetch(
            "select * from fuel_vouchers where " + f.where + " limit 500", *f.values
        ))
        loads = db_pg.rows_to_api(await conn.fetch(
            "select * from fuel_loads where " + f.where + " limit 1000", *f.values
        ))
    
    results = []
    for voucher in vouchers:
        voucher_loads = [l for l in loads if l.get("voucher_id") == voucher["id"]]
        total_loaded = sum(l.get("liters", 0) for l in voucher_loads)
        total_amount = sum(l.get("total_amount", 0) for l in voucher_loads)
        
        limit = voucher.get("limit_liters") or voucher.get("limit_amount")
        used = total_loaded if voucher.get("limit_liters") else total_amount
        
        results.append({
            "voucher": serialize_doc(voucher),
            "loads": [serialize_doc(l) for l in voucher_loads],
            "total_loaded_liters": total_loaded,
            "total_amount": total_amount,
            "limit": limit,
            "used_percentage": (used / limit * 100) if limit else 0,
            "over_limit": used > limit if limit else False
        })
    
    # Loads without voucher
    no_voucher_loads = [l for l in loads if not l.get("voucher_id")]
    
    return {
        "voucher_conciliation": results,
        "loads_without_voucher": [serialize_doc(l) for l in no_voucher_loads],
        "total_without_voucher": len(no_voucher_loads)
    }

@api_router.get("/fuel/kpis")
async def get_fuel_kpis(
    vehicle_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get fuel KPIs (km/gal, cost/km)"""
    company_id = current_user["company_id"]
    
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
    f.si(vehicle_id, "vehicle_id = $?", db_pg.as_uuid(vehicle_id))
    async with db_pg.tx(current_user) as conn:
        loads = db_pg.rows_to_api(await conn.fetch(
            "select * from fuel_loads where " + f.where
            + " order by load_date desc nulls last limit 1000", *f.values
        ))

    if not loads:
        return {"message": "No hay datos de combustible"}
    
    # Group by vehicle
    vehicles_data = {}
    for load in loads:
        vid = load.get("vehicle_id")
        if vid not in vehicles_data:
            vehicles_data[vid] = {"loads": [], "total_liters": 0, "total_amount": 0}
        vehicles_data[vid]["loads"].append(load)
        vehicles_data[vid]["total_liters"] += load.get("liters", 0)
        vehicles_data[vid]["total_amount"] += load.get("total_amount", 0)
    
    kpis = []
    for vid, data in vehicles_data.items():
        sorted_loads = sorted(data["loads"], key=lambda x: x.get("odometer", 0))
        if len(sorted_loads) >= 2:
            km_traveled = sorted_loads[-1].get("odometer", 0) - sorted_loads[0].get("odometer", 0)
            if km_traveled > 0 and data["total_liters"] > 0:
                km_per_liter = km_traveled / data["total_liters"]
                km_per_gallon = km_per_liter * 3.78541  # Convert to gallons
                cost_per_km = data["total_amount"] / km_traveled
                
                vehicle = await _vehiculo_pg(current_user["company_id"], vid)
                kpis.append({
                    "vehicle_id": vid,
                    "plate": vehicle.get("plate") if vehicle else "Unknown",
                    "km_traveled": km_traveled,
                    "total_liters": data["total_liters"],
                    "total_amount": data["total_amount"],
                    "km_per_gallon": round(km_per_gallon, 2),
                    "cost_per_km": round(cost_per_km, 2),
                    "loads_count": len(data["loads"])
                })
    
    # Sort by km/gal efficiency
    kpis.sort(key=lambda x: x.get("km_per_gallon", 0), reverse=True)
    
    return {"vehicle_kpis": kpis}

# ============== OCR ENDPOINT FOR FUEL VOUCHERS ==============
@api_router.post("/fuel/ocr")
async def extract_fuel_voucher_data(request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """
    Extract data from fuel voucher photo using AI vision
    Accepts base64 image data and returns extracted fields
    """
    try:
        from google import genai

        image_data = request.get("image_base64", "")
        if not image_data:
            raise HTTPException(status_code=400, detail="Se requiere imagen en base64")

        # Clean base64 data
        if "base64," in image_data:
            image_data = image_data.split("base64,")[1]

        api_key = os.environ.get("GOOGLE_API_KEY", "")
        if not api_key:
            raise HTTPException(status_code=500, detail="API key no configurada")

        client = genai.Client(api_key=api_key)

        system_message = """Eres un asistente especializado en extraer datos de vales y recibos de combustible peruanos.
Analiza la imagen del vale de combustible y extrae la siguiente información en formato JSON:
{
    "voucher_number": "número del vale o comprobante",
    "provider": "nombre del grifo/estación de servicio",
    "liters": número de litros (solo número),
    "price_per_liter": precio por litro (solo número),
    "total_amount": monto total (solo número),
    "date": "fecha en formato YYYY-MM-DD si es visible",
    "vehicle_plate": "placa del vehículo si aparece",
    "odometer": número del odómetro si aparece (solo número),
    "fuel_type": "tipo de combustible (diesel, gasolina, etc)"
}
Si algún campo no es legible o no aparece, devuelve null para ese campo.
Devuelve SOLO el JSON sin explicaciones adicionales."""

        import base64
        image_bytes = base64.b64decode(image_data)

        result = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[
                {"role": "user", "parts": [
                    {"text": system_message + "\n\nExtrae los datos del siguiente vale de combustible:"},
                    {"inline_data": {"mime_type": "image/jpeg", "data": image_data}}
                ]}
            ]
        )

        response = result.text
        
        # Try to parse JSON from response
        import json

        # Clean response - remove markdown code blocks if present
        cleaned_response = response.strip()
        if cleaned_response.startswith("```"):
            cleaned_response = re.sub(r'^```(?:json)?\n?', '', cleaned_response)
            cleaned_response = re.sub(r'\n?```$', '', cleaned_response)
        
        try:
            extracted_data = json.loads(cleaned_response)
        except json.JSONDecodeError:
            # Try to find JSON in response
            json_match = re.search(r'\{[^{}]*\}', cleaned_response, re.DOTALL)
            if json_match:
                extracted_data = json.loads(json_match.group())
            else:
                extracted_data = {"raw_response": response, "parse_error": True}
        
        return {
            "success": True,
            "extracted_data": extracted_data
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Módulo de IA no disponible")
    except Exception as e:
        logging.error(f"OCR error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al procesar imagen: {str(e)}")

# ============== DASHBOARD/REPORTS ROUTES ==============
@api_router.get("/dashboard/kpis")
async def get_dashboard_kpis(current_user: dict = Depends(get_current_user)):
    company_id = current_user["company_id"]
    
    # Count vehicles by status
    total_vehicles = await _contar_vehiculos(company_id)
    available_vehicles = await _contar_vehiculos(company_id, "disponible")
    in_trip_vehicles = await _contar_vehiculos(company_id, "en_viaje")
    in_maintenance = await _contar_vehiculos(company_id, "en_mantenimiento")
    
    # Count trips by status
    total_trips = await _contar_viajes(company_id)
    active_trips = await _contar_viajes(company_id, "en_curso")
    completed_trips = await _contar_viajes(company_id, "completado")
    
    # Count drivers
    total_drivers = await _contar_usuarios(company_id, "chofer")
    
    # Count active alerts
    async with db_pg.tx(current_user) as conn:
        active_alerts = await conn.fetchval(
            "select count(*) from alerts where company_id = $1 and not resolved",
            db_pg.as_uuid(company_id),
        )
        critical_alerts = await conn.fetchval(
            "select count(*) from alerts where company_id = $1 and not resolved "
            "and severity = 'critical'",
            db_pg.as_uuid(company_id),
        )
    
    # Count active blocks
    async with db_pg.tx(current_user) as conn:
        active_blocks = await conn.fetchval(
            "select count(*) from blocks where company_id = $1 and is_active",
            db_pg.as_uuid(company_id),
        )

        # Count expiring documents (next 30 days)
        expiring_docs = await conn.fetchval(
            "select count(*) from documents where company_id = $1 "
            "and expiry_date <= $2 "
            "and status = any(array['vigente','por_vencer']::document_status[])",
            db_pg.as_uuid(company_id),
            datetime.now(timezone.utc) + timedelta(days=30),
        )
    
    # Open work orders
    async with db_pg.tx({"company_id": company_id}) as conn:
        open_work_orders = await conn.fetchval(
            "select count(*) from work_orders where company_id = $1 "
            "and status in ('abierta', 'en_proceso')",
            db_pg.as_uuid(company_id),
        )
    
    return {
        "vehicles": {
            "total": total_vehicles,
            "available": available_vehicles,
            "in_trip": in_trip_vehicles,
            "in_maintenance": in_maintenance,
            "availability_rate": round((available_vehicles / total_vehicles * 100) if total_vehicles > 0 else 0, 1)
        },
        "trips": {
            "total": total_trips,
            "active": active_trips,
            "completed": completed_trips
        },
        "drivers": {
            "total": total_drivers
        },
        "alerts": {
            "total": active_alerts,
            "critical": critical_alerts
        },
        "blocks": {
            "active": active_blocks
        },
        "documents": {
            "expiring": expiring_docs
        },
        "maintenance": {
            "open_orders": open_work_orders
        }
    }

@api_router.get("/dashboard/recent-activity")
async def get_recent_activity(current_user: dict = Depends(get_current_user)):
    company_id = current_user["company_id"]
    
    # Get recent trips
    async with db_pg.tx(current_user) as conn:
        recent_trips = db_pg.rows_to_api(await conn.fetch(
            "select * from trips where company_id = $1 "
            "order by created_at desc nulls last limit 5",
            db_pg.as_uuid(company_id),
        ))
    
    # Get recent alerts
    async with db_pg.tx(current_user) as conn:
        recent_alerts = db_pg.rows_to_api(await conn.fetch(
            "select * from alerts where company_id = $1 "
            "order by created_at desc nulls last limit 5",
            db_pg.as_uuid(company_id),
        ))
    
    # Get recent work orders
    async with db_pg.tx({"company_id": company_id}) as conn:
        recent_orders = db_pg.rows_to_api(await conn.fetch(
            "select * from work_orders where company_id = $1 "
            "order by created_at desc limit 5",
            db_pg.as_uuid(company_id),
        ))
    
    return {
        "trips": [serialize_doc(t) for t in recent_trips],
        "alerts": [serialize_doc(a) for a in recent_alerts],
        "work_orders": [serialize_doc(o) for o in recent_orders]
    }

# ============== FILE UPLOAD ROUTE ==============
import boto3
from botocore.exceptions import ClientError
import base64

def get_s3_client():
    """Get S3 client if credentials are configured"""
    access_key = os.environ.get('AWS_ACCESS_KEY_ID', '')
    secret_key = os.environ.get('AWS_SECRET_ACCESS_KEY', '')
    region = os.environ.get('AWS_REGION', 'us-east-1')
    
    if access_key and secret_key:
        return boto3.client(
            's3',
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            region_name=region
        )
    return None

# ============== UPLOAD VALIDATION ==============
MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB
ALLOWED_UPLOAD_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf", ".webp", ".heic"}
ALLOWED_UPLOAD_CONTENT_TYPES = {
    "image/jpeg", "image/png", "image/webp", "image/heic", "image/heif", "application/pdf"
}
_ENTITY_TYPE_RE = re.compile(r"^[a-z_]+$")
_ENTITY_ID_RE = re.compile(r"^[0-9a-fA-F-]{36}$")


def validate_entity_path(entity_type: str, entity_id: str):
    """Sanitiza entity_type/entity_id contra path traversal (whitelist)."""
    if not entity_type or not _ENTITY_TYPE_RE.match(entity_type):
        raise HTTPException(status_code=400, detail="entity inválido")
    if not entity_id or not _ENTITY_ID_RE.match(entity_id):
        raise HTTPException(status_code=400, detail="entity inválido")


def safe_upload_filename(original_filename: str) -> str:
    """Devuelve un nombre de archivo seguro validando la extensión (whitelist)."""
    base = os.path.basename(original_filename or "")
    ext = Path(base).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Extensión de archivo no permitida")
    return f"{uuid.uuid4()}{ext}"

@api_router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    entity_type: str = Form(...),
    entity_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload file to S3 or local storage"""
    # Sanitizar entity_type/entity_id (path traversal) y filename
    validate_entity_path(entity_type, entity_id)
    filename = safe_upload_filename(file.filename)

    # Validar content-type
    if file.content_type not in ALLOWED_UPLOAD_CONTENT_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de contenido no permitido")

    file_key = f"{entity_type}/{entity_id}/{filename}"

    # Read file content y validar tamaño
    file_content = await file.read()
    if len(file_content) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande")

    # Try S3 first
    s3_client = get_s3_client()
    bucket_name = os.environ.get('S3_BUCKET_NAME', '')
    
    if s3_client and bucket_name:
        try:
            s3_client.put_object(
                Bucket=bucket_name,
                Key=file_key,
                Body=file_content,
                ContentType=file.content_type
            )
            # Generate presigned URL for access
            url = s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': bucket_name, 'Key': file_key},
                ExpiresIn=86400 * 7  # 7 days
            )
            return {"url": url, "filename": filename, "storage": "s3", "key": file_key}
        except ClientError as e:
            logging.error(f"S3 upload error: {e}")
    
    # Fall back to local storage
    entity_dir = UPLOAD_DIR / entity_type / entity_id
    entity_dir.mkdir(parents=True, exist_ok=True)
    file_path = entity_dir / filename
    
    with open(file_path, "wb") as buffer:
        buffer.write(file_content)
    
    relative_url = f"/uploads/{entity_type}/{entity_id}/{filename}"
    return {"url": relative_url, "filename": filename, "storage": "local"}

async def save_uploaded_content(
    file_content: bytes,
    entity_type: str,
    entity_id: str,
    content_type: str = "image/jpeg",
    ext: str = "jpg",
) -> dict:
    """Guarda bytes ya decodificados en S3 (si hay credenciales) o disco local.

    Función compartida: usada por /upload/base64 y por módulos nuevos
    (liquidacion_flete, whatsapp_bot) que necesitan guardar un archivo sin
    pasar por HTTP. entity_type/entity_id ya deben venir sanitizados por el
    caller (mismo whitelist de caracteres que usaba /upload/base64 antes).
    """
    if len(file_content) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande")

    filename = f"{uuid.uuid4()}.{ext}"
    file_key = f"{entity_type}/{entity_id}/{filename}"

    # Try S3 first
    s3_client = get_s3_client()
    bucket_name = os.environ.get('S3_BUCKET_NAME', '')

    if s3_client and bucket_name:
        try:
            s3_client.put_object(
                Bucket=bucket_name,
                Key=file_key,
                Body=file_content,
                ContentType=content_type
            )
            url = s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': bucket_name, 'Key': file_key},
                ExpiresIn=86400 * 7
            )
            return {"url": url, "filename": filename, "storage": "s3"}
        except ClientError as e:
            logging.error(f"S3 upload error: {e}")

    # Fall back to local storage
    entity_dir = UPLOAD_DIR / entity_type / entity_id
    entity_dir.mkdir(parents=True, exist_ok=True)
    file_path = entity_dir / filename

    with open(file_path, "wb") as buffer:
        buffer.write(file_content)

    relative_url = f"/uploads/{entity_type}/{entity_id}/{filename}"
    return {"url": relative_url, "filename": filename, "storage": "local"}


@api_router.post("/upload/base64")
async def upload_base64(request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Upload base64 encoded image (for camera captures)"""
    data = request.get("data", "")
    entity_type = request.get("entity_type", "general")
    entity_id = request.get("entity_id", "general")

    # Sanitizar segmentos de path contra traversal (whitelist de caracteres)
    entity_type = re.sub(r"[^0-9a-zA-Z_-]", "", str(entity_type)) or "general"
    entity_id = re.sub(r"[^0-9a-zA-Z_-]", "", str(entity_id)) or "general"

    # Parse base64 data
    if "base64," in data:
        data = data.split("base64,")[1]

    try:
        file_content = base64.b64decode(data)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid base64 data")

    return await save_uploaded_content(file_content, entity_type, entity_id, "image/jpeg", "jpg")

# ============== PUSH NOTIFICATIONS ==============
@api_router.get("/notifications/vapid-public-key")
async def get_vapid_public_key():
    """Clave pública VAPID para suscribir Web Push en el navegador."""
    return {"public_key": os.environ.get("VAPID_PUBLIC_KEY", "")}

@api_router.post("/notifications/subscribe")
async def subscribe_push(request: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """Subscribe to push notifications"""
    subscription = request.get("subscription", {})
    
    # Store subscription in user document
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update users set push_subscription = $1 where id = $2 and company_id = $3",
            subscription,
            db_pg.as_uuid(current_user["id"]),
            db_pg.as_uuid(current_user["company_id"]),
        )
    
    return {"message": "Subscripción registrada"}

@api_router.delete("/notifications/unsubscribe")
async def unsubscribe_push(current_user: dict = Depends(get_current_user)):
    """Unsubscribe from push notifications"""
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update users set push_subscription = null where id = $1 and company_id = $2",
            db_pg.as_uuid(current_user["id"]),
            db_pg.as_uuid(current_user["company_id"]),
        )
    
    return {"message": "Subscripción eliminada"}

@api_router.get("/notifications")
async def get_notifications(
    unread_only: bool = False,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Get user notifications"""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))

    # Role-based filtering
    if current_user["role"] == "chofer":
        f.agregar(
            "(user_id = $? or target_role = 'chofer' or target_role = 'all')",
            db_pg.as_uuid(current_user["id"]),
        )

    if unread_only:
        f.crudo("not is_read")

    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from notifications where " + f.where
            + " order by created_at desc nulls last limit " + str(int(limit)),
            *f.values
        ))

@api_router.post("/notifications")
async def create_notification(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    """Create a notification (admin only)"""
    
    notification = {
        "id": str(uuid.uuid4()),
        "company_id": current_user["company_id"],
        "title": request.get("title", ""),
        "message": request.get("message", ""),
        "type": request.get("type", "info"),  # info, warning, alert, success
        "target_role": request.get("target_role", "all"),  # all, chofer, admin, etc.
        "user_id": request.get("user_id"),  # specific user
        "entity_type": request.get("entity_type"),
        "entity_id": request.get("entity_id"),
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["id"]
    }
    
    await _crear_notificacion(current_user["company_id"], notification)

    # Send push notification to subscribed users
    await send_push_notifications(
        current_user["company_id"],
        notification["title"],
        notification["message"],
        notification.get("target_role"),
        notification.get("user_id")
    )
    
    return {"id": notification["id"], "message": "Notificación creada"}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Mark notification as read"""
    # Con el filtro por empresa, que la consulta de Mongo no tenia: sin el,
    # cualquiera podia marcar como leida una notificacion de otro tenant.
    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update notifications set is_read = true, read_at = now() "
            "where id = $1 and company_id = $2",
            db_pg.as_uuid(notification_id),
            db_pg.as_uuid(current_user["company_id"]),
        )
    return {"message": "Notificación marcada como leída"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    """Mark all notifications as read"""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.crudo("not is_read")
    if current_user["role"] == "chofer":
        f.agregar(
            "(user_id = $? or target_role = 'chofer' or target_role = 'all')",
            db_pg.as_uuid(current_user["id"]),
        )

    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update notifications set is_read = true, read_at = now() where "
            + f.where, *f.values
        )
    return {"message": "Todas las notificaciones marcadas como leídas"}

async def send_push_notifications(company_id: str, title: str, message: str, target_role: str = None, user_id: str = None):
    """Send push notifications to subscribed users"""
    try:
        from pywebpush import webpush, WebPushException
        
        vapid_private = os.environ.get('VAPID_PRIVATE_KEY', '')
        vapid_public = os.environ.get('VAPID_PUBLIC_KEY', '')
        
        if not vapid_private or not vapid_public:
            return
        
        # Find users with push subscriptions
        query = {"company_id": company_id, "push_subscription": {"$exists": True}}
        if user_id:
            query["id"] = user_id
        elif target_role and target_role != "all":
            query["role"] = target_role
        
        # El filtro por suscripcion no nula lo hace ahora la base, apoyado en
        # el indice parcial users_company_push_idx.
        async with db_pg.tx({"company_id": company_id}) as conn:
            users = db_pg.rows_to_api(await conn.fetch(
                "select push_subscription from users "
                "where company_id = $1 and push_subscription is not null limit 1000",
                db_pg.as_uuid(company_id),
            ))
        
        payload = {
            "title": title,
            "body": message,
            "icon": "/logo192.png",
            "badge": "/logo192.png"
        }
        
        for user in users:
            subscription = user.get("push_subscription")
            if subscription:
                try:
                    webpush(
                        subscription,
                        data=str(payload),
                        vapid_private_key=vapid_private,
                        vapid_claims={"sub": "mailto:admin@transperu.com"}
                    )
                except WebPushException as e:
                    logging.error(f"Push notification error: {e}")
    except ImportError:
        logging.warning("pywebpush not installed, skipping push notifications")
    except Exception as e:
        logging.error(f"Push notification error: {e}")

# ============== ALERTS AUTO-GENERATION ==============
@api_router.post("/alerts/generate")
async def generate_alerts(current_user: dict = Depends(require_roles("owner", "admin"))):
    """Generate alerts for expiring documents and other issues"""
    company_id = current_user["company_id"]
    alerts_created = await _generate_document_alerts(company_id)
    return {"message": f"{alerts_created} alertas generadas"}

# ============== SYSTEM STATUS (public, no auth) ==============
@api_router.get("/system/status")
async def system_status():
    """Check if the system has been initialized (owner exists)"""
    async with db_pg.tx_global("autenticacion: comprobaciones previas al login") as conn:
        owner_exists = await conn.fetchval(
            "select 1 from users where role = 'owner'::user_role limit 1"
        ) is not None
        company_count = await conn.fetchval("select count(*) from companies")
    return {
        "initialized": owner_exists,
        "companies": company_count
    }

# ============== INSTALL TOKEN GUARD ==============
async def require_install_token(x_install_token: Optional[str] = Header(None)):
    """Protege endpoints de instalación (bootstrap/seed).
    - Si INSTALL_TOKEN está seteado, exige que el header X-Install-Token coincida.
    - Si no está seteado, solo permite en el primer arranque (cuando no existe superadmin todavía).
    """
    configured = os.environ.get("INSTALL_TOKEN")
    if configured:
        if not x_install_token or not secrets.compare_digest(str(x_install_token), configured):
            raise HTTPException(status_code=403, detail="Token de instalación inválido")
    else:
        async with db_pg.tx_global("autenticacion: comprobaciones previas al login") as conn:
            existing = await conn.fetchval(
                "select 1 from users where role = 'superadmin'::user_role limit 1"
            )
        if existing:
            raise HTTPException(
                status_code=403,
                detail="Instalación deshabilitada: configure INSTALL_TOKEN para permitir esta operación"
            )
    return True

# ============== BOOTSTRAP: CREATE SUPERADMIN ==============
@api_router.post("/bootstrap")
async def bootstrap_superadmin(_: bool = Depends(require_install_token)):
    """
    Create the first superadmin user.
    Only works when NO superadmin user exists in the system.
    This is the entry point for multi-tenancy management.
    """
    async with db_pg.tx_global("autenticacion: comprobaciones previas al login") as conn:
        existing = await conn.fetchval(
            "select 1 from users where role = 'superadmin'::user_role limit 1"
        )
    if existing:
        raise HTTPException(
            status_code=400,
            detail="Ya existe un superadmin. Use las credenciales de superadmin para acceder."
        )

    # Empresa del sistema: se REUSA si ya existe.
    #
    # Antes se insertaba una nueva sin mirar, y la unica guarda del endpoint es
    # que no exista el USUARIO superadmin. Si ese usuario se borraba o el
    # bootstrap se corria de nuevo, quedaba una segunda 'Star Insights IT' con
    # el mismo RUC placeholder — indistinguible de la primera en cualquier
    # pantalla. Fue exactamente lo que paso en produccion.
    #
    # Se busca por RUC ademas de por nombre porque el nombre lo puede editar
    # un admin desde la UI, y el RUC 00000000000 es el marcador real de que
    # esta es la empresa interna y no la de un cliente.
    SYSTEM_RUC = "00000000000"
    async with db_pg.tx_global("autenticacion: comprobaciones previas al login") as conn:
        existing_company = db_pg.to_api(await conn.fetchrow(
            "select * from companies where ruc = $1", SYSTEM_RUC
        ))
    if existing_company:
        system_company_id = existing_company["id"]
        logger.info(
            "bootstrap: reusando la empresa del sistema existente " + system_company_id
        )
    else:
        system_company = Company(
            name="Star Insights IT",
            ruc=SYSTEM_RUC,
            address="Sistema",
            phone="",
            email="admin@starinsights.pe",
            config={},
        )
        company_doc = system_company.model_dump()
        for key, value in company_doc.items():
            if isinstance(value, datetime):
                company_doc[key] = value.isoformat()
        # La empresa del sistema tambien necesita slug: la columna es NOT NULL
        # y sin esto el bootstrap fallaria en la primera instalacion nueva. Que
        # tenga direccion propia no la vuelve accesible: para entrar por ella
        # siguen haciendo falta las credenciales del superadmin.
        company_doc["slug"] = await tenant_host.slug_libre(system_company.name, SYSTEM_RUC)
        async with db_pg.tx({"company_id": system_company.id}) as conn:
            sql, values = db_pg.build_insert("companies", COMPANY_COLS, company_doc)
            await conn.execute(sql, *values)
        system_company_id = system_company.id

    # Generar password aleatorio (se muestra UNA sola vez)
    generated_password = secrets.token_urlsafe(12)

    # Create superadmin user
    superadmin = User(
        company_id=system_company_id,
        email="superadmin@starinsights.pe",
        name="Super Administrador",
        role=UserRole.SUPERADMIN,
        password_hash=hash_password(generated_password),
        force_password_change=True,
    )
    sa_doc = superadmin.model_dump()
    for key, value in sa_doc.items():
        if isinstance(value, datetime):
            sa_doc[key] = value.isoformat()
    async with db_pg.tx({"company_id": system_company_id}) as conn:
        sql, values = db_pg.build_insert("users", USER_COLS, sa_doc)
        await conn.execute(sql, *values)

    return {
        "message": "SuperAdmin creado exitosamente",
        "credentials": {
            "email": "superadmin@starinsights.pe",
            "password": generated_password,
            "role": "superadmin"
        },
        "company_id": system_company_id,
        "instructions": "Guarde esta contraseña ahora; no se volverá a mostrar. Deberá cambiarla en el primer inicio de sesión."
    }

# ============== SEED DATA ROUTE (FOR DEMO) ==============
@api_router.post("/seed")
async def seed_demo_data(_: bool = Depends(require_install_token)):
    """Create demo data for testing"""
    # Check if company already exists
    async with db_pg.tx_global("seed: buscar la empresa de demo en todo el sistema") as conn:
        existing = db_pg.to_api(await conn.fetchrow(
            "select * from companies where ruc = $1", "20123456789"
        ))
    if existing:
        # No devolver credenciales si ya existía
        return {"message": "Demo data already exists", "company_id": existing["id"]}
    
    # Create company
    company = Company(
        name="TransPeru Logistics S.A.C.",
        ruc="20123456789",
        address="Av. Industrial 123, Lima",
        phone="+51 1 555 0123",
        email="admin@transperu.com",
        config={
            "lockout_minutes": 15,
            "max_failed_attempts": 5,
            "tire_critical_depth": 3,
            "tire_warning_depth": 5,
            "tire_review_km_threshold": 5000,
            "viatico_por_viaje": 540,
            "maintenance_anticipation_km": 500
        }
    )
    company_doc = company.model_dump()
    for key, value in company_doc.items():
        if isinstance(value, datetime):
            company_doc[key] = value.isoformat()
    company_doc["slug"] = await tenant_host.slug_libre(company.name, company.ruc)
    # El contexto se fija en la empresa que se crea: la politica RLS permite
    # insertar en companies cuando id coincide con la empresa actual.
    async with db_pg.tx({"company_id": company.id}) as conn:
        sql, values = db_pg.build_insert("companies", COMPANY_COLS, company_doc)
        await conn.execute(sql, *values)
    
    # Create admin user con password aleatorio (se muestra UNA sola vez)
    admin_password = secrets.token_urlsafe(12)
    admin = User(
        company_id=company.id,
        email="admin@transperu.com",
        name="Administrador Principal",
        role=UserRole.ADMIN,
        password_hash=hash_password(admin_password),
        force_password_change=True,
    )
    admin_doc = admin.model_dump()
    for key, value in admin_doc.items():
        if isinstance(value, datetime):
            admin_doc[key] = value.isoformat()
    async with db_pg.tx({"company_id": company.id}) as conn:
        sql, values = db_pg.build_insert("users", USER_COLS, admin_doc)
        await conn.execute(sql, *values)
    
    # Create drivers
    drivers = [
        {"dni": "12345678", "name": "Juan Pérez Rodríguez", "pin": "123456", "license": "Q12345678", "phone": "+51 987 654 321"},
        {"dni": "87654321", "name": "Carlos García López", "pin": "654321", "license": "Q87654321", "phone": "+51 987 654 322"},
        {"dni": "11223344", "name": "Pedro Mendoza Silva", "pin": "112233", "license": "Q11223344", "phone": "+51 987 654 323"},
    ]
    
    driver_ids = []
    for d in drivers:
        driver = User(
            company_id=company.id,
            dni=d["dni"],
            name=d["name"],
            role=UserRole.CHOFER,
            pin_hash=hash_password(d["pin"]),
            license_number=d["license"],
            license_expiry=datetime.now(timezone.utc) + timedelta(days=365),
            phone=d["phone"]
        )
        driver_doc = driver.model_dump()
        for key, value in driver_doc.items():
            if isinstance(value, datetime):
                driver_doc[key] = value.isoformat()
        async with db_pg.tx({"company_id": company.id}) as conn:
            sql, values = db_pg.build_insert("users", USER_COLS, driver_doc)
            await conn.execute(sql, *values)
        driver_ids.append(driver.id)
    
    # Create vehicles - Tractos
    tractos = [
        {"plate": "ABC-123", "brand": "Volvo", "model": "FH16", "year": 2022, "color": "Blanco"},
        {"plate": "DEF-456", "brand": "Scania", "model": "R500", "year": 2021, "color": "Rojo"},
        {"plate": "GHI-789", "brand": "Mercedes", "model": "Actros", "year": 2023, "color": "Azul"},
    ]
    
    tracto_ids = []
    for t in tractos:
        vehicle = Vehicle(
            company_id=company.id,
            plate=t["plate"],
            vehicle_type=VehicleType.TRACTO,
            brand=t["brand"],
            model=t["model"],
            year=t["year"],
            color=t["color"],
            fuel_capacity=400,
            tire_config="6"
        )
        vehicle_doc = vehicle.model_dump()
        for key, value in vehicle_doc.items():
            if isinstance(value, datetime):
                vehicle_doc[key] = value.isoformat()
        async with db_pg.tx({"company_id": company.id}) as conn:
            sql, values = db_pg.build_insert("vehicles", VEHICLE_COLS, vehicle_doc)
            await conn.execute(sql, *values)
        tracto_ids.append(vehicle.id)
    
    # Create vehicles - Carretas
    carretas = [
        {"plate": "T1A-001", "brand": "Facchini", "model": "Plataforma", "year": 2021},
        {"plate": "T1B-002", "brand": "Guerra", "model": "Cisterna", "year": 2020},
        {"plate": "T1C-003", "brand": "Randon", "model": "Furgón", "year": 2022},
    ]
    
    carreta_ids = []
    for c in carretas:
        vehicle = Vehicle(
            company_id=company.id,
            plate=c["plate"],
            vehicle_type=VehicleType.CARRETA,
            brand=c["brand"],
            model=c["model"],
            year=c["year"],
            tire_config="6"
        )
        vehicle_doc = vehicle.model_dump()
        for key, value in vehicle_doc.items():
            if isinstance(value, datetime):
                vehicle_doc[key] = value.isoformat()
        async with db_pg.tx({"company_id": company.id}) as conn:
            sql, values = db_pg.build_insert("vehicles", VEHICLE_COLS, vehicle_doc)
            await conn.execute(sql, *values)
        carreta_ids.append(vehicle.id)
    
    # Create document types
    doc_types = [
        {"name": "SOAT", "applies_to": "vehiculo", "is_critical": True, "block_rule": "bloquea_inicio"},
        {"name": "Revisión Técnica (CITV)", "applies_to": "vehiculo", "is_critical": True, "block_rule": "bloquea_inicio"},
        {"name": "Tarjeta de Propiedad", "applies_to": "vehiculo", "is_critical": True, "block_rule": "bloquea_asignacion"},
        {"name": "Tarjeta de Circulación", "applies_to": "vehiculo", "is_critical": False, "block_rule": "solo_alerta"},
        {"name": "Bonificación", "applies_to": "vehiculo", "is_critical": False, "block_rule": "solo_alerta"},
        {"name": "Póliza de Seguro", "applies_to": "vehiculo", "is_critical": False, "block_rule": "solo_alerta"},
        {"name": "Licencia de Conducir", "applies_to": "chofer", "is_critical": True, "block_rule": "bloquea_asignacion"},
        {"name": "DNI", "applies_to": "chofer", "is_critical": True, "block_rule": "bloquea_asignacion"},
        {"name": "Certificado Médico", "applies_to": "chofer", "is_critical": False, "block_rule": "solo_alerta"},
    ]
    
    doc_type_ids = {}
    for dt in doc_types:
        # Idempotente: no duplicar tipos de documento por nombre
        async with db_pg.tx({"company_id": company.id}) as conn:
            existing_dt = await conn.fetchval(
                "select id from document_types where company_id = $1 and name = $2",
                db_pg.as_uuid(company.id), dt["name"],
            )
        if existing_dt:
            doc_type_ids[dt["name"]] = str(existing_dt)
            continue
        doc_type = DocumentType(
            company_id=company.id,
            name=dt["name"],
            applies_to=dt["applies_to"],
            is_critical=dt["is_critical"],
            block_rule=BlockRule(dt["block_rule"])
        )
        sql, values = db_pg.build_insert(
            "document_types", DOCUMENT_TYPE_COLS,
            _modelo_a_fila(doc_type.model_dump()),
        )
        async with db_pg.tx({"company_id": company.id}) as conn:
            await conn.execute(sql, *values)
        doc_type_ids[dt["name"]] = doc_type.id
    
    # Create routes
    routes = [
        {"name": "Lima - Arequipa", "origin": "Lima", "destination": "Arequipa", "distance_km": 1010, "estimated_hours": 16, "toll_cost": 180},
        {"name": "Lima - Trujillo", "origin": "Lima", "destination": "Trujillo", "distance_km": 560, "estimated_hours": 8, "toll_cost": 95},
        {"name": "Lima - Cusco", "origin": "Lima", "destination": "Cusco", "distance_km": 1105, "estimated_hours": 20, "toll_cost": 210},
    ]
    
    route_ids = []
    for r in routes:
        route = Route(
            company_id=company.id,
            name=r["name"],
            origin=r["origin"],
            destination=r["destination"],
            distance_km=r["distance_km"],
            estimated_hours=r["estimated_hours"],
            toll_cost=r["toll_cost"]
        )
        sql, values = db_pg.build_insert(
            "routes", ROUTE_COLS, _modelo_a_fila(route.model_dump())
        )
        async with db_pg.tx({"company_id": company.id}) as conn:
            await conn.execute(sql, *values)
        route_ids.append(route.id)
    
    # Create some tires
    tire_brands = ["Michelin", "Bridgestone", "Goodyear", "Continental"]
    tire_positions_tracto = ["T-1L", "T-1R", "T-2L1", "T-2L2", "T-2R1", "T-2R2"]
    tire_positions_carreta = ["C-A-L", "C-A-R", "C-B-L", "C-B-R", "C-C-L", "C-C-R"]
    
    # Tractos y carretas se siembran por la MISMA tabla, a proposito. Antes
    # habia un solo bucle, escrito a mano para los tractos, y
    # tire_positions_carreta quedaba declarada y sin usar: las carretas se
    # sembraban sin una sola llanta pese a llevar tire_config="6". Duplicar el
    # cuerpo para arreglarlo habria sido repetir el error que lo causo.
    for prefijo, ids_unidad, posiciones in (
        ("TR", tracto_ids, tire_positions_tracto),
        ("CR", carreta_ids, tire_positions_carreta),
    ):
        for i, vehicle_id in enumerate(ids_unidad):
            for j, pos in enumerate(posiciones):
                tire = Tire(
                    company_id=company.id,
                    serial=f"{prefijo}{i+1}-{pos}-{j+1:03d}",
                    brand=tire_brands[j % len(tire_brands)],
                    model="XZA3",
                    dimension="295/80R22.5",
                    purchase_cost=450,
                    status=TireStatus.EN_USO,
                    current_vehicle_id=vehicle_id,
                    current_position=pos
                )
                # La llanta nace montada, asi que tambien nace su registro de
                # montaje. Sin el queda en_uso y ubicada pero sin historial: el
                # esquema de la unidad no puede calcular km_recorridos ni
                # cost_per_km, y un desmontaje posterior no encuentra que fila
                # cerrar y se va sin dejar rastro. Es lo que le paso a la
                # semilla anterior, cuyas diez llantas hubo que reconstruir a
                # mano.
                mount = TireMount(
                    company_id=company.id,
                    tire_id=tire.id,
                    vehicle_id=vehicle_id,
                    position_code=pos,
                    mount_odometer=0,
                )
                sql_llanta, val_llanta = db_pg.build_insert(
                    "tires", TIRE_COLS, _modelo_a_fila(tire.model_dump())
                )
                sql_montaje, val_montaje = db_pg.build_insert(
                    "tire_mounts", TIRE_MOUNT_COLS, _modelo_a_fila(mount.model_dump())
                )
                async with db_pg.tx({"company_id": company.id}) as conn:
                    await conn.execute(sql_llanta, *val_llanta)
                    await conn.execute(sql_montaje, *val_montaje)
    
    # Create a sample trip
    trip = Trip(
        company_id=company.id,
        tracto_id=tracto_ids[0],
        carreta_id=carreta_ids[0],
        driver_id=driver_ids[0],
        route_id=route_ids[0],
        client_name="Minera Sur S.A.",
        cargo_description="Maquinaria pesada",
        cargo_weight=25000,
        status=TripStatus.PROGRAMADO,
        scheduled_date=datetime.now(timezone.utc) + timedelta(days=1)
    )
    sql, values = db_pg.build_insert(
        "trips", TRIP_COLS, _modelo_a_fila(trip.model_dump())
    )
    async with db_pg.tx({"company_id": company.id}) as conn:
        await conn.execute(sql, *values)
    
    return {
        "message": "Demo data created successfully",
        "company_id": company.id,
        "admin_email": "admin@transperu.com",
        "admin_password": admin_password,
        "instructions": "Guarde esta contraseña ahora; no se volverá a mostrar. Deberá cambiarla en el primer inicio de sesión.",
        "sample_driver": {
            "dni": "12345678",
            "pin": "123456"
        }
    }

# ============== REPORTS ENDPOINTS ==============
from fastapi.responses import StreamingResponse
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

@api_router.get("/reports/trips")
async def get_trips_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    driver_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get trips report data"""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(start_date, "scheduled_date >= $?", db_pg.as_ts(start_date))
    f.si(end_date, "scheduled_date <= $?", db_pg.as_ts(end_date))
    # status::text y no un cast a trip_status: el valor lo elige el cliente, y
    # uno que no exista en el enum haria fallar el cast con un 500. Comparando
    # como texto un status desconocido devuelve lista vacia, igual que Mongo.
    f.si(status, "status::text = $?", status)
    f.si(driver_id, "driver_id = $?", db_pg.as_uuid(driver_id))
    async with db_pg.tx(current_user) as conn:
        trips = db_pg.rows_to_api(await conn.fetch(
            "select * from trips where " + f.where
            + " order by scheduled_date desc nulls last limit 500",
            *f.values,
        ))
    
    # Enrich with driver and vehicle info
    for trip in trips:
        driver = await _usuario_pg(current_user["company_id"], trip.get("driver_id"))
        tracto = await _vehiculo_pg(current_user["company_id"], trip.get("tracto_id"))
        trip["driver_name"] = driver.get("name") if driver else "-"
        trip["tracto_plate"] = tracto.get("plate") if tracto else "-"
    
    # Calculate totals
    total_km = sum((t.get("km_end", 0) or 0) - (t.get("km_start", 0) or 0) for t in trips)
    total_advances = sum(t.get("total_advance", 0) or 0 for t in trips)
    total_expenses = sum(t.get("total_expenses", 0) or 0 for t in trips)
    
    return {
        "trips": [serialize_doc(t) for t in trips],
        "totals": {
            "count": len(trips),
            "total_km": total_km,
            "total_advances": total_advances,
            "total_expenses": total_expenses,
            "balance": total_advances - total_expenses
        }
    }

@api_router.get("/reports/trips/export/excel")
async def export_trips_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export trips report to Excel"""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(start_date, "scheduled_date >= $?", db_pg.as_ts(start_date))
    f.si(end_date, "scheduled_date <= $?", db_pg.as_ts(end_date))
    async with db_pg.tx(current_user) as conn:
        trips = db_pg.rows_to_api(await conn.fetch(
            "select * from trips where " + f.where
            + " order by scheduled_date desc nulls last limit 500",
            *f.values,
        ))
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Viajes"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Fecha", "Tracto", "Carreta", "Chofer", "Cliente", "Carga", "Estado", "Km Inicio", "Km Fin", "Anticipo", "Gastos", "Saldo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Data
    for row, trip in enumerate(trips, 2):
        driver = await _usuario_pg(current_user["company_id"], trip.get("driver_id"))
        tracto = await _vehiculo_pg(current_user["company_id"], trip.get("tracto_id"))
        carreta = await _vehiculo_pg(current_user["company_id"], trip.get("carreta_id")) if trip.get("carreta_id") else None
        
        data = [
            trip.get("scheduled_date", "")[:10] if trip.get("scheduled_date") else "",
            tracto.get("plate") if tracto else "-",
            carreta.get("plate") if carreta else "-",
            driver.get("name") if driver else "-",
            trip.get("client_name", "-"),
            trip.get("cargo_description", "-"),
            trip.get("status", "-"),
            trip.get("km_start", 0),
            trip.get("km_end", 0),
            trip.get("total_advance", 0),
            trip.get("total_expenses", 0),
            (trip.get("total_advance", 0) or 0) - (trip.get("total_expenses", 0) or 0)
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_viajes.xlsx"}
    )

@api_router.get("/reports/settlements/export/pdf/{trip_id}")
async def export_settlement_pdf(trip_id: str, current_user: dict = Depends(get_current_user)):
    """Export settlement to PDF"""
    trip = await _viaje_pg(current_user["company_id"], trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")
    
    driver = await _usuario_pg(current_user["company_id"], trip.get("driver_id"))
    tracto = await _vehiculo_pg(current_user["company_id"], trip.get("tracto_id"))
    async with db_pg.tx(current_user) as conn:
        # Con el filtro por empresa, que la consulta de Mongo no tenia.
        advances = db_pg.rows_to_api(await conn.fetch(
            "select * from trip_advances where trip_id = $1 and company_id = $2 "
            "order by delivered_date desc nulls last limit 100",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        ))
        expenses = db_pg.rows_to_api(await conn.fetch(
            "select * from trip_expenses where trip_id = $1 and company_id = $2 "
            "order by expense_date desc nulls last limit 100",
            db_pg.as_uuid(trip_id), db_pg.as_uuid(current_user["company_id"]),
        ))
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
    elements.append(Paragraph("LIQUIDACIÓN DE VIAJE", title_style))
    elements.append(Spacer(1, 12))
    
    # Trip info
    info_data = [
        ["Fecha:", trip.get("scheduled_date", "")[:10] if trip.get("scheduled_date") else "-"],
        ["Chofer:", driver.get("name") if driver else "-"],
        ["Vehículo:", tracto.get("plate") if tracto else "-"],
        ["Cliente:", trip.get("client_name", "-")],
        ["Carga:", trip.get("cargo_description", "-")],
    ]
    info_table = Table(info_data, colWidths=[100, 300])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Advances table
    elements.append(Paragraph("ANTICIPOS", styles['Heading2']))
    if advances:
        adv_data = [["Fecha", "Método", "Monto"]]
        for adv in advances:
            adv_data.append([
                adv.get("delivered_date", "")[:10] if adv.get("delivered_date") else "-",
                adv.get("payment_method", "-"),
                f"S/ {adv.get('amount', 0):.2f}"
            ])
        adv_table = Table(adv_data, colWidths=[150, 150, 100])
        adv_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(adv_table)
    else:
        elements.append(Paragraph("Sin anticipos registrados", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Expenses table
    elements.append(Paragraph("GASTOS", styles['Heading2']))
    if expenses:
        exp_data = [["Categoría", "Descripción", "Proveedor", "Monto"]]
        for exp in expenses:
            exp_data.append([
                exp.get("category", "-"),
                exp.get("description", "-")[:30],
                exp.get("provider", "-"),
                f"S/ {exp.get('amount', 0):.2f}"
            ])
        exp_table = Table(exp_data, colWidths=[100, 150, 100, 80])
        exp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(exp_table)
    else:
        elements.append(Paragraph("Sin gastos registrados", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary
    total_advances = sum(a.get("amount", 0) for a in advances)
    total_expenses = sum(e.get("amount", 0) for e in expenses)
    balance = total_advances - total_expenses
    
    summary_data = [
        ["Total Anticipos:", f"S/ {total_advances:.2f}"],
        ["Total Gastos:", f"S/ {total_expenses:.2f}"],
        ["SALDO:", f"S/ {abs(balance):.2f} {'(A favor empresa)' if balance >= 0 else '(A favor chofer)'}"],
    ]
    summary_table = Table(summary_data, colWidths=[150, 200])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 2), (-1, 2), colors.lightgrey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 40))
    
    # Signatures
    sig_data = [["", ""], ["_______________________", "_______________________"], ["Chofer", "Contabilidad"]]
    sig_table = Table(sig_data, colWidths=[200, 200])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, 1), 40),
    ]))
    elements.append(sig_table)
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=liquidacion_{trip_id[:8]}.pdf"}
    )

@api_router.get("/reports/fuel")
async def get_fuel_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    vehicle_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get fuel consumption report"""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(start_date, "load_date >= $?", db_pg.as_ts(start_date))
    f.si(end_date, "load_date <= $?", db_pg.as_ts(end_date))
    f.si(vehicle_id, "vehicle_id = $?", db_pg.as_uuid(vehicle_id))
    async with db_pg.tx(current_user) as conn:
        loads = db_pg.rows_to_api(await conn.fetch(
            "select * from fuel_loads where " + f.where
            + " order by load_date desc nulls last limit 500", *f.values
        ))
    
    # Calculate totals and KPIs
    total_liters = sum(l.get("liters", 0) for l in loads)
    total_amount = sum(l.get("total_amount", 0) for l in loads)
    
    # Group by vehicle
    by_vehicle = {}
    for load in loads:
        vid = load.get("vehicle_id", "unknown")
        if vid not in by_vehicle:
            by_vehicle[vid] = {"liters": 0, "amount": 0, "loads": 0}
        by_vehicle[vid]["liters"] += load.get("liters", 0)
        by_vehicle[vid]["amount"] += load.get("total_amount", 0)
        by_vehicle[vid]["loads"] += 1

    # Group by driver (resuelve chofer vía viaje asociado; fallback a created_by)
    trip_ids = list({l.get("trip_id") for l in loads if l.get("trip_id")})
    trips_map = {}
    if trip_ids:
        async with db_pg.tx(current_user) as conn:
            trips = db_pg.rows_to_api(await conn.fetch(
                "select id, driver_id from trips "
                "where company_id = $1 and id = any($2::uuid[]) limit 1000",
                db_pg.as_uuid(current_user["company_id"]),
                [u for u in (db_pg.as_uuid(t) for t in trip_ids) if u],
            ))
        trips_map = {t["id"]: t.get("driver_id") for t in trips}

    by_driver = {}
    driver_ids_seen = set()
    for load in loads:
        did = trips_map.get(load.get("trip_id")) or load.get("created_by") or "unknown"
        driver_ids_seen.add(did)
        if did not in by_driver:
            by_driver[did] = {"driver_id": did, "driver_name": None, "liters": 0, "amount": 0, "loads": 0}
        by_driver[did]["liters"] += load.get("liters", 0)
        by_driver[did]["amount"] += load.get("total_amount", 0)
        by_driver[did]["loads"] += 1

    real_driver_ids = [d for d in driver_ids_seen if d and d != "unknown"]
    if real_driver_ids:
        async with db_pg.tx(current_user) as conn:
            drivers = db_pg.rows_to_api(await conn.fetch(
                "select id, name from users where company_id = $1 and id = any($2::uuid[])",
                db_pg.as_uuid(current_user["company_id"]),
                [db_pg.as_uuid(x) for x in real_driver_ids],
            ))
        name_map = {u["id"]: u.get("name") for u in drivers}
        for did, agg in by_driver.items():
            agg["driver_name"] = name_map.get(did, "N/A" if did == "unknown" else did)

    return {
        "loads": [serialize_doc(l) for l in loads],
        "totals": {
            "total_liters": total_liters,
            "total_amount": total_amount,
            "total_loads": len(loads),
            "avg_price_per_liter": total_amount / total_liters if total_liters > 0 else 0
        },
        "by_vehicle": by_vehicle,
        "by_driver": by_driver
    }

@api_router.get("/reports/maintenance")
async def get_maintenance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    vehicle_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get maintenance report"""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(start_date, "created_at >= $?", db_pg.as_ts(start_date))
    f.si(end_date, "created_at <= $?", db_pg.as_ts(end_date))
    f.si(vehicle_id, "vehicle_id = $?", db_pg.as_uuid(vehicle_id))
    async with db_pg.tx(current_user) as conn:
        work_orders = db_pg.rows_to_api(await conn.fetch(
            "select * from work_orders where " + f.where
            + " order by created_at desc limit 500", *f.values
        ))
    
    # Calculate totals
    total_cost = sum(wo.get("total_cost", 0) for wo in work_orders)
    by_status = {}
    by_type = {}
    
    for wo in work_orders:
        status = wo.get("status", "unknown")
        by_status[status] = by_status.get(status, 0) + 1
        
        order_type = wo.get("order_type", "unknown")
        by_type[order_type] = by_type.get(order_type, 0) + 1
    
    return {
        "work_orders": [serialize_doc(wo) for wo in work_orders],
        "totals": {
            "count": len(work_orders),
            "total_cost": total_cost
        },
        "by_status": by_status,
        "by_type": by_type
    }

# ============== CONFIGURATION ENDPOINTS ==============
@api_router.get("/config/document-types")
async def get_document_types_config(current_user: dict = Depends(get_current_user)):
    """Get document types configuration"""
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from document_types where company_id = $1 "
            "order by name limit 100",
            db_pg.as_uuid(current_user["company_id"]),
        ))

@api_router.post("/config/document-types")
async def create_document_type_config(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    """Create new document type"""
    
    doc_type = DocumentType(
        company_id=current_user["company_id"],
        name=request.get("name"),
        applies_to=request.get("applies_to", "vehiculo"),
        is_critical=request.get("is_critical", False),
        requires_expiry=request.get("requires_expiry", True),
        alert_days=request.get("alert_days", [60, 30, 15, 7, 3, 1]),
        block_rule=BlockRule(request.get("block_rule", "solo_alerta"))
    )
    
    sql, values = db_pg.build_insert(
        "document_types", DOCUMENT_TYPE_COLS, _modelo_a_fila(doc_type.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)

    return {"id": doc_type.id, "message": "Tipo de documento creado"}

@api_router.put("/config/document-types/{doc_type_id}")
async def update_document_type_config(doc_type_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    """Update document type"""
    
    update_data = {}
    if "name" in request:
        update_data["name"] = request["name"]
    if "is_critical" in request:
        update_data["is_critical"] = request["is_critical"]
    if "requires_expiry" in request:
        update_data["requires_expiry"] = request["requires_expiry"]
    if "alert_days" in request:
        update_data["alert_days"] = request["alert_days"]
    if "block_rule" in request:
        update_data["block_rule"] = request["block_rule"]
    
    update_data["id"] = doc_type_id
    update_data["company_id"] = current_user["company_id"]
    sql, values = db_pg.build_update(
        "document_types", DOCUMENT_TYPE_COLS, update_data, ["id", "company_id"]
    )
    if sql:
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)

    return {"message": "Tipo de documento actualizado"}

@api_router.delete("/config/document-types/{doc_type_id}")
async def delete_document_type_config(doc_type_id: str, current_user: dict = Depends(require_roles("owner", "admin"))):
    """Delete document type"""
    
    # Check if any documents use this type
    async with db_pg.tx(current_user) as conn:
        docs_count = await conn.fetchval(
            "select count(*) from documents "
            "where company_id = $1 and document_type_id = $2",
            db_pg.as_uuid(current_user["company_id"]), db_pg.as_uuid(doc_type_id),
        )

        if docs_count > 0:
            raise HTTPException(status_code=400, detail=f"No se puede eliminar: {docs_count} documentos usan este tipo")

        await conn.execute(
            "delete from document_types where id = $1 and company_id = $2",
            db_pg.as_uuid(doc_type_id), db_pg.as_uuid(current_user["company_id"]),
        )
    
    return {"message": "Tipo de documento eliminado"}

@api_router.get("/config/checklist-templates")
async def get_checklist_templates_config(current_user: dict = Depends(get_current_user)):
    """Get checklist templates"""
    async with db_pg.tx(current_user) as conn:
        return db_pg.rows_to_api(await conn.fetch(
            "select * from checklist_templates where company_id = $1 "
            "order by name limit 100",
            db_pg.as_uuid(current_user["company_id"]),
        ))

@api_router.post("/config/checklist-templates")
async def create_checklist_template_config(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    """Create checklist template"""
    
    template = ChecklistTemplate(
        company_id=current_user["company_id"],
        name=request.get("name"),
        vehicle_type=request.get("vehicle_type"),
        items=request.get("items", []),
        is_active=request.get("is_active", True),
        created_by=current_user["id"]
    )
    
    sql, values = db_pg.build_insert(
        "checklist_templates", CHECKLIST_TEMPLATE_COLS,
        _modelo_a_fila(template.model_dump()),
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)

    return {"id": template.id, "message": "Plantilla creada"}

@api_router.put("/config/checklist-templates/{template_id}")
async def update_checklist_template_config(template_id: str, request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    """Update checklist template"""
    
    update_data = {}
    if "name" in request:
        update_data["name"] = request["name"]
    if "vehicle_type" in request:
        update_data["vehicle_type"] = request["vehicle_type"]
    if "items" in request:
        update_data["items"] = request["items"]
    if "is_active" in request:
        update_data["is_active"] = request["is_active"]
    
    update_data["id"] = template_id
    update_data["company_id"] = current_user["company_id"]
    sql, values = db_pg.build_update(
        "checklist_templates", CHECKLIST_TEMPLATE_COLS, update_data, ["id", "company_id"]
    )
    if sql:
        async with db_pg.tx(current_user) as conn:
            await conn.execute(sql, *values)

    return {"message": "Plantilla actualizada"}

@api_router.get("/config/company")
async def get_company_config(current_user: dict = Depends(get_current_user)):
    """Get company configuration"""
    company = await _empresa_pg(current_user["company_id"])
    return serialize_doc(company)

@api_router.put("/config/company")
async def update_company_config(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):
    """Update company configuration"""
    
    update_data = {}
    for field in ["name", "address", "phone", "email", "logo_url", "brand_color", "config"]:
        if field in request:
            update_data[field] = request[field]
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    update_data["id"] = current_user["company_id"]
    sql, values = db_pg.build_update("companies", COMPANY_COLS, update_data, ["id"])
    async with db_pg.tx(current_user) as conn:
        if sql:
            await conn.execute(sql, *values)
    
    return {"message": "Configuración actualizada"}

# ============== GUÍAS DE TRANSPORTISTA Y FACTURAS (SUNAT) ==============
class GuiaTransportistaStatus(str, Enum):
    BORRADOR = "borrador"
    EMITIDA = "emitida"
    ANULADA = "anulada"
    ERROR = "error"

class FacturaStatus(str, Enum):
    BORRADOR = "borrador"
    EMITIDA = "emitida"
    PAGADA = "pagada"
    ANULADA = "anulada"
    ERROR = "error"

class GuiaTransportista(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    trip_id: Optional[str] = None
    serie: str = "T001"
    numero: Optional[int] = None
    fecha_emision: Optional[str] = None
    # Remitente
    remitente_ruc: Optional[str] = None
    remitente_razon_social: Optional[str] = None
    # Destinatario
    destinatario_ruc: Optional[str] = None
    destinatario_razon_social: Optional[str] = None
    # Transportista
    transportista_ruc: Optional[str] = None
    transportista_razon_social: Optional[str] = None
    # Ruta
    punto_partida: Optional[str] = None
    punto_partida_ubigeo: Optional[str] = None
    punto_llegada: Optional[str] = None
    punto_llegada_ubigeo: Optional[str] = None
    # Vehículo / Conductor
    vehiculo_placa: Optional[str] = None
    conductor_dni: Optional[str] = None
    conductor_nombre: Optional[str] = None
    conductor_licencia: Optional[str] = None
    # Carga
    descripcion_carga: Optional[str] = None
    peso_bruto: Optional[float] = None
    unidad_peso: str = "KGM"
    num_bultos: Optional[int] = None
    # SUNAT
    sunat_response: Optional[Dict[str, Any]] = None
    sunat_ticket: Optional[str] = None
    sunat_cdr: Optional[str] = None
    pdf_url: Optional[str] = None
    status: GuiaTransportistaStatus = GuiaTransportistaStatus.BORRADOR
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class Factura(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    trip_id: Optional[str] = None
    serie: str = "F001"
    numero: Optional[int] = None
    fecha_emision: Optional[str] = None
    # Cliente
    cliente_ruc: Optional[str] = None
    cliente_razon_social: Optional[str] = None
    cliente_direccion: Optional[str] = None
    # Detalle
    items: List[Dict[str, Any]] = []  # [{descripcion, cantidad, precio_unitario, igv, total}]
    subtotal: float = 0
    igv: float = 0
    total: float = 0
    moneda: str = "PEN"
    # SUNAT
    sunat_response: Optional[Dict[str, Any]] = None
    sunat_ticket: Optional[str] = None
    sunat_cdr: Optional[str] = None
    pdf_url: Optional[str] = None
    status: FacturaStatus = FacturaStatus.BORRADOR
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

# ============== TABLAS EN POSTGRES: FACTURACION / SUNAT / CAJA ==============
# Estas cuatro tablas ya cortaron (db/migrations/002_corte_facturacion.sql):
# Postgres es su fuente de verdad y NO se escriben mas en Mongo. Lo que sigan
# consultando de companies/trips/vehicles se lee de Mongo, que es correcto:
# esas todavia no cortaron.
#
# El mapa {columna: tipo} es la lista blanca de columnas escribibles (nada que
# venga del request puede sumar una) y le dice a db_pg como convertir cada
# valor y cuando hace falta un cast SQL.

GUIA_COLS = {
    "id": "uuid", "company_id": "uuid", "trip_id": "uuid",
    "serie": "text", "numero": "int", "fecha_emision": "text",
    "remitente_ruc": "text", "remitente_razon_social": "text",
    "destinatario_ruc": "text", "destinatario_razon_social": "text",
    "transportista_ruc": "text", "transportista_razon_social": "text",
    "punto_partida": "text", "punto_partida_ubigeo": "text",
    "punto_llegada": "text", "punto_llegada_ubigeo": "text",
    "vehiculo_placa": "text", "conductor_dni": "text",
    "conductor_nombre": "text", "conductor_licencia": "text",
    "descripcion_carga": "text", "peso_bruto": "float",
    "unidad_peso": "text", "num_bultos": "int",
    "sunat_response": "json", "sunat_ticket": "text", "sunat_cdr": "text",
    "pdf_url": "text", "status": "enum:guia_transportista_status",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}

FACTURA_COLS = {
    "id": "uuid", "company_id": "uuid", "trip_id": "uuid",
    "serie": "text", "numero": "int", "fecha_emision": "text",
    "cliente_ruc": "text", "cliente_razon_social": "text", "cliente_direccion": "text",
    "items": "json", "subtotal": "float", "igv": "float", "total": "float",
    "moneda": "text",
    "sunat_response": "json", "sunat_ticket": "text", "sunat_cdr": "text",
    "pdf_url": "text", "status": "enum:factura_status",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}

DETRACCION_COLS = {
    "id": "uuid", "company_id": "uuid", "factura_id": "uuid", "trip_id": "uuid",
    "client_ruc": "text", "client_name": "text",
    "comprobante_serie": "text", "comprobante_numero": "text",
    "fecha_emision": "text",
    "base_amount": "float", "rate": "float", "amount": "float",
    "codigo_bien_servicio": "text", "constancia_number": "text",
    "deposit_date": "text", "status": "enum:detraccion_status", "notes": "text",
    "anulada_at": "ts", "anulada_by": "uuid",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}

CASH_MOVEMENT_COLS = {
    "id": "uuid", "company_id": "uuid", "movement_number": "text", "date": "text",
    "type": "enum:cash_movement_type", "concept": "text", "category": "text",
    "amount": "float", "payment_method": "enum:cash_payment_method",
    "reference": "text", "trip_id": "uuid", "vehicle_id": "uuid",
    "client_ruc": "text", "supplier": "text", "receipt_url": "text", "notes": "text",
    "deleted": "bool", "deleted_at": "ts", "deleted_by": "uuid",
    "created_at": "ts", "updated_at": "ts", "created_by": "uuid",
}


# --- Guía Endpoints ---
@api_router.get("/guias")
async def get_guias(
    trip_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(trip_id, "trip_id = $?", db_pg.as_uuid(trip_id))
    if status:
        # Un estado fuera del enum reventaria el cast en Postgres; con Mongo
        # simplemente no casaba con nada. Se conserva esa forma: lista vacia.
        if status not in [e.value for e in GuiaTransportistaStatus]:
            return []
        f.agregar("status = $?::guia_transportista_status", status)
    async with db_pg.tx(current_user) as conn:
        rows = await conn.fetch(
            "select * from guias_transportista where " + f.where
            + " order by created_at desc limit 500",
            *f.values,
        )
    return db_pg.rows_to_api(rows)

@api_router.get("/guias/{guia_id}")
async def get_guia(guia_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        row = await conn.fetchrow(
            "select * from guias_transportista where id = $1 and company_id = $2",
            db_pg.as_uuid(guia_id), db_pg.as_uuid(current_user["company_id"]),
        )
    if not row:
        raise HTTPException(status_code=404, detail="Guía no encontrada")
    return db_pg.to_api(row)

@api_router.post("/guias")
async def create_guia(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "operaciones", "contabilidad"))):

    serie = request.get("serie", "T001")
    async with db_pg.tx(current_user) as conn:
        # Correlativo por (empresa, serie). Mantiene la misma carrera que tenia
        # con Mongo: dos altas simultaneas de la misma serie pueden calcular el
        # mismo numero. Se arregla con una secuencia cuando toque, no ahora.
        ultimo = await conn.fetchval(
            "select max(numero) from guias_transportista "
            "where company_id = $1 and serie = $2",
            db_pg.as_uuid(current_user["company_id"]), serie,
        )
        next_num = (ultimo or 0) + 1

        guia = GuiaTransportista(
            company_id=current_user["company_id"],
            trip_id=request.get("trip_id"),
            serie=serie,
            numero=next_num,
            fecha_emision=request.get("fecha_emision", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            remitente_ruc=request.get("remitente_ruc"),
            remitente_razon_social=request.get("remitente_razon_social"),
            destinatario_ruc=request.get("destinatario_ruc"),
            destinatario_razon_social=request.get("destinatario_razon_social"),
            transportista_ruc=request.get("transportista_ruc"),
            transportista_razon_social=request.get("transportista_razon_social"),
            punto_partida=request.get("punto_partida"),
            punto_partida_ubigeo=request.get("punto_partida_ubigeo"),
            punto_llegada=request.get("punto_llegada"),
            punto_llegada_ubigeo=request.get("punto_llegada_ubigeo"),
            vehiculo_placa=request.get("vehiculo_placa"),
            conductor_dni=request.get("conductor_dni"),
            conductor_nombre=request.get("conductor_nombre"),
            conductor_licencia=request.get("conductor_licencia"),
            descripcion_carga=request.get("descripcion_carga"),
            peso_bruto=request.get("peso_bruto"),
            num_bultos=request.get("num_bultos"),
            created_by=current_user["id"]
        )
        sql, values = db_pg.build_insert(
            "guias_transportista", GUIA_COLS, _modelo_a_fila(guia.model_dump())
        )
        await conn.execute(sql, *values)

    return {"id": guia.id, "numero": f"{guia.serie}-{next_num:08d}", "message": "Guía creada"}

@api_router.post("/guias/{guia_id}/emit")
async def emit_guia_sunat(guia_id: str, current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))):
    """Emit guía to SUNAT - requires SUNAT API credentials in company config"""
    async with db_pg.tx(current_user) as conn:
        existe = await conn.fetchval(
            "select 1 from guias_transportista where id = $1 and company_id = $2",
            db_pg.as_uuid(guia_id), db_pg.as_uuid(current_user["company_id"]),
        )
        if not existe:
            raise HTTPException(status_code=404, detail="Guía no encontrada")

        # companies sigue en Mongo: lectura cruzada, correcta mientras esa
        # tabla no haya cortado.
        company = await _empresa_pg(current_user["company_id"])
        sunat_config = company.get("sunat_config", {}) if company else {}
        if not sunat_config.get("api_token"):
            raise HTTPException(
                status_code=400,
                detail="Configuración SUNAT no encontrada. Configure su token API en Configuración > SUNAT"
            )

        # TODO: Integrate with SUNAT API (Nubefact, PSE, or direct SUNAT)
        await conn.execute(
            "update guias_transportista set status = $1::guia_transportista_status, "
            "sunat_response = $2, updated_at = now() where id = $3 and company_id = $4",
            GuiaTransportistaStatus.EMITIDA.value,
            {"message": "Pendiente integración SUNAT API"},
            db_pg.as_uuid(guia_id),
            db_pg.as_uuid(current_user["company_id"]),
        )

    return {"message": "Guía marcada como emitida. Configure la API SUNAT para emisión electrónica."}
# --- Factura Endpoints ---
@api_router.get("/facturas")
async def get_facturas(
    trip_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(trip_id, "trip_id = $?", db_pg.as_uuid(trip_id))
    if status:
        if status not in [e.value for e in FacturaStatus]:
            return []
        f.agregar("status = $?::factura_status", status)
    async with db_pg.tx(current_user) as conn:
        rows = await conn.fetch(
            "select * from facturas where " + f.where
            + " order by created_at desc limit 500",
            *f.values,
        )
    return db_pg.rows_to_api(rows)

@api_router.post("/facturas")
async def create_factura(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))):

    serie = request.get("serie", "F001")
    items = request.get("items", [])
    subtotal = sum(i.get("cantidad", 0) * i.get("precio_unitario", 0) for i in items)
    igv = round(subtotal * 0.18, 2)
    total = round(subtotal + igv, 2)

    async with db_pg.tx(current_user) as conn:
        # Correlativo por (empresa, serie), misma carrera que con Mongo.
        ultimo = await conn.fetchval(
            "select max(numero) from facturas where company_id = $1 and serie = $2",
            db_pg.as_uuid(current_user["company_id"]), serie,
        )
        next_num = (ultimo or 0) + 1

        factura = Factura(
            company_id=current_user["company_id"],
            trip_id=request.get("trip_id"),
            serie=serie,
            numero=next_num,
            fecha_emision=request.get("fecha_emision", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            cliente_ruc=request.get("cliente_ruc"),
            cliente_razon_social=request.get("cliente_razon_social"),
            cliente_direccion=request.get("cliente_direccion"),
            items=items,
            subtotal=subtotal,
            igv=igv,
            total=total,
            created_by=current_user["id"]
        )
        sql, values = db_pg.build_insert(
            "facturas", FACTURA_COLS, _modelo_a_fila(factura.model_dump())
        )
        await conn.execute(sql, *values)

    return {"id": factura.id, "numero": f"{factura.serie}-{next_num:08d}", "total": total, "message": "Factura creada"}

@api_router.post("/facturas/{factura_id}/emit")
async def emit_factura_sunat(factura_id: str, current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))):
    """Emit factura to SUNAT"""
    async with db_pg.tx(current_user) as conn:
        existe = await conn.fetchval(
            "select 1 from facturas where id = $1 and company_id = $2",
            db_pg.as_uuid(factura_id), db_pg.as_uuid(current_user["company_id"]),
        )
        if not existe:
            raise HTTPException(status_code=404, detail="Factura no encontrada")

        # companies sigue en Mongo: lectura cruzada.
        company = await _empresa_pg(current_user["company_id"])
        sunat_config = company.get("sunat_config", {}) if company else {}
        if not sunat_config.get("api_token"):
            raise HTTPException(
                status_code=400,
                detail="Configuración SUNAT no encontrada. Configure su token API en Configuración > SUNAT"
            )

        # TODO: SUNAT API integration placeholder
        await conn.execute(
            "update facturas set status = $1::factura_status, "
            "sunat_response = $2, updated_at = now() where id = $3 and company_id = $4",
            FacturaStatus.EMITIDA.value,
            {"message": "Pendiente integración SUNAT API"},
            db_pg.as_uuid(factura_id),
            db_pg.as_uuid(current_user["company_id"]),
        )

    return {"message": "Factura marcada como emitida. Configure la API SUNAT para emisión electrónica."}
# --- SUNAT Config ---
@api_router.get("/config/sunat")
async def get_sunat_config(current_user: dict = Depends(require_roles("owner", "admin"))):
    company = await _empresa_pg(current_user["company_id"])
    config = company.get("sunat_config", {}) if company else {}
    # Mask token
    if config.get("api_token"):
        config["api_token_masked"] = config["api_token"][:8] + "..." + config["api_token"][-4:]
        del config["api_token"]
    return config

@api_router.put("/config/sunat")
async def update_sunat_config(request: dict = Body(...), current_user: dict = Depends(require_roles("owner", "admin"))):

    sunat_config = {
        "ruc": request.get("ruc"),
        "razon_social": request.get("razon_social"),
        "api_provider": request.get("api_provider", "nubefact"),  # nubefact, efact, sunat
        "api_url": request.get("api_url"),
        "api_token": request.get("api_token"),
        "certificate_password": request.get("certificate_password"),
        "production": request.get("production", False),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    async with db_pg.tx(current_user) as conn:
        await conn.execute(
            "update companies set sunat_config = $1, updated_at = now() where id = $2",
            sunat_config, db_pg.as_uuid(current_user["company_id"]),
        )

    return {"message": "Configuración SUNAT actualizada"}

# ============== REPORTES AVANZADOS (P3) ==============
def _date_range_query(field: str, date_from: Optional[str], date_to: Optional[str]) -> dict:
    """Construye un sub-query de rango de fechas (ISO strings) para `field`."""
    cond = {}
    if date_from:
        cond["$gte"] = date_from
    if date_to:
        # inclusivo hasta el fin del día si sólo viene la fecha
        cond["$lte"] = date_to if len(date_to) > 10 else date_to + "T23:59:59.999999"
    return {field: cond} if cond else {}


def _rango_fechas_pg(f, columna, desde, hasta):
    """Version SQL de _date_range_query: agrega el rango a un db_pg.Filtros.

    Conserva la misma regla que la de Mongo: si `hasta` viene como fecha sola
    (10 caracteres) el limite se estira al final de ese dia, para que el filtro
    siga siendo inclusivo.
    """
    if desde:
        f.agregar(columna + " >= $?", db_pg.as_ts(desde))
    if hasta:
        limite = hasta if len(hasta) > 10 else hasta + "T23:59:59.999999"
        f.agregar(columna + " <= $?", db_pg.as_ts(limite))
    return f


@api_router.get("/reports/cost-per-km")
async def get_cost_per_km_report(
    vehicle_id: Optional[str] = None,
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Costo por km por unidad: combustible + llantas + mantenimiento / km recorridos."""
    company_id = current_user["company_id"]

    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
    f.si(vehicle_id, "id = $?", db_pg.as_uuid(vehicle_id))
    async with db_pg.tx({"company_id": company_id}) as conn:
        vehicles = db_pg.rows_to_api(await conn.fetch(
            "select * from vehicles where " + f.where + " limit 1000", *f.values
        ))

    rows = []
    tot_fuel = tot_tires = tot_maint = tot_km = 0.0
    for v in vehicles:
        vid = v["id"]

        # Combustible
        f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
        f.agregar("vehicle_id = $?", db_pg.as_uuid(vid))
        _rango_fechas_pg(f, "load_date", from_, to)
        async with db_pg.tx(current_user) as conn:
            loads = db_pg.rows_to_api(await conn.fetch(
                "select * from fuel_loads where " + f.where + " limit 2000", *f.values
            ))
        fuel = sum(l.get("total_amount", 0) or 0 for l in loads)

        # Llantas (montajes en el periodo -> costo de compra de la llanta)
        # El JOIN reemplaza al bucle que pedia la llanta de cada montaje una
        # por una: era una consulta por montaje, hasta 2000 por vehiculo.
        fm = db_pg.Filtros("m.company_id = $?", db_pg.as_uuid(company_id))
        fm.agregar("m.vehicle_id = $?", db_pg.as_uuid(vid))
        fm.si(from_, "m.mount_date >= $?", db_pg.as_ts(from_))
        fm.si(to, "m.mount_date <= $?", db_pg.as_ts(to))

        fw = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
        fw.agregar("vehicle_id = $?", db_pg.as_uuid(vid))
        fw.si(from_, "created_at >= $?", db_pg.as_ts(from_))
        fw.si(to, "created_at <= $?", db_pg.as_ts(to))

        async with db_pg.tx({"company_id": company_id}) as conn:
            tires = float(await conn.fetchval(
                "select coalesce(sum(t.purchase_cost), 0) from tire_mounts m "
                "join tires t on t.id = m.tire_id and t.company_id = m.company_id "
                "where " + fm.where, *fm.values
            ) or 0)

            # Mantenimiento
            work_orders = db_pg.rows_to_api(await conn.fetch(
                "select * from work_orders where " + fw.where + " limit 2000",
                *fw.values
            ))
        maintenance = sum(wo.get("total_cost", 0) or 0 for wo in work_orders)

        # Km recorridos (rango de odómetro observado en el periodo)
        readings = [l.get("odometer") for l in loads if l.get("odometer")]
        for wo in work_orders:
            if wo.get("odometer_at_service"):
                readings.append(wo["odometer_at_service"])
        f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
        f.agregar("(tracto_id = $? or carreta_id = $?)", db_pg.as_uuid(vid))
        _rango_fechas_pg(f, "scheduled_date", from_, to)
        async with db_pg.tx(current_user) as conn:
            trips = db_pg.rows_to_api(await conn.fetch(
                "select km_start, km_end from trips where " + f.where + " limit 2000",
                *f.values,
            ))
        for t in trips:
            if t.get("km_start"):
                readings.append(t["km_start"])
            if t.get("km_end"):
                readings.append(t["km_end"])
        readings = [r for r in readings if isinstance(r, (int, float))]
        km = (max(readings) - min(readings)) if len(readings) >= 2 else 0
        km = km if km > 0 else 0

        total = fuel + tires + maintenance
        rows.append({
            "vehicle_id": vid,
            "plate": v.get("plate"),
            "fuel": round(fuel, 2),
            "tires": round(tires, 2),
            "maintenance": round(maintenance, 2),
            "total": round(total, 2),
            "km": km,
            "cost_per_km": round(total / km, 4) if km > 0 else 0
        })
        tot_fuel += fuel
        tot_tires += tires
        tot_maint += maintenance
        tot_km += km

    grand_total = tot_fuel + tot_tires + tot_maint
    return {
        "rows": rows,
        "totals": {
            "fuel": round(tot_fuel, 2),
            "tires": round(tot_tires, 2),
            "maintenance": round(tot_maint, 2),
            "total": round(grand_total, 2),
            "km": tot_km,
            "cost_per_km": round(grand_total / tot_km, 4) if tot_km > 0 else 0
        }
    }


@api_router.get("/reports/documents-expiring")
async def get_documents_expiring_report(
    days: int = 90,
    current_user: dict = Depends(get_current_user)
):
    """Documentos vencidos y por vencer, ordenados por days_remaining."""
    company_id = current_user["company_id"]
    now = datetime.now(timezone.utc)

    async with db_pg.tx(current_user) as conn:
        documents = db_pg.rows_to_api(await conn.fetch(
            "select * from documents where company_id = $1 "
            "and expiry_date is not null limit 5000",
            db_pg.as_uuid(company_id),
        ))

    dt_ids = list({d.get("document_type_id") for d in documents if d.get("document_type_id")})
    doc_types = {}
    if dt_ids:
        async with db_pg.tx(current_user) as conn:
            filas = db_pg.rows_to_api(await conn.fetch(
                "select * from document_types where company_id = $1 "
                "and id = any($2::uuid[]) limit 1000",
                db_pg.as_uuid(company_id),
                [u for u in (db_pg.as_uuid(x) for x in dt_ids) if u],
            ))
        for dt in filas:
            doc_types[dt["id"]] = dt

    rows = []
    for d in documents:
        expiry = d.get("expiry_date")
        if not expiry:
            continue
        exp = expiry
        if isinstance(exp, str):
            try:
                exp = datetime.fromisoformat(exp.replace("Z", "+00:00"))
            except ValueError:
                continue
        days_remaining = (exp - now).days
        dt = doc_types.get(d.get("document_type_id"), {})

        # Entidad + regla Revisión Técnica (no aplica a unidades <= 4 años)
        entity = None
        etype = d.get("entity_type")
        if etype == "vehicle":
            entity = await _vehiculo_pg(company_id, d.get("entity_id"))
        else:
            entity = await _usuario_pg(company_id, d.get("entity_id"))
        if _revision_tecnica_no_aplica(dt, entity, etype):
            continue

        if days_remaining > days:
            continue

        entity_name = (entity or {}).get("plate") or (entity or {}).get("name") or "N/A"
        status = "vencido" if days_remaining <= 0 else ("por_vencer" if days_remaining <= 30 else "vigente")
        rows.append({
            "entity_type": etype,
            "entity_name": entity_name,
            "document_type": dt.get("name", "Documento"),
            "expiry_date": exp.isoformat() if isinstance(exp, datetime) else expiry,
            "days_remaining": days_remaining,
            "status": status
        })

    rows.sort(key=lambda r: r["days_remaining"])
    return {"rows": rows}


@api_router.get("/reports/viaticos")
async def get_viaticos_report(
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    driver_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Consolidado de viáticos por chofer: presupuesto, gastado, saldo, viajes."""
    company_id = current_user["company_id"]

    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
    f.si(driver_id, "driver_id = $?", db_pg.as_uuid(driver_id))
    _rango_fechas_pg(f, "scheduled_date", from_, to)
    async with db_pg.tx(current_user) as conn:
        trips = db_pg.rows_to_api(await conn.fetch(
            "select * from trips where " + f.where + " limit 5000", *f.values,
        ))

    # Gasto real por viaje, en UNA consulta agrupada. La version Mongo hacia una
    # consulta por viaje dentro del bucle: con 5000 viajes eran 5000 viajes a la
    # base para sumar unos pocos importes.
    gasto_por_viaje = {}
    ids_viajes = [u for u in (db_pg.as_uuid(t.get("id")) for t in trips) if u]
    if ids_viajes:
        async with db_pg.tx(current_user) as conn:
            filas = await conn.fetch(
                "select trip_id, coalesce(sum(amount), 0) as total from trip_expenses "
                "where company_id = $1 and trip_id = any($2::uuid[]) group by trip_id",
                db_pg.as_uuid(company_id), ids_viajes,
            )
        gasto_por_viaje = {str(f["trip_id"]): float(f["total"] or 0) for f in filas}

    by_driver = {}
    for t in trips:
        did = t.get("driver_id") or "unknown"
        if did not in by_driver:
            by_driver[did] = {"driver_id": did, "driver_name": None, "budget": 0.0, "spent": 0.0, "remaining": 0.0, "trips": 0}
        by_driver[did]["budget"] += t.get("viatico_budget", 0) or 0
        by_driver[did]["trips"] += 1
        by_driver[did]["spent"] += gasto_por_viaje.get(t.get("id"), 0.0)

    real_ids = [d for d in by_driver if d and d != "unknown"]
    name_map = {}
    if real_ids:
        async with db_pg.tx({"company_id": company_id}) as conn:
            filas = db_pg.rows_to_api(await conn.fetch(
                "select id, name from users where company_id = $1 and id = any($2::uuid[])",
                db_pg.as_uuid(company_id), [db_pg.as_uuid(x) for x in real_ids],
            ))
        for u in filas:
            name_map[u["id"]] = u.get("name")

    rows = []
    tot_budget = tot_spent = 0.0
    for did, agg in by_driver.items():
        agg["driver_name"] = name_map.get(did, "N/A" if did == "unknown" else did)
        agg["budget"] = round(agg["budget"], 2)
        agg["spent"] = round(agg["spent"], 2)
        agg["remaining"] = round(agg["budget"] - agg["spent"], 2)
        tot_budget += agg["budget"]
        tot_spent += agg["spent"]
        rows.append(agg)

    rows.sort(key=lambda r: r["spent"], reverse=True)
    return {
        "rows": rows,
        "totals": {
            "budget": round(tot_budget, 2),
            "spent": round(tot_spent, 2),
            "remaining": round(tot_budget - tot_spent, 2)
        }
    }


# ============== UNIDADES (agrupación tracto + carreta + chofer + EPP) ==============
class Unit(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    tracto_id: str
    carreta_id: Optional[str] = None
    driver_id: Optional[str] = None
    status: str = "activa"
    epp_items: List[Dict[str, Any]] = Field(default_factory=list)
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None


async def _resolve_unit(company_id: str, unit: dict) -> dict:
    """Enriquece una unidad con placas y nombre de chofer."""
    out = serialize_doc(unit)
    tracto = await _vehiculo_pg(company_id, unit.get("tracto_id"))
    out["tracto_plate"] = (tracto or {}).get("plate")
    if unit.get("carreta_id"):
        carreta = await _vehiculo_pg(company_id, unit.get("carreta_id"))
        out["carreta_plate"] = (carreta or {}).get("plate")
    else:
        out["carreta_plate"] = None
    if unit.get("driver_id"):
        driver = await _usuario_pg(company_id, unit.get("driver_id"))
        out["driver_name"] = (driver or {}).get("name")
    else:
        out["driver_name"] = None
    return out


@api_router.get("/units")
async def get_units(
    include_inactive: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Lista de unidades con placas y chofer resueltos."""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    if not include_inactive:
        f.crudo("active")
    async with db_pg.tx(current_user) as conn:
        units = db_pg.rows_to_api(await conn.fetch(
            "select * from units where " + f.where
            + " order by created_at desc nulls last limit 1000",
            *f.values,
        ))
    return [await _resolve_unit(current_user["company_id"], u) for u in units]


@api_router.post("/units")
async def create_unit(
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "flota"))
):
    """Crea una unidad (capa aditiva sobre vehículos/chofer)."""
    company_id = current_user["company_id"]
    tracto_id = request.get("tracto_id")
    if not tracto_id:
        raise HTTPException(status_code=400, detail="tracto_id es obligatorio")

    tracto = await _vehiculo_pg(company_id, tracto_id)
    if not tracto:
        raise HTTPException(status_code=404, detail="Tracto no encontrado")
    if request.get("carreta_id"):
        carreta = await _vehiculo_pg(company_id, request["carreta_id"])
        if not carreta:
            raise HTTPException(status_code=404, detail="Carreta no encontrada")

    unit = Unit(
        company_id=company_id,
        tracto_id=tracto_id,
        carreta_id=request.get("carreta_id"),
        driver_id=request.get("driver_id"),
        status=request.get("status", "activa"),
        epp_items=request.get("epp_items", []) or [],
        created_by=current_user["id"]
    )
    doc = unit.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["updated_at"] = doc["updated_at"].isoformat()
    sql, values = db_pg.build_insert("units", UNIT_COLS, doc)
    async with db_pg.tx({"company_id": company_id}) as conn:
        await conn.execute(sql, *values)

    # Conveniencia: sincroniza el chofer asignado del tracto (no rompe el coupling existente)
    if request.get("driver_id"):
        await _actualizar_vehiculo(
            company_id, tracto_id, {"assigned_driver_id": request["driver_id"]}
        )

    return await _resolve_unit(company_id, doc)


@api_router.put("/units/{unit_id}")
async def update_unit(
    unit_id: str,
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "flota"))
):
    """Actualiza una unidad."""
    company_id = current_user["company_id"]
    async with db_pg.tx({"company_id": company_id}) as conn:
        existing = db_pg.to_api(await conn.fetchrow(
            "select * from units where id = $1 and company_id = $2",
            db_pg.as_uuid(unit_id), db_pg.as_uuid(company_id),
        ))
    if not existing:
        raise HTTPException(status_code=404, detail="Unidad no encontrada")

    update_data = {}
    for field in ["tracto_id", "carreta_id", "driver_id", "status", "epp_items", "active"]:
        if field in request:
            update_data[field] = request[field]
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    update_data["id"] = unit_id
    update_data["company_id"] = company_id
    sql, values = db_pg.build_update("units", UNIT_COLS, update_data, ["id", "company_id"])
    async with db_pg.tx({"company_id": company_id}) as conn:
        if sql:
            await conn.execute(sql, *values)

    if request.get("driver_id"):
        tracto_id = request.get("tracto_id", existing.get("tracto_id"))
        await _actualizar_vehiculo(
            company_id, tracto_id, {"assigned_driver_id": request["driver_id"]}
        )

    async with db_pg.tx({"company_id": company_id}) as conn:
        updated = db_pg.to_api(await conn.fetchrow(
            "select * from units where id = $1 and company_id = $2",
            db_pg.as_uuid(unit_id), db_pg.as_uuid(company_id),
        ))
    return await _resolve_unit(company_id, updated)


@api_router.delete("/units/{unit_id}")
async def delete_unit(
    unit_id: str,
    current_user: dict = Depends(require_roles("owner", "admin", "flota"))
):
    """Soft-delete de una unidad (active=False)."""
    company_id = current_user["company_id"]
    datos = {
        "active": False,
        "status": "inactiva",
        "updated_at": datetime.now(timezone.utc),
        "id": unit_id,
        "company_id": company_id,
    }
    sql, values = db_pg.build_update("units", UNIT_COLS, datos, ["id", "company_id"])
    async with db_pg.tx({"company_id": company_id}) as conn:
        # returning id ocupa el lugar de matched_count: None = no existia.
        existia = await conn.fetchval(sql + " returning id", *values)
    if existia is None:
        raise HTTPException(status_code=404, detail="Unidad no encontrada")
    return {"message": "Unidad eliminada"}


# ============== DETRACCIONES (SPOT) Y CAJA ==============
# Módulos financieros: detracción del IGV (SPOT) y caja chica / kardex.

def _modelo_a_fila(doc: dict) -> dict:
    """model_dump() -> dict listo para db_pg.build_insert.

    Los datetime se dejan como datetime en vez de pasarlos a string:
    Postgres los quiere como datetime, y convertirlos para que db_pg los
    vuelva a parsear seria trabajo de ida y vuelta. Los Enum si se bajan a su
    .value, que es lo que espera el cast a enum de Postgres.
    """
    out = {}
    for key, value in doc.items():
        out[key] = value.value if isinstance(value, Enum) else value
    return out


def _to_float(value, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


# --- Detracciones ---
class DetraccionStatus(str, Enum):
    PENDIENTE = "pendiente"
    DEPOSITADA = "depositada"
    ANULADA = "anulada"


class Detraccion(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    factura_id: Optional[str] = None
    trip_id: Optional[str] = None
    # Cliente / adquirente que efectúa el depósito
    client_ruc: Optional[str] = None
    client_name: Optional[str] = None
    # Comprobante que origina la detracción
    comprobante_serie: Optional[str] = None
    comprobante_numero: Optional[str] = None
    fecha_emision: Optional[str] = None
    # Importes: amount = base_amount * rate / 100 (calculado en el servidor)
    base_amount: float = 0
    rate: float = DEFAULT_DETRACCION_RATE
    amount: float = 0
    codigo_bien_servicio: str = DEFAULT_DETRACCION_CODIGO
    # Constancia de depósito (Banco de la Nación)
    constancia_number: Optional[str] = None
    deposit_date: Optional[str] = None
    status: DetraccionStatus = DetraccionStatus.PENDIENTE
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None


def _calc_detraccion_amount(base_amount, rate) -> float:
    """Fórmula SPOT: importe de detracción = base (total del comprobante) * tasa% / 100."""
    return round(_to_float(base_amount) * _to_float(rate) / 100.0, 2)


async def _detraccion_defaults(company_id: str) -> Dict[str, float]:
    """Tasa y monto mínimo de detracción desde la config de la empresa."""
    config = await _company_config(company_id)
    return {
        "rate": _to_float(config.get("detraccion_rate"), DEFAULT_DETRACCION_RATE) or DEFAULT_DETRACCION_RATE,
        "min_amount": _to_float(config.get("detraccion_min_amount"), DEFAULT_DETRACCION_MIN_AMOUNT),
        "codigo": config.get("detraccion_codigo_bien_servicio") or DEFAULT_DETRACCION_CODIGO,
    }


@api_router.get("/detracciones")
async def get_detracciones(
    status: Optional[str] = None,
    client_ruc: Optional[str] = None,
    factura_id: Optional[str] = None,
    trip_id: Optional[str] = None,
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista de detracciones de la empresa (fecha de emisión desc)."""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    if status:
        if status not in [e.value for e in DetraccionStatus]:
            return []
        f.agregar("status = $?::detraccion_status", status)
    f.si(client_ruc, "client_ruc = $?", client_ruc)
    f.si(factura_id, "factura_id = $?", db_pg.as_uuid(factura_id))
    f.si(trip_id, "trip_id = $?", db_pg.as_uuid(trip_id))
    # fecha_emision es texto YYYY-MM-DD, igual que en Mongo, asi que el rango
    # se compara como cadena y da el mismo resultado. Ya no hace falta el
    # T23:59:59 que _date_range_query le pegaba al limite superior.
    f.si(from_, "fecha_emision >= $?", from_)
    f.si(to, "fecha_emision <= $?", to)

    async with db_pg.tx(current_user) as conn:
        rows = await conn.fetch(
            "select * from detracciones where " + f.where
            + " order by fecha_emision desc, created_at desc limit 1000",
            *f.values,
        )
    return db_pg.rows_to_api(rows)


@api_router.get("/detracciones/summary")
async def get_detracciones_summary(
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    client_ruc: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Totales del periodo: pendientes, depositadas, anuladas y total.

    La suma la hace Postgres agrupando por estado, en vez de traer hasta 5000
    filas al proceso para sumarlas en Python."""
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(current_user["company_id"]))
    f.si(client_ruc, "client_ruc = $?", client_ruc)
    f.si(from_, "fecha_emision >= $?", from_)
    f.si(to, "fecha_emision <= $?", to)

    async with db_pg.tx(current_user) as conn:
        rows = await conn.fetch(
            "select status::text as status, count(*) as n, "
            "coalesce(sum(amount), 0) as monto, "
            "coalesce(sum(base_amount), 0) as base "
            "from detracciones where " + f.where + " group by status",
            *f.values,
        )

    buckets = {
        "pendientes": {"count": 0, "amount": 0.0},
        "depositadas": {"count": 0, "amount": 0.0},
        "anuladas": {"count": 0, "amount": 0.0},
    }
    key_by_status = {
        DetraccionStatus.PENDIENTE.value: "pendientes",
        DetraccionStatus.DEPOSITADA.value: "depositadas",
        DetraccionStatus.ANULADA.value: "anuladas",
    }

    total_count = 0
    total_amount = 0.0
    total_base = 0.0
    for r in rows:
        key = key_by_status.get(r["status"])
        if key:
            buckets[key]["count"] = r["n"]
            buckets[key]["amount"] = round(_to_float(r["monto"]), 2)
        # El total del periodo no considera las anuladas
        if r["status"] != DetraccionStatus.ANULADA.value:
            total_count += r["n"]
            total_amount += _to_float(r["monto"])
            total_base += _to_float(r["base"])

    return {
        "from": from_,
        "to": to,
        "pendientes": buckets["pendientes"],
        "depositadas": buckets["depositadas"],
        "anuladas": buckets["anuladas"],
        "total": {
            "count": total_count,
            "amount": round(total_amount, 2),
            "base_amount": round(total_base, 2),
        },
    }

@api_router.get("/detracciones/{detraccion_id}")
async def get_detraccion(detraccion_id: str, current_user: dict = Depends(get_current_user)):
    async with db_pg.tx(current_user) as conn:
        row = await conn.fetchrow(
            "select * from detracciones where id = $1 and company_id = $2",
            db_pg.as_uuid(detraccion_id), db_pg.as_uuid(current_user["company_id"]),
        )
    if not row:
        raise HTTPException(status_code=404, detail="Detracción no encontrada")
    return db_pg.to_api(row)


@api_router.post("/detracciones")
async def create_detraccion(
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))
):
    """Crea una detracción. El importe se calcula SIEMPRE en el servidor."""
    company_id = current_user["company_id"]
    defaults = await _detraccion_defaults(company_id)

    base_amount = _to_float(request.get("base_amount"))
    if base_amount <= 0:
        raise HTTPException(status_code=400, detail="El importe base debe ser mayor a 0")

    rate = _to_float(request.get("rate"), defaults["rate"]) if request.get("rate") is not None else defaults["rate"]
    if rate <= 0 or rate > 100:
        raise HTTPException(status_code=400, detail="La tasa de detracción debe estar entre 0 y 100")

    # Con Mongo un status invalido se guardaba tal cual; el enum de Postgres lo
    # rechazaria con un 500. Se valida antes para responder 400, igual que ya
    # hacia el PUT de este mismo recurso.
    estado = request.get("status") or DetraccionStatus.PENDIENTE.value
    if estado not in [e.value for e in DetraccionStatus]:
        raise HTTPException(status_code=400, detail="Estado de detracción inválido")

    detraccion = Detraccion(
        company_id=company_id,
        factura_id=request.get("factura_id"),
        trip_id=request.get("trip_id"),
        client_ruc=request.get("client_ruc"),
        client_name=request.get("client_name"),
        comprobante_serie=request.get("comprobante_serie"),
        comprobante_numero=request.get("comprobante_numero"),
        fecha_emision=request.get("fecha_emision") or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        base_amount=round(base_amount, 2),
        rate=rate,
        amount=_calc_detraccion_amount(base_amount, rate),
        codigo_bien_servicio=request.get("codigo_bien_servicio") or defaults["codigo"],
        constancia_number=request.get("constancia_number"),
        deposit_date=request.get("deposit_date"),
        status=estado,
        notes=request.get("notes"),
        created_by=current_user["id"],
    )

    sql, values = db_pg.build_insert(
        "detracciones", DETRACCION_COLS, _modelo_a_fila(detraccion.model_dump())
    )
    async with db_pg.tx(current_user) as conn:
        await conn.execute(sql, *values)
        row = await conn.fetchrow(
            "select * from detracciones where id = $1", db_pg.as_uuid(detraccion.id)
        )
    return db_pg.to_api(row)

@api_router.put("/detracciones/{detraccion_id}")
async def update_detraccion(
    detraccion_id: str,
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))
):
    """Actualiza una detracción y recalcula el importe si cambia la base o la tasa."""
    company_id = current_user["company_id"]
    async with db_pg.tx(current_user) as conn:
        row = await conn.fetchrow(
            "select * from detracciones where id = $1 and company_id = $2",
            db_pg.as_uuid(detraccion_id), db_pg.as_uuid(company_id),
        )
        if not row:
            raise HTTPException(status_code=404, detail="Detracción no encontrada")
        existing = db_pg.to_api(row)

        update_data = {}
        for field in ["factura_id", "trip_id", "client_ruc", "client_name", "comprobante_serie",
                      "comprobante_numero", "fecha_emision", "codigo_bien_servicio",
                      "constancia_number", "deposit_date", "notes"]:
            if field in request:
                update_data[field] = request[field]

        if "status" in request and request["status"]:
            if request["status"] not in [e.value for e in DetraccionStatus]:
                raise HTTPException(status_code=400, detail="Estado de detracción inválido")
            update_data["status"] = request["status"]

        base_amount = _to_float(request["base_amount"]) if "base_amount" in request else _to_float(existing.get("base_amount"))
        rate = _to_float(request["rate"]) if "rate" in request else _to_float(existing.get("rate"), DEFAULT_DETRACCION_RATE)

        if "base_amount" in request:
            if base_amount <= 0:
                raise HTTPException(status_code=400, detail="El importe base debe ser mayor a 0")
            update_data["base_amount"] = round(base_amount, 2)
        if "rate" in request:
            if rate <= 0 or rate > 100:
                raise HTTPException(status_code=400, detail="La tasa de detracción debe estar entre 0 y 100")
            update_data["rate"] = rate
        if "base_amount" in request or "rate" in request:
            update_data["amount"] = _calc_detraccion_amount(base_amount, rate)

        update_data["updated_at"] = datetime.now(timezone.utc)
        update_data["id"] = detraccion_id
        update_data["company_id"] = company_id
        sql, values = db_pg.build_update(
            "detracciones", DETRACCION_COLS, update_data, ["id", "company_id"]
        )
        if sql:
            await conn.execute(sql, *values)

        actualizado = await conn.fetchrow(
            "select * from detracciones where id = $1 and company_id = $2",
            db_pg.as_uuid(detraccion_id), db_pg.as_uuid(company_id),
        )
    return db_pg.to_api(actualizado)


@api_router.post("/detracciones/{detraccion_id}/register-deposit")
async def register_detraccion_deposit(
    detraccion_id: str,
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))
):
    """Registra la constancia de depósito y marca la detracción como depositada."""
    company_id = current_user["company_id"]

    constancia_number = (request.get("constancia_number") or "").strip()
    if not constancia_number:
        raise HTTPException(status_code=400, detail="El número de constancia es obligatorio")

    async with db_pg.tx(current_user) as conn:
        estado_actual = await conn.fetchval(
            "select status::text from detracciones where id = $1 and company_id = $2",
            db_pg.as_uuid(detraccion_id), db_pg.as_uuid(company_id),
        )
        if estado_actual is None:
            raise HTTPException(status_code=404, detail="Detracción no encontrada")
        if estado_actual == DetraccionStatus.ANULADA.value:
            raise HTTPException(status_code=400, detail="La detracción está anulada")

        update_data = {
            "constancia_number": constancia_number,
            "deposit_date": request.get("deposit_date") or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "status": DetraccionStatus.DEPOSITADA.value,
            "updated_at": datetime.now(timezone.utc),
            "id": detraccion_id,
            "company_id": company_id,
        }
        if request.get("notes"):
            update_data["notes"] = request["notes"]

        sql, values = db_pg.build_update(
            "detracciones", DETRACCION_COLS, update_data, ["id", "company_id"]
        )
        await conn.execute(sql, *values)
        actualizado = await conn.fetchrow(
            "select * from detracciones where id = $1 and company_id = $2",
            db_pg.as_uuid(detraccion_id), db_pg.as_uuid(company_id),
        )
    return {"message": "Depósito de detracción registrado", "detraccion": db_pg.to_api(actualizado)}

@api_router.delete("/detracciones/{detraccion_id}")
async def delete_detraccion(
    detraccion_id: str,
    current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))
):
    """Anula la detracción (soft-delete): se conserva la trazabilidad contable."""
    company_id = current_user["company_id"]
    ahora = datetime.now(timezone.utc)
    async with db_pg.tx(current_user) as conn:
        # UPDATE ... RETURNING: una sola ida a la base para anular y saber si
        # la fila existia, en vez de un SELECT previo.
        anulada = await conn.fetchval(
            "update detracciones set status = $1::detraccion_status, "
            "anulada_at = $2, anulada_by = $3, updated_at = $2 "
            "where id = $4 and company_id = $5 returning id",
            DetraccionStatus.ANULADA.value,
            ahora,
            db_pg.as_uuid(current_user["id"]),
            db_pg.as_uuid(detraccion_id),
            db_pg.as_uuid(company_id),
        )
    if not anulada:
        raise HTTPException(status_code=404, detail="Detracción no encontrada")
    return {"message": "Detracción anulada"}


@api_router.post("/detracciones/from-factura/{factura_id}")
async def create_detraccion_from_factura(
    factura_id: str,
    request: dict = Body(default={}),
    current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))
):
    """Genera la detracción a partir de una factura existente.
    No aplica si el total de la factura es menor al mínimo configurado."""
    company_id = current_user["company_id"]
    defaults = await _detraccion_defaults(company_id)

    # Factura y detraccion cortaron juntas, asi que leer la factura, comprobar
    # que no haya una detraccion previa y crear la nueva ocurre todo dentro de
    # la misma transaccion: dos llamadas simultaneas ya no pueden generar dos
    # detracciones para la misma factura.
    async with db_pg.tx(current_user) as conn:
        factura = await conn.fetchrow(
            "select * from facturas where id = $1 and company_id = $2",
            db_pg.as_uuid(factura_id), db_pg.as_uuid(company_id),
        )
        if not factura:
            raise HTTPException(status_code=404, detail="Factura no encontrada")
        factura = db_pg.to_api(factura)

        existente = await conn.fetchrow(
            "select * from detracciones where company_id = $1 and factura_id = $2 "
            "and status <> $3::detraccion_status",
            db_pg.as_uuid(company_id), db_pg.as_uuid(factura_id),
            DetraccionStatus.ANULADA.value,
        )
        if existente:
            return {
                "applies": True,
                "created": False,
                "message": "La factura ya tiene una detracción registrada",
                "detraccion": db_pg.to_api(existente),
            }

        base_amount = _to_float(factura.get("total"))
        min_amount = defaults["min_amount"]

        if base_amount <= 0:
            raise HTTPException(status_code=400, detail="La factura no tiene importe total")
        if base_amount < min_amount:
            return {
                "applies": False,
                "created": False,
                "base_amount": round(base_amount, 2),
                "min_amount": min_amount,
                "message": (f"No aplica detracción: el total de la factura (S/ {round(base_amount, 2)}) "
                            f"es menor al mínimo de S/ {min_amount}"),
            }

        rate = _to_float(request.get("rate"), defaults["rate"]) if request.get("rate") is not None else defaults["rate"]

        numero = factura.get("numero")
        detraccion = Detraccion(
            company_id=company_id,
            factura_id=factura_id,
            trip_id=factura.get("trip_id"),
            client_ruc=factura.get("cliente_ruc"),
            client_name=factura.get("cliente_razon_social"),
            comprobante_serie=factura.get("serie"),
            comprobante_numero=f"{numero:08d}" if isinstance(numero, int) else (str(numero) if numero else None),
            fecha_emision=factura.get("fecha_emision") or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            base_amount=round(base_amount, 2),
            rate=rate,
            amount=_calc_detraccion_amount(base_amount, rate),
            codigo_bien_servicio=request.get("codigo_bien_servicio") or defaults["codigo"],
            notes=request.get("notes"),
            created_by=current_user["id"],
        )

        sql, values = db_pg.build_insert(
            "detracciones", DETRACCION_COLS, _modelo_a_fila(detraccion.model_dump())
        )
        await conn.execute(sql, *values)
        creada = await conn.fetchrow(
            "select * from detracciones where id = $1", db_pg.as_uuid(detraccion.id)
        )

    return {
        "applies": True,
        "created": True,
        "message": "Detracción generada desde la factura",
        "detraccion": db_pg.to_api(creada),
    }

# --- Caja (ingresos / egresos, kardex, reportes por rubro) ---
class CashMovementType(str, Enum):
    INGRESO = "ingreso"
    EGRESO = "egreso"


class CashPaymentMethod(str, Enum):
    EFECTIVO = "efectivo"
    TRANSFERENCIA = "transferencia"
    DEPOSITO = "deposito"
    YAPE_PLIN = "yape_plin"
    OTRO = "otro"


# Rubros sugeridos (el campo `category` acepta texto libre)
CASHBOX_SUGGESTED_CATEGORIES = [
    "combustible", "peajes", "viaticos", "mantenimiento",
    "planilla", "cobranza", "aporte", "retiro", "otros",
]


class CashMovement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    movement_number: Optional[str] = None  # correlativo por empresa: MOV-00001
    date: Optional[str] = None
    type: CashMovementType = CashMovementType.EGRESO
    concept: Optional[str] = None
    category: str = "otros"
    amount: float = 0
    payment_method: CashPaymentMethod = CashPaymentMethod.EFECTIVO
    reference: Optional[str] = None
    trip_id: Optional[str] = None
    vehicle_id: Optional[str] = None
    client_ruc: Optional[str] = None
    supplier: Optional[str] = None
    receipt_url: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None


async def _next_movement_number(conn, company_id: str) -> str:
    """Correlativo por empresa (MOV-00001).

    Cuenta tambien los movimientos eliminados: el correlativo es contable y no
    se reutiliza, igual que antes."""
    ultimo = await conn.fetchval(
        "select movement_number from cash_movements "
        "where company_id = $1 and movement_number is not null "
        "order by movement_number desc limit 1",
        db_pg.as_uuid(company_id),
    )
    next_num = 1
    if ultimo:
        try:
            next_num = int(str(ultimo).split("-")[-1]) + 1
        except (ValueError, IndexError):
            total = await conn.fetchval(
                "select count(*) from cash_movements where company_id = $1",
                db_pg.as_uuid(company_id),
            )
            next_num = total + 1

    # Verificacion defensiva contra duplicados (se conserva del codigo Mongo).
    for _ in range(50):
        candidate = f"MOV-{next_num:05d}"
        choque = await conn.fetchval(
            "select 1 from cash_movements where company_id = $1 and movement_number = $2",
            db_pg.as_uuid(company_id), candidate,
        )
        if not choque:
            return candidate
        next_num += 1
    return f"MOV-{next_num:05d}"


def _cashbox_filtros(company_id: str, from_=None, to=None):
    """Filtro base de caja: la empresa y los movimientos NO eliminados.

    El soft-delete vive en la columna `deleted` (agregada en la migracion 002).
    Cualquier consulta de caja que se olvide de excluirlos devuelve un saldo
    equivocado, asi que el filtro se arma en un solo lugar.

    `date` es una columna de texto YYYY-MM-DD, igual que en Mongo, asi que el
    rango se compara como cadena y da el mismo resultado de siempre.
    """
    f = db_pg.Filtros("company_id = $?", db_pg.as_uuid(company_id))
    f.crudo("not deleted")
    f.si(from_, "date >= $?", from_)
    f.si(to, "date <= $?", to)
    return f


async def _cashbox_saldo_inicial(conn, company_id: str, date_from: Optional[str]) -> float:
    """Saldo acumulado de todos los movimientos anteriores al inicio del rango.

    La suma la hace Postgres; antes se traian hasta 20000 filas al proceso
    solo para sumarlas."""
    if not date_from:
        return 0.0
    f = _cashbox_filtros(company_id)
    f.agregar("date < $?", date_from)
    saldo = await conn.fetchval(
        "select coalesce(sum(case when type = " + chr(39) + "ingreso" + chr(39) + " then amount else -amount end), 0) "
        "from cash_movements where " + f.where,
        *f.values,
    )
    return round(_to_float(saldo), 2)


@api_router.get("/cashbox/categories")
async def get_cashbox_categories(current_user: dict = Depends(get_current_user)):
    """Rubros sugeridos + los ya usados por la empresa."""
    f = _cashbox_filtros(current_user["company_id"])
    async with db_pg.tx(current_user) as conn:
        rows = await conn.fetch(
            "select distinct category from cash_movements where " + f.where,
            *f.values,
        )
    categories = list(CASHBOX_SUGGESTED_CATEGORIES)
    for r in rows:
        c = r["category"]
        if c and c not in categories:
            categories.append(c)
    return {"suggested": CASHBOX_SUGGESTED_CATEGORIES, "categories": categories}


@api_router.get("/cashbox/movements")
async def get_cash_movements(
    type: Optional[str] = None,
    category: Optional[str] = None,
    payment_method: Optional[str] = None,
    trip_id: Optional[str] = None,
    vehicle_id: Optional[str] = None,
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Movimientos de caja (fecha desc)."""
    if type and type not in [t.value for t in CashMovementType]:
        return []
    if payment_method and payment_method not in [p.value for p in CashPaymentMethod]:
        return []

    f = _cashbox_filtros(current_user["company_id"], from_, to)
    f.si(type, "type = $?::cash_movement_type", type)
    f.si(category, "category = $?", category)
    f.si(payment_method, "payment_method = $?::cash_payment_method", payment_method)
    f.si(trip_id, "trip_id = $?", db_pg.as_uuid(trip_id))
    f.si(vehicle_id, "vehicle_id = $?", db_pg.as_uuid(vehicle_id))

    async with db_pg.tx(current_user) as conn:
        rows = await conn.fetch(
            "select * from cash_movements where " + f.where
            + " order by date desc, created_at desc limit 2000",
            *f.values,
        )
    return db_pg.rows_to_api(rows)

@api_router.get("/cashbox/balance")
async def get_cashbox_balance(
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Saldo de caja del periodo: ingresos, egresos, saldo inicial y saldo final."""
    company_id = current_user["company_id"]
    f = _cashbox_filtros(company_id, from_, to)

    async with db_pg.tx(current_user) as conn:
        r = await conn.fetchrow(
            "select count(*) as n, "
            "coalesce(sum(case when type = 'ingreso' then amount else 0 end), 0) as ingresos, "
            "coalesce(sum(case when type <> 'ingreso' then amount else 0 end), 0) as egresos "
            "from cash_movements where " + f.where,
            *f.values,
        )
        saldo_inicial = await _cashbox_saldo_inicial(conn, company_id, from_)

    total_ingresos = round(_to_float(r["ingresos"]), 2)
    total_egresos = round(_to_float(r["egresos"]), 2)
    saldo = round(total_ingresos - total_egresos, 2)
    return {
        "from": from_,
        "to": to,
        "count": r["n"],
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "saldo": saldo,
        "saldo_inicial": saldo_inicial,
        "saldo_final": round(saldo_inicial + saldo, 2),
    }


@api_router.get("/cashbox/kardex")
async def get_cashbox_kardex(
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Kardex clasico: movimientos en orden cronologico ascendente con saldo corriente.

    El saldo corriente se sigue calculando en Python: es un acumulado fila a
    fila que arranca del saldo inicial, y hacerlo aca deja el redondeo por
    linea exactamente igual que antes."""
    company_id = current_user["company_id"]
    f = _cashbox_filtros(company_id, from_, to)

    async with db_pg.tx(current_user) as conn:
        movimientos = await conn.fetch(
            "select * from cash_movements where " + f.where
            + " order by date, created_at limit 20000",
            *f.values,
        )
        saldo_inicial = await _cashbox_saldo_inicial(conn, company_id, from_)

    saldo = saldo_inicial
    total_ingresos = 0.0
    total_egresos = 0.0
    rows = []
    for fila in movimientos:
        m = db_pg.to_api(fila)
        amount = _to_float(m.get("amount"))
        is_ingreso = m.get("type") == CashMovementType.INGRESO.value
        ingreso = round(amount, 2) if is_ingreso else 0.0
        egreso = 0.0 if is_ingreso else round(amount, 2)
        saldo = round(saldo + ingreso - egreso, 2)
        total_ingresos += ingreso
        total_egresos += egreso
        rows.append({
            "id": m.get("id"),
            "date": m.get("date"),
            "movement_number": m.get("movement_number"),
            "concept": m.get("concept"),
            "category": m.get("category"),
            "type": m.get("type"),
            "payment_method": m.get("payment_method"),
            "reference": m.get("reference"),
            "ingreso": ingreso,
            "egreso": egreso,
            "saldo": saldo,
        })

    return {
        "from": from_,
        "to": to,
        "saldo_inicial": saldo_inicial,
        "rows": rows,
        "totals": {
            "ingresos": round(total_ingresos, 2),
            "egresos": round(total_egresos, 2),
            "count": len(rows),
        },
        "saldo_final": saldo,
    }

@api_router.get("/cashbox/report-by-category")
async def get_cashbox_report_by_category(
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Reporte por rubro: ingresos, egresos y neto agrupados por category.

    El agrupado lo hace Postgres. Antes se traian hasta 20000 movimientos al
    proceso para armar el mismo resumen en un diccionario."""
    f = _cashbox_filtros(current_user["company_id"], from_, to)

    async with db_pg.tx(current_user) as conn:
        filas = await conn.fetch(
            "select coalesce(category, 'otros') as category, count(*) as n, "
            "coalesce(sum(case when type = 'ingreso' then amount else 0 end), 0) as ingresos, "
            "coalesce(sum(case when type <> 'ingreso' then amount else 0 end), 0) as egresos "
            "from cash_movements where " + f.where + " group by 1",
            *f.values,
        )

    rows = []
    tot_ingresos = 0.0
    tot_egresos = 0.0
    total_count = 0
    for r in filas:
        ingresos = round(_to_float(r["ingresos"]), 2)
        egresos = round(_to_float(r["egresos"]), 2)
        tot_ingresos += ingresos
        tot_egresos += egresos
        total_count += r["n"]
        rows.append({
            "category": r["category"],
            "ingresos": ingresos,
            "egresos": egresos,
            "neto": round(ingresos - egresos, 2),
            "count": r["n"],
        })
    rows.sort(key=lambda r: (r["egresos"] + r["ingresos"]), reverse=True)

    return {
        "from": from_,
        "to": to,
        "rows": rows,
        "totals": {
            "ingresos": round(tot_ingresos, 2),
            "egresos": round(tot_egresos, 2),
            "neto": round(tot_ingresos - tot_egresos, 2),
            "count": total_count,
        },
    }


@api_router.get("/cashbox/movements/{movement_id}")
async def get_cash_movement(movement_id: str, current_user: dict = Depends(get_current_user)):
    f = _cashbox_filtros(current_user["company_id"])
    f.agregar("id = $?", db_pg.as_uuid(movement_id))
    async with db_pg.tx(current_user) as conn:
        row = await conn.fetchrow(
            "select * from cash_movements where " + f.where, *f.values
        )
    if not row:
        raise HTTPException(status_code=404, detail="Movimiento de caja no encontrado")
    return db_pg.to_api(row)

@api_router.post("/cashbox/movements")
async def create_cash_movement(
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))
):
    """Registra un ingreso o egreso de caja con correlativo por empresa."""
    company_id = current_user["company_id"]

    amount = _to_float(request.get("amount"))
    if amount <= 0:
        raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0")

    mov_type = request.get("type") or CashMovementType.EGRESO.value
    if mov_type not in [t.value for t in CashMovementType]:
        raise HTTPException(status_code=400, detail="Tipo de movimiento inválido (ingreso|egreso)")

    payment_method = request.get("payment_method") or CashPaymentMethod.EFECTIVO.value
    if payment_method not in [p.value for p in CashPaymentMethod]:
        raise HTTPException(status_code=400, detail="Método de pago inválido")

    concept = (request.get("concept") or "").strip()
    if not concept:
        raise HTTPException(status_code=400, detail="El concepto es obligatorio")

    # El correlativo y el alta van en la MISMA transaccion: antes, entre
    # calcular el numero y guardar el movimiento, otra alta podia colarse.
    async with db_pg.tx(current_user) as conn:
        movement = CashMovement(
            company_id=company_id,
            movement_number=await _next_movement_number(conn, company_id),
            date=request.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            type=mov_type,
            concept=concept,
            category=(request.get("category") or "otros"),
            amount=round(amount, 2),
            payment_method=payment_method,
            reference=request.get("reference"),
            trip_id=request.get("trip_id"),
            vehicle_id=request.get("vehicle_id"),
            client_ruc=request.get("client_ruc"),
            supplier=request.get("supplier"),
            receipt_url=request.get("receipt_url"),
            notes=request.get("notes"),
            created_by=current_user["id"],
        )
        sql, values = db_pg.build_insert(
            "cash_movements", CASH_MOVEMENT_COLS, _modelo_a_fila(movement.model_dump())
        )
        await conn.execute(sql, *values)
        row = await conn.fetchrow(
            "select * from cash_movements where id = $1", db_pg.as_uuid(movement.id)
        )
    return db_pg.to_api(row)

@api_router.put("/cashbox/movements/{movement_id}")
async def update_cash_movement(
    movement_id: str,
    request: dict = Body(...),
    current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))
):
    """Actualiza un movimiento de caja (el correlativo no cambia)."""
    company_id = current_user["company_id"]
    async with db_pg.tx(current_user) as conn:
        f = _cashbox_filtros(company_id)
        f.agregar("id = $?", db_pg.as_uuid(movement_id))
        existe = await conn.fetchval(
            "select 1 from cash_movements where " + f.where, *f.values
        )
        if not existe:
            raise HTTPException(status_code=404, detail="Movimiento de caja no encontrado")

        update_data = {}
        for field in ["date", "concept", "category", "reference", "trip_id", "vehicle_id",
                      "client_ruc", "supplier", "receipt_url", "notes"]:
            if field in request:
                update_data[field] = request[field]

        if "amount" in request:
            amount = _to_float(request.get("amount"))
            if amount <= 0:
                raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0")
            update_data["amount"] = round(amount, 2)

        if "type" in request and request["type"]:
            if request["type"] not in [t.value for t in CashMovementType]:
                raise HTTPException(status_code=400, detail="Tipo de movimiento inválido (ingreso|egreso)")
            update_data["type"] = request["type"]

        if "payment_method" in request and request["payment_method"]:
            if request["payment_method"] not in [p.value for p in CashPaymentMethod]:
                raise HTTPException(status_code=400, detail="Método de pago inválido")
            update_data["payment_method"] = request["payment_method"]

        if "concept" in request and not (request.get("concept") or "").strip():
            raise HTTPException(status_code=400, detail="El concepto es obligatorio")

        update_data["updated_at"] = datetime.now(timezone.utc)
        update_data["id"] = movement_id
        update_data["company_id"] = company_id
        sql, values = db_pg.build_update(
            "cash_movements", CASH_MOVEMENT_COLS, update_data, ["id", "company_id"]
        )
        if sql:
            await conn.execute(sql, *values)

        actualizado = await conn.fetchrow(
            "select * from cash_movements where id = $1 and company_id = $2",
            db_pg.as_uuid(movement_id), db_pg.as_uuid(company_id),
        )
    return db_pg.to_api(actualizado)


@api_router.delete("/cashbox/movements/{movement_id}")
async def delete_cash_movement(
    movement_id: str,
    current_user: dict = Depends(require_roles("owner", "admin", "contabilidad"))
):
    """Elimina un movimiento de caja (soft-delete: no altera los correlativos)."""
    company_id = current_user["company_id"]
    ahora = datetime.now(timezone.utc)
    async with db_pg.tx(current_user) as conn:
        # El "not deleted" del WHERE hace que borrar dos veces devuelva 404 la
        # segunda vez, igual que el filtro deleted != true que usaba con Mongo.
        borrado = await conn.fetchval(
            "update cash_movements set deleted = true, deleted_at = $1, "
            "deleted_by = $2, updated_at = $1 "
            "where id = $3 and company_id = $4 and not deleted returning id",
            ahora,
            db_pg.as_uuid(current_user["id"]),
            db_pg.as_uuid(movement_id),
            db_pg.as_uuid(company_id),
        )
    if not borrado:
        raise HTTPException(status_code=404, detail="Movimiento de caja no encontrado")
    return {"message": "Movimiento de caja eliminado"}

# Include router (DESPUÉS de definir TODAS las rutas @api_router, incluidas SUNAT)
app.include_router(api_router)

# Módulo Liquidación de Flete (proveedores, tipos de carga, liquidaciones y sus líneas)
from liquidacion_flete import router as liquidacion_router
app.include_router(liquidacion_router)

# Bot de WhatsApp (webhook + bandeja de documentos pendientes)
from whatsapp_bot import router as whatsapp_router
app.include_router(whatsapp_router)

# Serve uploaded files
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

# Serve React frontend build
FRONTEND_BUILD = ROOT_DIR.parent / "frontend" / "build"
if FRONTEND_BUILD.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_BUILD / "static")), name="react-static")

    @app.get("/{full_path:path}")
    async def serve_react(full_path: str):
        """Serve React SPA - all non-API routes return index.html"""
        if full_path.startswith("api/") or full_path.startswith("uploads/"):
            raise HTTPException(status_code=404, detail="Not found")
        file_path = FRONTEND_BUILD / full_path
        if full_path and file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        return FileResponse(str(FRONTEND_BUILD / "index.html"))

# CORS
_env = os.environ.get("ENV", "development").lower()
_cors_raw = os.environ.get("CORS_ORIGINS", "").strip()
# Lista explícita solo si NO es "*" (comodín inseguro junto a credenciales)
_cors_explicit = [o.strip() for o in _cors_raw.split(",") if o.strip() and o.strip() != "*"]
if _cors_explicit:
    _cors_origins = _cors_explicit
elif _env == "production":
    # En producción CORS_ORIGINS es obligatorio y no puede ser "*"
    raise RuntimeError("CORS_ORIGINS no configurado (o '*') en producción")
else:
    # Desarrollo: permitir localhost por defecto
    _cors_origins = [
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:5173",
        "http://127.0.0.1:5173",
    ]
# Cada inquilino es un origen distinto (<slug>.sisac.pe), y son tantos como
# clientes haya: enumerarlos en CORS_ORIGINS significaria reiniciar el backend
# en cada alta. La expresion cubre un solo nivel bajo el dominio base, que es
# exactamente lo que cubre el certificado comodin.
#
# En la practica casi no se usa -el frontend se sirve del mismo origen que la
# API-, pero el dia que algo pida desde fuera, esto es lo que decide.
_tenant_regex = (
    r"^https://[a-z0-9][a-z0-9-]*[a-z0-9]\." + re.escape(tenant_host.DOMINIO_BASE) + r"$"
    if tenant_host.DOMINIO_BASE
    else None
)

# Nunca allow_origins=["*"] junto con allow_credentials=True
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=_cors_origins,
    allow_origin_regex=_tenant_regex,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Cabeceras de seguridad en todas las respuestas (defensa en profundidad).
# El token sigue siendo bearer en localStorage (compatible con el deploy
# cross-origin frontend/backend); estas cabeceras reducen la superficie XSS/clickjacking.
@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers.setdefault("Cache-Control", "no-store")
    return response

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def create_indexes():
    """Crea índices idempotentes en background para consultas multi-tenant frecuentes."""
    try:
        # Los indices de users ya no se crean aca: la tabla corto a Postgres y
        # sus indices (incluido el unico parcial sobre whatsapp_number) viven
        # en db/schema.sql, que es donde se versionan.
        # Los indices de vehicles viven ahora en db/schema.sql y en la
        # migracion 006, que es donde se versionan.
        # El indice de trips ya no se crea aca: la tabla corto a Postgres y sus
        # indices viven en db/schema.sql y en la migracion 007.
        # El indice de tires ya no se crea aca: la tabla corto a Postgres y sus
        # indices viven en db/schema.sql y en la migracion 013.
        # El indice de documents ya no se crea aca: la tabla corto a Postgres y
        # sus indices viven en db/schema.sql y en la migracion 009.
        logger.info("Índices de MongoDB verificados/creados")
    except Exception as e:
        logger.error(f"Error creando índices MongoDB: {e}")

@app.on_event("startup")
async def ensure_default_document_types():
    """Asegura los tipos de documento estándar en empresas existentes (idempotente por nombre).
    Repara instalaciones donde el seed no volvió a correr (p. ej. faltaban Tarjeta de
    Circulación y Bonificación)."""
    defaults = [
        {"name": "SOAT", "applies_to": "vehiculo", "is_critical": True, "block_rule": "bloquea_inicio"},
        {"name": "Revisión Técnica (CITV)", "applies_to": "vehiculo", "is_critical": True, "block_rule": "bloquea_inicio"},
        {"name": "Tarjeta de Propiedad", "applies_to": "vehiculo", "is_critical": True, "block_rule": "bloquea_asignacion"},
        {"name": "Tarjeta de Circulación", "applies_to": "vehiculo", "is_critical": False, "block_rule": "solo_alerta"},
        {"name": "Bonificación", "applies_to": "vehiculo", "is_critical": False, "block_rule": "solo_alerta"},
        {"name": "Póliza de Seguro", "applies_to": "vehiculo", "is_critical": False, "block_rule": "solo_alerta"},
        {"name": "Licencia de Conducir", "applies_to": "chofer", "is_critical": True, "block_rule": "bloquea_asignacion"},
        {"name": "DNI", "applies_to": "chofer", "is_critical": True, "block_rule": "bloquea_asignacion"},
        {"name": "Certificado Médico", "applies_to": "chofer", "is_critical": False, "block_rule": "solo_alerta"},
    ]
    try:
        async with db_pg.tx_global("arranque: sembrar tipos de documento en cada empresa") as conn:
            companies = db_pg.rows_to_api(await conn.fetch("select id from companies"))
        for comp in companies:
            cid = comp.get("id")
            if not cid:
                continue
            for dt in defaults:
                async with db_pg.tx({"company_id": cid}) as conn:
                    exists = await conn.fetchval(
                        "select id from document_types "
                        "where company_id = $1 and name = $2",
                        db_pg.as_uuid(cid), dt["name"],
                    )
                if exists:
                    continue
                doc_type = DocumentType(
                    company_id=cid, name=dt["name"], applies_to=dt["applies_to"],
                    is_critical=dt["is_critical"], block_rule=BlockRule(dt["block_rule"]),
                )
                sql, values = db_pg.build_insert(
                    "document_types", DOCUMENT_TYPE_COLS,
                    _modelo_a_fila(doc_type.model_dump()),
                )
                async with db_pg.tx({"company_id": cid}) as conn:
                    await conn.execute(sql, *values)
        logger.info("Tipos de documento por defecto verificados")
    except Exception as e:
        logger.error(f"Error asegurando tipos de documento: {e}")

async def _scheduler_loop():
    """Barrido periódico (documentos + mantenimiento + llantas + viáticos) cada N horas."""
    try:
        interval_hours = int(os.environ.get("SCHEDULER_INTERVAL_HOURS", "6"))
    except (TypeError, ValueError):
        interval_hours = 6
    interval = max(interval_hours, 1) * 3600
    # Pequeño retraso inicial para no competir con el arranque
    await asyncio.sleep(120)
    while True:
        try:
            await run_maintenance_sweep()
            logger.info("Barrido de alertas/mantenimiento ejecutado")
        except Exception as e:
            logger.error(f"Error en barrido programado: {e}")
        await asyncio.sleep(interval)

@app.on_event("startup")
async def start_scheduler():
    """Lanza el scheduler en background (deshabilitable con DISABLE_SCHEDULER=1)."""
    if os.environ.get("DISABLE_SCHEDULER", "").lower() in ("1", "true", "yes"):
        logger.info("Scheduler deshabilitado por DISABLE_SCHEDULER")
        return
    try:
        asyncio.create_task(_scheduler_loop())
        logger.info("Scheduler de mantenimiento/alertas iniciado")
    except Exception as e:
        logger.error(f"No se pudo iniciar el scheduler: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
